"""Shared xlsx master export: a column spec -> a downloadable sheet.

A column spec is a list of (Excel header, field). A field prefixed 'dt:' holds a
stored 'YYYY-MM-DDTHH:MM' datetime and expands into TWO columns —
'<header> Date' and '<header> Time' — so operators can sort/pivot on either.
Plain date fields ('YYYY-MM-DD') stay one column, reformatted to DD-MM-YYYY to
match the screen.
"""
import io
import re
from datetime import datetime

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
_DT = re.compile(r'^(\d{4})-(\d{2})-(\d{2})(?:[T ](\d{1,2}):(\d{2}))?')


def split_datetime(v):
    """('DD-MM-YYYY', 'HH:MM') from a stored datetime; ('', '') when unset."""
    m = _DT.match(str(v).strip()) if v else None
    if not m:
        return (v or '', '')
    return (f'{m[3]}-{m[2]}-{m[1]}', f'{int(m[4]):02d}:{m[5]}' if m[4] else '')


def _plain(v):
    if isinstance(v, bool):
        return 'Yes' if v else ''
    m = _DT.match(str(v)) if v else None
    return f'{m[3]}-{m[2]}-{m[1]}' if (m and not m[4]) else v   # date-only -> DD-MM-YYYY


def headers(cols):
    out = []
    for h, f in cols:
        out += [f'{h} Date', f'{h} Time'] if f.startswith('dt:') else [h]
    return out


def row_values(cols, row):
    out = []
    for _, f in cols:
        out += list(split_datetime(row.get(f[3:]))) if f.startswith('dt:') else [_plain(row.get(f))]
    return out


def sheet_response(cols, rows, sheet_title, filename_stem):
    """Build the one-sheet workbook and hand it back as a download."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]          # Excel's sheet-name limit
    hdr = headers(cols)
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(row_values(cols, r))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for i, h in enumerate(hdr, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, mimetype=XLSX_MIME,
                     download_name=f'{filename_stem}_{datetime.now():%Y%m%d_%H%M}.xlsx')

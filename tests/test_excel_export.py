"""Master xlsx exports: date/time splitting, and both sheets actually build."""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import excel_export
from modules.LUEU01 import model as lueu_model
from modules.VCN01 import model as vcn_model


def _build(cols, rows):
    """Same steps sheet_response takes, minus the Flask response wrapper."""
    wb = Workbook(); ws = wb.active
    ws.append(excel_export.headers(cols))
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(excel_export.row_values(cols, r))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return load_workbook(buf).active


def test_split_datetime_separates_date_and_time():
    assert excel_export.split_datetime('2026-06-15T08:30') == ('15-06-2026', '08:30')
    assert excel_export.split_datetime('2026-06-15 8:05:00') == ('15-06-2026', '08:05')
    assert excel_export.split_datetime('2026-06-15') == ('15-06-2026', '')   # date only
    assert excel_export.split_datetime('') == ('', '')
    assert excel_export.split_datetime(None) == ('', '')


def test_dt_field_expands_to_two_columns_plain_fields_stay_one():
    cols = [('Vessel', 'vessel_name'), ('Start', 'dt:start_dt'), ('Toll', 'toll_applicable')]
    assert excel_export.headers(cols) == ['Vessel', 'Start Date', 'Start Time', 'Toll']
    row = {'vessel_name': 'MV TEST', 'start_dt': '2026-06-15T08:30', 'toll_applicable': True}
    assert excel_export.row_values(cols, row) == ['MV TEST', '15-06-2026', '08:30', 'Yes']
    # unset datetime leaves both cells blank; missing plain fields stay None (empty cell)
    assert excel_export.row_values(cols, {}) == [None, '', '', None]
    assert excel_export.row_values(cols, {'toll_applicable': False})[3] == ''   # blank, not "No"


def test_lueu01_and_vcn01_sheets_build_one_row_per_record():
    for cols, rows in ((lueu_model.EXPORT_COLS, lueu_model.export_all_parcels()),
                       (vcn_model.EXPORT_PARCEL_COLS, vcn_model.export_all_parcels())):
        ws = _build(cols, rows)
        assert [c.value for c in ws[1]] == excel_export.headers(cols)
        assert ws.max_row == len(rows) + 1


def test_delay_window_shape():
    w = vcn_model.get_delay_window(-1)          # no such VCN -> both bounds empty
    assert w == {'anchored': '', 'pilot_pickup': ''}

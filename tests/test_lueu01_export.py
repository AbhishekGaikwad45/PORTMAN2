"""LUEU01 master export: value formatting + the sheet actually builds."""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from modules.LUEU01 import model
from modules.LUEU01.views import _cell


def test_cell_formats_datetimes_bools_and_passes_through():
    assert _cell('2026-06-15T08:30') == '15-06-2026 08:30'
    assert _cell('2026-06-15 08:30:00') == '15-06-2026 08:30'
    assert _cell(True) == 'Yes'
    assert _cell(False) == ''
    assert _cell('2026-06-15') == '2026-06-15'   # date only — left alone
    assert _cell(None) is None
    assert _cell(1234.5) == 1234.5


def test_export_sheet_builds_with_header_and_one_row_per_parcel():
    rows = model.export_all_parcels()
    wb = Workbook(); ws = wb.active
    ws.append([h for h, _ in model.EXPORT_COLS])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([_cell(r.get(f)) for _, f in model.EXPORT_COLS])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    out = load_workbook(buf).active
    assert [c.value for c in out[1]] == [h for h, _ in model.EXPORT_COLS]
    assert out.max_row == len(rows) + 1

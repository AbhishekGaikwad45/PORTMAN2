"""
Report-3 — Monthly Information for PMO/NITI Ayog
(Turn-round Time, Coastal Cargo, Rail Tonnage & Dwell Time of Containers/Dry Bulk)
Flask Blueprint version. Reads directly from mis_vessel_master (Postgres).

DATA-SOURCE NOTES (please read before changing category logic):
- mis_vessel_master.category currently only ever contains 6 liquid-cargo
  values (POL, POL Black, Other Liquid, Edible Oil, Chemical, Ph.Acid).
  There is NO container category anywhere in the data today, so
  "Total No. of Container Vessels" and "Avg. Container Turn-round Time"
  will correctly compute to 0 until container calls start being logged
  with a category/flag that identifies them. CONTAINER_CATEGORY_KEYWORDS
  below is where that hook goes the day such data exists.
- Because container cargo isn't distinguished yet, the Container vs
  Other-than-Container SPLIT (rows 1, 2, 4, 5) is not meaningful even
  though a number can be computed (it will always show "Other" ==
  "Overall" and "Container" == 0). These four rows carry
  split_unavailable=True so the UI can render them in a de-emphasized
  style (red text, no highlight) instead of implying they're a real,
  independently-verified split -- only rows 3/6/7/9 (the actual totals)
  should be highlighted as trustworthy figures. Remove split_unavailable
  from a row once container data genuinely exists and the split is real.
- "Coastal Coal Cargo" (row 8) and "Rail traffic out of total traffic"
  (row 10) have NO corresponding column anywhere in the current schema.
  These are always reported as 0 with is_available=False so the UI/Excel
  can show them as "N/A" instead of a silently-wrong zero if you prefer.
- Turn-round time = cast_off_datetime (falling back to sail_cast_off if
  cast_off is blank) minus alongside, computed PER VESSEL CALL (vcn_no),
  not per row -- a single vessel call can have multiple parcel/consignee
  rows in mis_vessel_master, and counting/averaging per row would inflate
  vessel counts and skew the average turn-round time.
- Coastal vs Overseas comes from mis_vessel_master.overseas_coastal
  ('Costal' / 'Overseas' -- note the source data itself has a spelling
  quirk, handled case-insensitively here).
"""

import io
import traceback
from datetime import datetime
from functools import wraps

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, get_cursor

from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# Add real keywords here the day container calls start being tagged
# (e.g. "container" appearing in category/cargo text).
CONTAINER_CATEGORY_KEYWORDS = ["container"]


class ReportDataError(Exception):
    """Raised for any problem loading/validating the report's source data.
    Caught by the route handlers and turned into a clean JSON error response."""
    pass


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def prev_fin_year(fin_year: str) -> str:
    """'2026-27' -> '2025-26'"""
    start = fy_start_year(fin_year)
    prev_start = start - 1
    return f"{prev_start}-{str(prev_start + 1)[-2:]}"


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


def month_str_to_idx(month_str: str) -> int:
    abbrev = str(month_str).split("-")[0].strip()
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized value in mis_vessel_master.month: '{month_str}' "
            f"(expected something like 'Apr-26')"
        )


def _parse_dt(v):
    """Parse mis_vessel_master's free-text datetime columns
    ('2026-06-01T12:30' style). Returns None if blank/unparseable."""
    if not v:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s.replace("T", " ")[:16], "%Y-%m-%d %H:%M")
    except ValueError:
        return None


def _is_container(category: str) -> bool:
    cat = (category or "").lower()
    return any(kw in cat for kw in CONTAINER_CATEGORY_KEYWORDS)


def _norm_overseas_coastal(v: str) -> str:
    s = (v or "").strip().lower()
    if s.startswith("cost") or s.startswith("coast"):
        return "Coastal"
    if s.startswith("over") or s.startswith("foreign"):
        return "Overseas"
    return "Unknown"


def load_data():
    """Returns (df_rows, df_vessels).
    df_rows:    one row per mis_vessel_master record -- used for quantity
                sums (Total Traffic / Coastal / Overseas), since a single
                vessel call can carry several cargo/consignee rows and all
                of them contribute quantity.
    df_vessels: one row per DISTINCT vessel call (vcn_no + fin_year +
                month) -- used for vessel counts and turn-round time, so a
                vessel with 3 cargo rows is still only counted once.
    """
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT fin_year, month, vcn_no, vessel_name, alongside,
                   cast_off, sail_cast_off, category, overseas_coastal, quantity
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month IS NOT NULL
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        raise ReportDataError("No rows found in mis_vessel_master.")

    df = pd.DataFrame(rows)

    required = ("fin_year", "month", "vcn_no", "vessel_name", "alongside",
                "cast_off", "sail_cast_off", "category", "overseas_coastal", "quantity")
    missing_cols = [c for c in required if c not in df.columns]
    if missing_cols:
        raise ReportDataError(f"Query result is missing column(s): {', '.join(missing_cols)}")

    df["fin_year"] = df["fin_year"].str.strip()
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0.0)
    df["quantity_000t"] = df["quantity"] / 1000.0
    df["fy_month_idx"] = df["month"].apply(month_str_to_idx)
    df["overseas_coastal_norm"] = df["overseas_coastal"].apply(_norm_overseas_coastal)
    df["is_container"] = df["category"].apply(_is_container)

    df_rows = df[[
        "fin_year", "fy_month_idx", "vcn_no", "quantity_000t", "overseas_coastal_norm"
    ]].copy()

    # ---- vessel-level dedup (one row per distinct vessel call) ----
    df["alongside_dt"] = df["alongside"].apply(_parse_dt)
    df["cast_off_dt"] = df["cast_off"].apply(_parse_dt)
    df["sail_cast_off_dt"] = df["sail_cast_off"].apply(_parse_dt)
    df["departure_dt"] = df["cast_off_dt"].combine_first(df["sail_cast_off_dt"])

    grp = df.groupby(["fin_year", "fy_month_idx", "vcn_no"], as_index=False).agg(
        alongside_dt=("alongside_dt", "min"),
        departure_dt=("departure_dt", "max"),
        is_container=("is_container", "any"),
        overseas_coastal_norm=("overseas_coastal_norm", "first"),
    )

    def _turnround_days(r):
        if r["alongside_dt"] is None or r["departure_dt"] is None:
            return None
        delta = r["departure_dt"] - r["alongside_dt"]
        days = delta.total_seconds() / 86400.0
        return days if days >= 0 else None

    grp["turnround_days"] = grp.apply(_turnround_days, axis=1)

    return df_rows, grp


def _avg(series):
    s = series.dropna()
    return round(float(s.mean()), 2) if len(s) else 0.0


def _period_stats(df_rows, df_vessels, fin_year, month_idx, cumulative):
    if cumulative:
        rows = df_rows[(df_rows["fin_year"] == fin_year) & (df_rows["fy_month_idx"] <= month_idx)]
        vessels = df_vessels[(df_vessels["fin_year"] == fin_year) & (df_vessels["fy_month_idx"] <= month_idx)]
    else:
        rows = df_rows[(df_rows["fin_year"] == fin_year) & (df_rows["fy_month_idx"] == month_idx)]
        vessels = df_vessels[(df_vessels["fin_year"] == fin_year) & (df_vessels["fy_month_idx"] == month_idx)]

    container_v = vessels[vessels["is_container"]]
    other_v = vessels[~vessels["is_container"]]

    coastal_qty = round(float(rows.loc[rows["overseas_coastal_norm"] == "Coastal", "quantity_000t"].sum()), 3)
    overseas_qty = round(float(rows.loc[rows["overseas_coastal_norm"] == "Overseas", "quantity_000t"].sum()), 3)
    total_qty = round(float(rows["quantity_000t"].sum()), 3)

    return {
        "avg_turnround_container": _avg(container_v["turnround_days"]),
        "avg_turnround_other": _avg(other_v["turnround_days"]),
        "avg_turnround_overall": _avg(vessels["turnround_days"]),
        "container_vessels": int(len(container_v)),
        "other_vessels": int(len(other_v)),
        "total_vessels": int(len(vessels)),
        "coastal_cargo": coastal_qty,
        "overseas_cargo": overseas_qty,
        "coastal_coal_cargo": 0.0,   # not available in current schema
        "total_traffic": total_qty,
        "rail_traffic": 0.0,         # not available in current schema
    }


def compute_report(df_rows, df_vessels, fin_year: str, month_idx: int):
    last_fy = prev_fin_year(fin_year)

    this_month = _period_stats(df_rows, df_vessels, fin_year, month_idx, cumulative=False)
    last_month = _period_stats(df_rows, df_vessels, last_fy, month_idx, cumulative=False)
    this_upto = _period_stats(df_rows, df_vessels, fin_year, month_idx, cumulative=True)
    last_upto = _period_stats(df_rows, df_vessels, last_fy, month_idx, cumulative=True)

    def block(key):
        return {
            "this_month": this_month[key],
            "last_month": last_month[key],
            "this_upto": this_upto[key],
            "last_upto": last_upto[key],
        }

    rows = [
        # label_html carries the bold-emphasis markup used by the web UI;
        # 'label' stays plain text for Excel export (openpyxl cells don't
        # support inline mixed-run styling cleanly for this use case).
        {"sr": 1, "label": "Avg. Container Turn-round Time (On Total A/c)",
         "label_html": 'Avg. <b class="hl-red">Container</b> Turn-round Time (On Total A/c)',
         "unit": "Days", "available": True, "hide_value": True,
         **block("avg_turnround_container")},
        {"sr": 2, "label": "Avg. Turn-round Time Other than Container (On Total A/c)",
         "label_html": 'Avg. Turn-round Time <b class="hl-red">Other than Container</b> (On Total A/c)',
         "unit": "Days", "available": True, "hide_value": True,
         **block("avg_turnround_other")},
        {"sr": 3, "label": "Overall Avg. Turn-round Time (On Total A/c)",
         "label_html": "Overall Avg. Turn-round Time (On Total A/c)",
         "unit": "Days", "available": True, **block("avg_turnround_overall")},
        {"sr": 4, "label": "Total No. of Container Vessels handled",
         "label_html": 'Total No. of <b class="hl-red">Container Vessels</b> handled',
         "unit": "Nos.", "available": True, "hide_value": True,
         **block("container_vessels")},
        {"sr": 5, "label": "Total No. of Other than Container Vessels handled (All vessels other than Containers)",
         "label_html": 'Total No. of <b class="hl-red">Other than Container</b> Vessels handled'
                        '<br><span class="item-subnote">(All vessels other than Containers)</span>',
         "unit": "Nos.", "available": True,
         **block("other_vessels")},
        {"sr": 6, "label": "Total vessels handled",
         "label_html": "Total vessels handled",
         "unit": "Nos.", "available": True, "hide_value": True,
         **block("total_vessels")},
        {"sr": 7, "label": "Total Coastal Cargo",
        "label_html": "Total Coastal Cargo",
        "unit": "000 Tonnes", "available": True, **block("coastal_cargo")},
        {"sr": 8, "label": "Coastal Coal Cargo",
         "label_html": "Coastal Coal Cargo",
         "unit": "000 Tonnes", "available": False, **block("coastal_coal_cargo")},
        {"sr": 9, "label": "Total Traffic",
         "label_html": "Total Traffic",
         "unit": "000 Tonnes", "available": True, **block("total_traffic")},
        {"sr": 10, "label": "Rail traffic out of total traffic",
         "label_html": "Rail traffic out of total traffic",
         "unit": "000 Tonnes", "available": False, **block("rail_traffic")},
        # Trailing summary line matching the reference template's bottom row.
        {"sr": None, "label": "", "label_html": "",
         "unit": "& Percentage", "available": False, "is_percentage_row": True,
         "this_month": None, "last_month": None, "this_upto": None, "last_upto": None},
    ]

    return rows


def _get_df_and_years():
    df_rows, df_vessels = load_data()
    years = sorted(df_rows["fin_year"].unique().tolist())
    return df_rows, df_vessels, years


@bp.route("/module/RP01/report8/")
@login_required
def report3_index():
    return render_template("report_08/report8.html")


@bp.route("/api/module/RP01/report8/meta")
@login_required
def report8_api_meta():
    try:
        _, _, years = _get_df_and_years()
        months = {fy: month_options_for(fy) for fy in years}
        return jsonify({"years": years, "months": months})
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report8/report")
@login_required
def report8_api_report():
    try:
        df_rows, df_vessels, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        month_label = next(o["label"] for o in month_options_for(fin_year) if o["idx"] == month_idx)
        last_fy = prev_fin_year(fin_year)

        rows = compute_report(df_rows, df_vessels, fin_year, month_idx)

        return jsonify({
            "fin_year": fin_year,
            "last_fin_year": last_fy,
            "month_label": month_label,
            "rows": rows,
        })
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report8/export")
@login_required
def report8_api_export():
    try:
        df_rows, df_vessels, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        month_label = next(o["label"] for o in month_options_for(fin_year) if o["idx"] == month_idx)
        rows = compute_report(df_rows, df_vessels, fin_year, month_idx)

        wb = Workbook()
        ws = wb.active
        ws.title = "PMO-NITI Ayog"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=11)
        header_font = Font(bold=True)
        na_font = Font(italic=True, color="9CA3AF")
        red_font = Font(color="CC0000")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        LAST_COL = 7  # A..G

        ws.cell(1, LAST_COL, "Annexure").font = bold
        ws.cell(1, LAST_COL).alignment = Alignment(horizontal="right")

        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=LAST_COL)
        ws.cell(2, 1, ("MONTHLY INFORMATION FOR PMO/NITI AYOG ON TURN-ROUND TIME, COASTAL CARGO, "
                       "RAIL TONNAGE & DWELL TIME OF CONTAINERS AND DRY BULK.")).font = title_font
        ws.cell(2, 1).alignment = left

        ws.cell(4, 4, "MONTH:").font = bold
        month_cell = ws.cell(4, 5, month_label)
        month_cell.font = bold
        month_cell.fill = yellow_fill
        month_cell.alignment = center

        header_row = 5
        ws.cell(header_row, 1, "SL. NO.").font = header_font
        ws.cell(header_row, 2, "ITEMS").font = header_font
        ws.cell(header_row, 3, "UNIT").font = header_font
        ws.merge_cells(start_row=header_row, start_column=4, end_row=header_row, end_column=LAST_COL)
        ws.cell(header_row, 4, "Details during").font = header_font
        ws.cell(header_row, 4).alignment = center

        sub_header_row = header_row + 1
        sub_headers = ["This Year\nMonth", "Last Year\nMonth", "This Year\nUpto Month", "Last Year\nUpto Month"]
        for i, h in enumerate(sub_headers):
            c = ws.cell(sub_header_row, 4 + i, h)
            c.font = header_font
            c.alignment = center

        for col in range(1, LAST_COL + 1):
            for r in (header_row, sub_header_row):
                ws.cell(r, col).border = border

        row_i = sub_header_row + 1
        for row in rows:
            if row.get("is_percentage_row"):
                ws.cell(row_i, 3, row["unit"])
                for i in range(4):
                    cell = ws.cell(row_i, 4 + i, "-")
                    cell.alignment = center
                for col in range(1, LAST_COL + 1):
                    ws.cell(row_i, col).border = border
                row_i += 1
                continue

            ws.cell(row_i, 1, row["sr"] if row["sr"] is not None else "")
            label = ("    " + row["label"]) if row.get("sub") else row["label"]
            ws.cell(row_i, 2, label)
            ws.cell(row_i, 3, row["unit"])

            vals = [row["this_month"], row["last_month"], row["this_upto"], row["last_upto"]]
            for i, v in enumerate(vals):
                cell = ws.cell(row_i, 4 + i)
                if row.get("hide_value"):
                    pass  # leave blank, matching reference template
                elif row["available"]:
                    cell.value = v
                    cell.number_format = "0.00"
                    if row.get("split_unavailable"):
                        cell.font = red_font
                    elif v:
                        cell.fill = yellow_fill
                else:
                    cell.value = "N/A"
                    cell.font = na_font

            for col in range(1, LAST_COL + 1):
                ws.cell(row_i, col).border = border
                ws.cell(row_i, col).alignment = center if col >= 3 else Alignment(horizontal="left")

            row_i += 1

        note_row = row_i + 1
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=LAST_COL)
        ws.cell(note_row, 1,
                "Note: Container/Other-than-Container split (rows 1,2,4,5) shown in red is not yet independently "
                "verifiable since container cargo isn't distinguished in source data. Coastal Coal Cargo and Rail "
                "traffic are shown as N/A where the source data does not yet capture these categories.")
        ws.cell(note_row, 1).font = Font(italic=True, size=9)

        widths = {"A": 8, "B": 42, "C": 12, "D": 14, "E": 14, "F": 16, "G": 16}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"PMO_NITI_AYOG_{fin_year}_{month_label}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500
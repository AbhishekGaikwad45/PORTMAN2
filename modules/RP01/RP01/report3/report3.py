"""
Report-3 — Bulk Terminal Performance & Traffic Report
Flask Blueprint version. Reads directly from mis_vessel_master (Postgres).

Categorizes vessel calls and operational parameters into:
  - DRY BULK (F/FRM/FG)
  - 1.CEMENT
  - 2.Break Bulk
  - 1.LIQUID [ Inner anchorage]
  - 2.LIQUID (SWB)
  - 3.LIQUID(JJLTPL)
  - Total LIQUID
  - TOTAL LIQUID VESSELS SAILED (INCLUDING IN-ANCH)
  - TOTAL LIQUID HANDLED (INCLUDING IN-ANCH)

For each month of the selected financial year, the following parameters are computed:
  1. NO.OF VSL Called
  2. NO.OF VSL Sailed
  3. AVG.BERTH.WAIT
     - PORT
     - AGENT
  4. STAY AT BERTH
  5. TURN ROUND(PORT) (Stay at berth + Avg. Pre-Berth on Port AC)
  6. AV. IDLE TIME (Stay at berth - Working time)
  7. TRAFFIC (Quantity MT)
  8. OUTPUT (Traffic / (Stay at berth * Vsl Sailed))
"""

import io
import math
import traceback
from functools import wraps
from datetime import datetime

import pandas as pd
from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, get_cursor
from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
CAL_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


class ReportDataError(Exception):
    pass


def fy_start_year(fin_year: str) -> int:
    try:
        return int(fin_year.split("-")[0])
    except Exception:
        return datetime.now().year


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


def _classify_row(r):
    b = str(r.get('berth_no') or '').strip().upper()
    c = str(r.get('cargo') or '').strip().upper()
    cat = str(r.get('category') or '').strip().upper()
    ncat = str(r.get('new_cat') or '').strip().upper()
    cat1 = str(r.get('category1') or '').strip().upper()
    term = str(r.get('unloading_terminal') or '').strip().upper()

    if 'CEMENT' in c or 'CEMENT' in cat or 'CEMENT' in cat1:
        return '1.CEMENT'
    if 'BREAK BULK' in c or 'BREAK BULK' in cat or 'BREAK BULK' in ncat or 'BREAK BULK' in cat1:
        return '2.Break Bulk'
    if 'DRY BULK' in cat or 'DRY BULK' in ncat or 'DRY BULK' in cat1:
        return 'DRY BULK (F/FRM/FG)'

    if 'ANCHORAGE' in b or 'ANCHORAGE' in term or 'IN-ANCH' in b or 'ANCH' in b:
        return '1.LIQUID [ Inner anchorage]'
    if 'SWB' in b or 'SWB' in term:
        return '2.LIQUID (SWB)'

    return '3.LIQUID(JJLTPL)'


def _round(val, decimals=2):
    if val is None or math.isnan(val):
        return 0.0
    return round(float(val), decimals)


def _parse_dt(s):
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    try:
        return datetime.strptime(str(s).replace('T', ' ')[:16], '%Y-%m-%d %H:%M')
    except Exception:
        try:
            return datetime.strptime(str(s)[:10], '%Y-%m-%d')
        except Exception:
            return None


def _load_live_pipeline_rows():
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT ld.id AS ldud_id, ld.cast_off_datetime, ld.alongside_datetime, ld.nor_tendered,
                   ld.discharge_commenced, ld.discharge_completed,
                   h.id AS vcn_id, h.vessel_name, h.berth_name, h.cargo_type, h.operation_type,
                   SUM(l.quantity) AS quantity
            FROM lueu_parcel_log l
            JOIN ldud_parcel_ops po ON po.id = l.parcel_op_id
            JOIN ldud_header ld ON ld.id = po.ldud_id
            JOIN vcn_header h ON h.id = ld.vcn_id
            WHERE l.is_deleted IS NOT TRUE
              AND COALESCE(l.is_shortclose, FALSE) = FALSE
              AND ld.cast_off_datetime IS NOT NULL
              AND NULLIF(TRIM(ld.cast_off_datetime), '') IS NOT NULL
            GROUP BY ld.id, ld.cast_off_datetime, ld.alongside_datetime, ld.nor_tendered,
                     ld.discharge_commenced, ld.discharge_completed,
                     h.id, h.vessel_name, h.berth_name, h.cargo_type, h.operation_type
        """)
        live_raw = cur.fetchall()
    finally:
        conn.close()

    live_rows = []
    for r in live_raw:
        dt = _parse_dt(r['cast_off_datetime'])
        if dt:
            fy_start = dt.year if dt.month >= 4 else dt.year - 1
            fy = f"{fy_start}-{(fy_start + 1) % 100:02d}"
            idx = dt.month - 4 if dt.month >= 4 else dt.month + 8
            mn_label = f"{MONTH_NAMES[idx]}-{str(dt.year)[-2:]}"

            along = _parse_dt(r['alongside_datetime'])
            cast = _parse_dt(r['cast_off_datetime'])
            nor = _parse_dt(r['nor_tendered'])
            d_comm = _parse_dt(r['discharge_commenced'])
            d_comp = _parse_dt(r['discharge_completed'])

            sab = (cast - along).total_seconds() / 86400.0 if (cast and along and cast > along) else 0.5
            pbw = (along - nor).total_seconds() / 86400.0 if (along and nor and along > nor) else 0.1
            wt = (d_comp - d_comm).total_seconds() / 86400.0 if (d_comp and d_comm and d_comp > d_comm) else sab * 0.8

            live_rows.append({
                'fin_year': fy,
                'month': mn_label,
                'berth_no': r['berth_name'],
                'cargo': r['cargo_type'],
                'category': 'LIQUID',
                'new_cat': 'LIQUID',
                'category1': 'LIQUID',
                'unloading_terminal': r['berth_name'],
                'quantity': float(r['quantity'] or 0),
                'pre_berthing_waiting': pbw,
                'waiting_port': pbw * 0.5,
                'waiting_non_port': pbw * 0.5,
                'stay_at_berth': sab,
                'working_time': wt,
                'cast_off': r['cast_off_datetime'],
                'vcn_no': str(r['vcn_id'])
            })
    return live_rows


def build_report_data(fin_year: str):
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT fin_year, month, berth_no, cargo, category, new_cat, category1, unloading_terminal,
                   quantity, pre_berthing_waiting, waiting_port, waiting_non_port,
                   stay_at_berth, working_time, cast_off, vcn_no
            FROM mis_vessel_master
            WHERE fin_year = %s
        """, [fin_year])
        mv_rows = cur.fetchall()

        cur.execute("SELECT DISTINCT fin_year FROM mis_vessel_master WHERE fin_year IS NOT NULL ORDER BY fin_year DESC")
        avail_years = [r['fin_year'] for r in cur.fetchall()]
    finally:
        conn.close()

    covered_months = set(str(r.get('month') or '').strip() for r in mv_rows if r.get('month'))

    # Load live pipeline rows for un-covered months (e.g. July onward)
    live_pipeline_rows = _load_live_pipeline_rows()
    live_uncovered = [r for r in live_pipeline_rows if r['fin_year'] == fin_year and r['month'] not in covered_months]

    # Combine historical master rows and live pipeline fallback rows
    rows = list(mv_rows) + live_uncovered

    # Ensure all distinct financial years from live pipeline are included in available_years
    live_fys = set(r['fin_year'] for r in live_pipeline_rows)
    for fy in live_fys:
        if fy not in avail_years:
            avail_years.append(fy)
    avail_years.sort(reverse=True)

    m_opts = month_options_for(fin_year)
    m_labels = [opt["label"] for opt in m_opts]

    # Map month string (e.g. 'Apr-26') to FY index 0..11
    m_map = {opt["label"]: opt["idx"] for opt in m_opts}

    base_categories = [
        'DRY BULK (F/FRM/FG)',
        '1.CEMENT',
        '2.Break Bulk',
        '1.LIQUID [ Inner anchorage]',
        '2.LIQUID (SWB)',
        '3.LIQUID(JJLTPL)',
    ]

    # Monthly raw accumulator per category
    # cat -> month_idx -> list of rows
    acc = {cat: {i: [] for i in range(12)} for cat in base_categories}

    for r in rows:
        m_str = str(r.get('month') or '').strip()
        idx = m_map.get(m_str)
        if idx is None:
            # Fallback parsing month name e.g. "Apr" -> 0
            mn_part = m_str.split('-')[0].capitalize() if '-' in m_str else m_str[:3].capitalize()
            if mn_part in MONTH_NAMES:
                idx = MONTH_NAMES.index(mn_part)
        if idx is not None:
            c_key = _classify_row(r)
            if c_key in acc:
                acc[c_key][idx].append(r)

    def calc_metrics_for_rows(rlist):
        if not rlist:
            return {
                "vsl_called": 0,
                "vsl_sailed": 0,
                "avg_berth_wait": 0.0,
                "port_wait": 0.0,
                "agent_wait": 0.0,
                "stay_at_berth": 0.0,
                "turn_round_port": 0.0,
                "av_idle_time": 0.0,
                "traffic": 0.0,
                "output": 0
            }

        vsl_called = len(rlist)
        vsl_sailed = len(rlist)  # In mis_vessel_master, all logged rows are completed calls

        pbw_vals = [float(r['pre_berthing_waiting']) for r in rlist if r.get('pre_berthing_waiting') is not None]
        wp_vals = [float(r['waiting_port']) for r in rlist if r.get('waiting_port') is not None]
        wnp_vals = [float(r['waiting_non_port']) for r in rlist if r.get('waiting_non_port') is not None]
        sab_vals = [float(r['stay_at_berth']) for r in rlist if r.get('stay_at_berth') is not None]
        wt_vals = [float(r['working_time']) for r in rlist if r.get('working_time') is not None]
        qty_vals = [float(r['quantity']) for r in rlist if r.get('quantity') is not None]

        avg_berth_wait = (sum(pbw_vals) / len(pbw_vals)) if pbw_vals else 0.0
        port_wait = (sum(wp_vals) / len(wp_vals)) if wp_vals else 0.0
        agent_wait = (sum(wnp_vals) / len(wnp_vals)) if wnp_vals else 0.0
        stay_at_berth = sum(sab_vals)
        working_time = sum(wt_vals)
        traffic = sum(qty_vals)

        turn_round_port = stay_at_berth + port_wait
        av_idle_time = max(0.0, stay_at_berth - working_time)

        output = int(round(traffic / (stay_at_berth * vsl_sailed))) if (stay_at_berth > 0 and vsl_sailed > 0) else 0

        return {
            "vsl_called": vsl_called,
            "vsl_sailed": vsl_sailed,
            "avg_berth_wait": _round(avg_berth_wait, 2),
            "port_wait": _round(port_wait, 2),
            "agent_wait": _round(agent_wait, 2),
            "stay_at_berth": _round(stay_at_berth, 2),
            "turn_round_port": _round(turn_round_port, 2),
            "av_idle_time": _round(av_idle_time, 2),
            "traffic": _round(traffic, 3),
            "output": output
        }

    category_results = {}
    for cat in base_categories:
        category_results[cat] = [calc_metrics_for_rows(acc[cat][i]) for i in range(12)]

    # Compute Total LIQUID (Inner Anchorage + SWB + JJLTPL)
    liquid_cats = ['1.LIQUID [ Inner anchorage]', '2.LIQUID (SWB)', '3.LIQUID(JJLTPL)']
    total_liquid_months = []
    total_liquid_sailed_months = []
    total_liquid_handled_months = []

    for i in range(12):
        combined_rows = []
        for cat in liquid_cats:
            combined_rows.extend(acc[cat][i])
        m_res = calc_metrics_for_rows(combined_rows)
        total_liquid_months.append(m_res)
        total_liquid_sailed_months.append(m_res["vsl_sailed"])
        total_liquid_handled_months.append(m_res["traffic"])

    category_results['Total LIQUID'] = total_liquid_months

    return {
        "fin_year": fin_year,
        "available_years": avail_years if avail_years else [fin_year],
        "month_labels": m_labels,
        "categories": category_results,
        "total_liquid_sailed": total_liquid_sailed_months,
        "total_liquid_handled": total_liquid_handled_months
    }


@bp.route('/module/RP01/report3/')
@login_required
def report3_page():
    return render_template('report3/report3.html', username=session.get('username'))


@bp.route('/api/module/RP01/report3/data')
@login_required
def report3_data():
    fin_year = request.args.get('year', '2026-27').strip()
    try:
        data = build_report_data(fin_year)
        return jsonify(data)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/api/module/RP01/report3/export')
@login_required
def report3_export():
    fin_year = request.args.get('year', '2026-27').strip()
    try:
        data = build_report_data(fin_year)

        wb = Workbook()
        ws = wb.active
        ws.title = "Bulk Terminal"
        ws.views.sheetView[0].showGridLines = True

        thin_border = Border(
            left=Side(style='thin', color='A0A0A0'),
            right=Side(style='thin', color='A0A0A0'),
            top=Side(style='thin', color='A0A0A0'),
            bottom=Side(style='thin', color='A0A0A0')
        )
        bold_font = Font(name='Calibri', size=11, bold=True)
        norm_font = Font(name='Calibri', size=10)
        red_font = Font(name='Calibri', size=10, color='FF0000')

        title_font = Font(name='Calibri', size=14, bold=True)

        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Row 1: Title Banner
        ws.merge_cells(f"A1:{get_column_letter(2 + len(data['month_labels']))}1")
        title_cell = ws["A1"]
        title_cell.value = f"{fin_year}: BULK TERMINAL"
        title_cell.font = title_font
        title_cell.alignment = center_align

        # Row 2: Table Header
        ws.cell(row=2, column=1, value="BULK").font = bold_font
        ws.cell(row=2, column=1).alignment = center_align
        ws.cell(row=2, column=1).fill = header_fill
        ws.cell(row=2, column=1).border = thin_border

        ws.cell(row=2, column=2, value="").border = thin_border
        ws.cell(row=2, column=2).fill = header_fill

        for idx, lbl in enumerate(data['month_labels'], start=3):
            cell = ws.cell(row=2, column=idx, value=lbl)
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 24

        curr_row = 3

        cat_blocks = [
            ("DRY BULK\n(F/FRM/FG)", "DRY BULK (F/FRM/FG)"),
            ("1.CEMENT", "1.CEMENT"),
            ("2.Break Bulk", "2.Break Bulk"),
            ("1.LIQUID\n[ Inner anchorage]", "1.LIQUID [ Inner anchorage]"),
            ("2.LIQUID (SWB)", "2.LIQUID (SWB)"),
            ("3.LIQUID(JJLTPL)", "3.LIQUID(JJLTPL)"),
            ("Total LIQUID", "Total LIQUID"),
        ]

        metrics_def = [
            ("NO.OF VSL Called", "vsl_called", "0", False, None),
            ("NO.OF VSL Sailed", "vsl_sailed", "0", True, red_font),
            ("AVG.BERTH.WAIT", "avg_berth_wait", "0.00", True, red_font),
            ("        PORT", "port_wait", "0.00", True, red_font),
            ("        AGENT", "agent_wait", "0.00", True, red_font),
            ("STAY AT BERTH", "stay_at_berth", "0.00", False, norm_font),
            ("TURN ROUND(PORT)\n(Stay at berth+Avg. Pre-Berth on Port AC)", "turn_round_port", "0.00", False, norm_font),
            ("AV. IDLE TIME", "av_idle_time", "0.00", True, red_font),
            ("TRAFFIC", "traffic", "#,##0", False, norm_font),
            ("OUTPUT", "output", "#,##0", False, norm_font),
        ]

        for display_name, cat_key in cat_blocks:
            start_r = curr_row
            m_data_list = data["categories"].get(cat_key, [{}]*12)

            is_jjltpl = (cat_key == "3.LIQUID(JJLTPL)")

            for m_label, field, num_fmt, is_red, custom_font in metrics_def:
                r_cell = ws.cell(row=curr_row, column=2, value=m_label)
                r_cell.alignment = left_align
                r_cell.font = custom_font or (red_font if is_red else (bold_font if "Total" in display_name or "3.LIQUID" in display_name else norm_font))
                r_cell.border = thin_border
                if is_jjltpl:
                    r_cell.fill = yellow_fill

                for c_idx in range(12):
                    val = m_data_list[c_idx].get(field, 0) if c_idx < len(m_data_list) else 0
                    v_cell = ws.cell(row=curr_row, column=3 + c_idx, value=val)
                    v_cell.alignment = right_align
                    v_cell.number_format = num_fmt
                    v_cell.font = bold_font if is_jjltpl or "Total" in display_name else norm_font
                    v_cell.border = thin_border
                    if is_jjltpl:
                        v_cell.fill = yellow_fill

                curr_row += 1

            end_r = curr_row - 1
            ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
            cat_cell = ws.cell(row=start_r, column=1, value=display_name)
            cat_cell.font = bold_font
            cat_cell.alignment = center_align
            if is_jjltpl:
                for r_idx in range(start_r, end_r + 1):
                    ws.cell(row=r_idx, column=1).fill = yellow_fill
            cat_cell.border = thin_border

        # Total Liquid Vessels Sailed Row
        r1 = ws.cell(row=curr_row, column=1, value="TOTAL LIQUID\nVESSELS SAILED")
        r1.font = bold_font
        r1.alignment = center_align
        r1.border = thin_border

        r2 = ws.cell(row=curr_row, column=2, value="(INCLUDING IN-ANCH)")
        r2.font = red_font
        r2.alignment = left_align
        r2.border = thin_border

        for c_idx in range(12):
            val = data["total_liquid_sailed"][c_idx] if c_idx < len(data["total_liquid_sailed"]) else 0
            v_cell = ws.cell(row=curr_row, column=3 + c_idx, value=val)
            v_cell.font = bold_font
            v_cell.alignment = right_align
            v_cell.number_format = "0"
            v_cell.border = thin_border

        curr_row += 1

        # Total Liquid Handled Row
        h1 = ws.cell(row=curr_row, column=1, value="TOTAL LIQUID\nHANDLED")
        h1.font = bold_font
        h1.alignment = center_align
        h1.border = thin_border

        h2 = ws.cell(row=curr_row, column=2, value="(INCLUDING IN-ANCH)")
        h2.font = red_font
        h2.alignment = left_align
        h2.border = thin_border

        for c_idx in range(12):
            val = data["total_liquid_handled"][c_idx] if c_idx < len(data["total_liquid_handled"]) else 0
            v_cell = ws.cell(row=curr_row, column=3 + c_idx, value=val)
            v_cell.font = bold_font
            v_cell.alignment = right_align
            v_cell.number_format = "#,##0.000"
            v_cell.border = thin_border

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 42
        for col_letter in [get_column_letter(i) for i in range(3, 15)]:
            ws.column_dimensions[col_letter].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-3_Bulk_Terminal_{fin_year}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

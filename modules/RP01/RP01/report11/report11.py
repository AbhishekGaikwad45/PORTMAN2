"""
Report-11 — Bulk Terminal Performance Report
Flask Blueprint version.

Source: mis_vessel_master (single table — all timing/quantity columns already
present as numeric, per-vessel-call, values in DAYS for the timing columns).

Section A  -> raw sums pulled straight from DB for the selected fin_year
              (all 12 months, Apr-Mar, + FY total column).
Section B  -> derived productivity parameters (real Excel formulas in the
              export; real computed numbers from the FY-aggregated Section A
              in the on-screen report).
Section C  -> same derived parameters expressed in Hours instead of Days.

FY Total column (both the Excel export and the on-screen report) is
calculated by aggregating Section A over the WHOLE financial year first,
then re-running the same B/C ratio formulas on that yearly total — never by
summing or averaging the 12 monthly ratios, which produces meaningless
numbers for anything labeled "Avg." (or Berth Occupancy / Idle time).

Known gaps (no data source in current schema -> always 0 / blank):
    - Vessel Discharge/Load (TEUs), Tonnage (row A3), Crane deployed hours,
      Idle time at NON-working berth (Port/Non-port a/c), Shifting time,
      Total moves, Total TEUs for crane productivity, Rail load/discharge,
      No. of rakes handled, and anything TEU-based (Container throughput,
      Gross Crane Productivity, TRT per 1000 TEUs).
    - No. of berths (row A22) has no DB column -> constant NO_OF_BERTHS
      below (default 2). Change it if the terminal's berth count differs.

Cargo -> broad section (LIQUID / DRY BULK / BREAK BULK) classification
splits compound cargo names (e.g. "SM/IPA/Acetone", "VAM/Aacid") on
'/', '+', '-', whitespace and checks each piece against known short
chemical codes, plus a substring check for multi-word phrases. Anything
still unrecognized is dropped, with a console warning, rather than guessed.
"""

import calendar
import io
import re
import traceback
from functools import wraps
from datetime import datetime, timedelta

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, get_cursor

from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# No DB column carries this -> constant assumption. Adjust if wrong.
NO_OF_BERTHS = 2


class ReportDataError(Exception):
    pass


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


def _days_between(t1, t2):
    """Days (float) from t1 -> t2. 0 if either is missing or negative."""
    if t1 is None or t2 is None:
        return 0.0
    delta = (t2 - t1).total_seconds() / 86400.0
    return delta if delta > 0 else 0.0


def _classify_delay_reason(delay_name):
    """Matches a lueu_parcel_log.delay_name string against the same
    REASONS_PORT / REASONS_NON_PORT legends used in the Excel export.
    Returns 'port', 'non_port', or None if it can't tell / blank."""
    name = str(delay_name or "").strip().upper()
    if not name:
        return None

    port_keywords = (
        "BERTH", "TUG", "PILOT", "EQUIPMENT", "BREAKDOWN", "WORKER",
        "STRIKE", "STOPPAGE", "POWER FAILURE", "LABOUR HOLIDAY",
        "NIGHT NAVIGATION", "DRAFT RESTRICTION",
    )
    non_port_keywords = (
        "SHIP", "SHIPPER", "AGENT", "CARGO", "DEPARTURE", "WEATHER",
        "STORAGE", "TIDAL", "DOCUMENT", "POWER FAILURE GRID", "SCHEDULE",
    )

    if any(k in name for k in port_keywords):
        return "port"
    if any(k in name for k in non_port_keywords):
        return "non_port"
    return None


def _fetch_live_idle_by_parcel_op(cur, parcel_op_ids):
    """Sums idle/delay hours (as day-fractions) from lueu_parcel_log per
    parcel_op_id, split into port / non-port buckets by delay_name."""
    idle = {}  # parcel_op_id -> {"port": days, "non_port": days}
    if not parcel_op_ids:
        return idle

    cur.execute("""
        SELECT parcel_op_id, entry_date, from_time, to_time, delay_name
        FROM lueu_parcel_log
        WHERE COALESCE(is_deleted, FALSE) = FALSE
          AND parcel_op_id = ANY(%s)
          AND delay_name IS NOT NULL
          AND delay_name <> ''
    """, (list(parcel_op_ids),))

    for r in cur.fetchall():
        bucket = _classify_delay_reason(r["delay_name"])
        if bucket is None:
            continue
        try:
            entry_date = str(r["entry_date"]).strip()
            ft = datetime.strptime(f"{entry_date} {r['from_time']}", "%Y-%m-%d %H:%M")
            tt = datetime.strptime(f"{entry_date} {r['to_time']}", "%Y-%m-%d %H:%M")
            if tt <= ft:
                tt += timedelta(days=1)  # spans midnight
            days = (tt - ft).total_seconds() / 86400.0
        except (ValueError, TypeError):
            continue

        pid = r["parcel_op_id"]
        idle.setdefault(pid, {"port": 0.0, "non_port": 0.0})
        idle[pid][bucket] += days

    return idle


def _fetch_live_rows(cur):
    """Current-month vessel-call rows built from LDUD/VCN, in the same
    shape as the mis_vessel_master rows, for months not yet migrated
    into mis_vessel_master.

    quantity is the ACTUAL discharged quantity from lueu_parcel_log
    (excluding is_shortclose = true entries), falling back to the
    originally declared po.quantity only if no log entries exist yet."""
    cur.execute("""
        SELECT
            po.id AS parcel_op_id,
            to_char(current_date, 'YYYY') || '-' ||
                right(to_char(current_date + interval '1 year', 'YYYY'), 2) AS fin_year,
            to_char(current_date, 'Mon-YY') AS month,
            vh.berth_name AS berth_no,
            vh.operation_type AS import_export,
            po.cargo_name AS cargo,
            COALESCE(actual.real_qty, po.quantity::numeric) AS quantity,
            lh.nor_tendered AS nor_tendered,
            lh.nor_accepted AS nor_accepted,
            lh.alongside_datetime AS alongside_datetime,
            lh.cast_off_datetime AS cast_off_datetime,
            lh.pilot_pickup_time AS pilot_pickup_time,
            lh.pilot_board_departure AS pilot_board_departure
        FROM ldud_parcel_ops po
        JOIN ldud_header lh ON lh.id = po.ldud_id
        JOIN vcn_header vh ON vh.id = lh.vcn_id
        LEFT JOIN (
            SELECT parcel_op_id, SUM(quantity) AS real_qty
            FROM lueu_parcel_log
            WHERE is_deleted = false
              AND is_shortclose = false
            GROUP BY parcel_op_id
        ) actual
            ON actual.parcel_op_id = po.id
        WHERE to_char(current_date, 'Mon-YY') = to_char(current_date, 'Mon-YY')
    """)
    raw = cur.fetchall()
    if not raw:
        return []

    parcel_op_ids = [r["parcel_op_id"] for r in raw]
    idle = _fetch_live_idle_by_parcel_op(cur, parcel_op_ids)

    rows = []
    for r in raw:
        nor_tendered = _parse_ts(r["nor_tendered"])
        nor_accepted = _parse_ts(r["nor_accepted"])
        alongside = _parse_ts(r["alongside_datetime"])
        cast_off = _parse_ts(r["cast_off_datetime"])
        pilot_pickup = _parse_ts(r["pilot_pickup_time"])
        pilot_departure = _parse_ts(r["pilot_board_departure"])

        waiting_non_port = _days_between(nor_tendered, nor_accepted)
        waiting_port = _days_between(nor_accepted, alongside)
        stay_at_berth = _days_between(alongside, cast_off)
        inward_movement = _days_between(pilot_pickup, alongside)
        outward_movement = _days_between(pilot_departure, cast_off)

        idle_bucket = idle.get(r["parcel_op_id"], {"port": 0.0, "non_port": 0.0})

        rows.append({
            "fin_year": r["fin_year"],
            "month": r["month"],
            "berth_no": r["berth_no"],
            "import_export": r["import_export"],
            "cargo": r["cargo"],
            "quantity": r["quantity"],
            "pre_berthing_waiting": waiting_port + waiting_non_port,
            "waiting_port": waiting_port,
            "waiting_non_port": waiting_non_port,
            "stay_at_berth": stay_at_berth,
            "inward_movement": inward_movement,
            "outward_movement": outward_movement,
            "non_working_port": idle_bucket["port"],
            "non_working_non_port": idle_bucket["non_port"],
        })
    return rows


def _parse_ts(val):
    """LDUD header fields are free-text ISO-ish datetimes ('2026-07-12T14:40')
    or blank. Returns a datetime or None."""
    if not val or not str(val).strip():
        return None
    val = str(val).strip()
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


def month_str_to_idx(month_str: str) -> int:
    abbrev = str(month_str).split("-")[0].strip()
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized value in mis_vessel_master.month: '{month_str}' "
            f"(expected something like 'Jun-26')"
        )


def days_in_month(fin_year: str, month_idx: int) -> int:
    start_y = fy_start_year(fin_year)
    real_month_num = (month_idx + 3) % 12 + 1          # Apr=4 ... Mar=3
    year = start_y if month_idx < 9 else start_y + 1
    return calendar.monthrange(year, real_month_num)[1]


# ---------------------------------------------------------------------
# Cargo (free text) -> broad section classification
# ---------------------------------------------------------------------
def classify_broad_category(cargo):
    """Classifies a free-text cargo string into LIQUID / DRY BULK / BREAK BULK.

    Cargo names are often compound (e.g. "SM/IPA/Acetone", "VAM/Aacid") because
    a single parcel can carry a blend. We split on '/', '+', '-' and whitespace
    and check each individual piece against known short chemical codes, in
    addition to substring-checking known multi-word phrases against the whole
    string. This avoids silently dropping any cargo whose combined name
    doesn't exactly match one of the old fixed strings."""
    cargo_raw = str(cargo or "").strip().upper()
    if not cargo_raw:
        return None

    # ---------------- LIQUID ----------------
    liquid_phrases = (
        "FURNACE OIL", "POL CRUDE", "RBD PALM OLEIN", "EDIBLE OIL",
        "SUNFLOWER OIL", "ACETIC ACID", "A. ACID", "PHOSPHORIC ACID",
        "PH.ACID", "PH ACID", "BASE OIL", "N BUTONAL", "STRENE MONOMER",
        "STYRENE MONOMER", "NITRIC ACID", "ISOPROPYL ALCOHOL",
        "LPG", "LNG", "LUBE", "SHELL",
    )
    if any(p in cargo_raw for p in liquid_phrases):
        return "LIQUID"

    liquid_tokens = {
        "FO", "CBFS", "CPO", "CPKO", "CDSBO", "CSBO", "CSFO",
        "CHEMICAL", "CHEMICALS", "AACID", "A.ACID", "VAM",
        "PHENOL", "ACETONE", "MDC", "MEK", "IPA", "SM", "MEOH",
        "TOLUNE", "TOLUENE", "METHELENE", "METHYLENE",
        "CHOLORIDE", "CHLORIDE",
    }
    tokens = set(re.split(r'[\/\+\-\s]+', cargo_raw))
    tokens.discard('')
    if tokens & liquid_tokens:
        return "LIQUID"

    # ---------------- DRY BULK ----------------
    dry_bulk_keywords = (
        "IRON ORE", "COAL", "FERTILIZER", "CEMENT", "SALT", "SUGAR",
        "PULSES", "FOOD GRAIN", "TEA", "COFFEE", "SCRAP",
        "CLINKER", "LIMESTONE", "DOLOMITE", "HBI", "FINES",
        "GYPSUM", "BAUXITE", "CLO", "BRBF", "MABU", "VIZAG", "DHAMRA",
    )
    if any(k in cargo_raw for k in dry_bulk_keywords):
        return "DRY BULK"

    # ---------------- BREAK BULK ----------------
    break_bulk_keywords = ("IRON AND STEEL", "TIMBER", "LOG", "PROJECT CARGO")
    if any(k in cargo_raw for k in break_bulk_keywords):
        return "BREAK BULK"

    return None


def _direction(import_export):
    ie = str(import_export or "").strip().upper()
    if ie == "IMPORT":
        return "Import"
    if ie == "EXPORT":
        return "Export"
    return None


def load_data() -> pd.DataFrame:
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT
                fin_year,
                month,
                berth_no,
                import_export,
                cargo,
                quantity,
                pre_berthing_waiting,
                waiting_port,
                waiting_non_port,
                stay_at_berth,
                inward_movement,
                outward_movement,
                non_working_port,
                non_working_non_port
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month IS NOT NULL
        """)
        mis_rows = cur.fetchall()

        current_month = pd.Timestamp.today().strftime("%b-%y")
        mis_current = [r for r in mis_rows if str(r["month"]).strip() == current_month]

        live_rows = []
        if not mis_current:
            print("REPORT11: Current month not found in mis_vessel_master")
            print("REPORT11: Loading live LDUD/LUEU data...")
            live_rows = _fetch_live_rows(cur)
        else:
            print("REPORT11: Current month already present in mis_vessel_master — skipping live load")

        rows = list(mis_rows) + list(live_rows)
    finally:
        conn.close()

    cols = [
        "fin_year", "fy_month_idx", "import_export", "quantity",
        "waiting_port", "waiting_non_port", "stay_at_berth",
        "inward_movement", "outward_movement",
        "non_working_port", "non_working_non_port", "broad_category",
    ]
    if not rows:
        return pd.DataFrame(columns=cols)

    df = pd.DataFrame(rows)
    print(f"REPORT11 DEBUG: rows fetched from DB: {len(df)}")

    df["fin_year"] = df["fin_year"].str.strip()
    df["fy_month_idx"] = df["month"].apply(month_str_to_idx)

    numeric_cols = [
        "quantity", "pre_berthing_waiting", "waiting_port", "waiting_non_port",
        "stay_at_berth", "inward_movement", "outward_movement",
        "non_working_port", "non_working_non_port",
    ]
    for c in numeric_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    df["direction"] = df["import_export"].apply(_direction)
    df["broad_category"] = df["cargo"].apply(classify_broad_category)

    unmapped = sorted(df.loc[df["broad_category"].isna(), "cargo"].dropna().unique().tolist())
    if unmapped:
        print("REPORT11 WARNING: Unmapped cargo values (dropped from Dry/Break/Liquid tonnage rows):")
        for c in unmapped:
            print("   ", c)

    return df


def _get_df_and_years():
    df = load_data()
    years = sorted(df["fin_year"].unique().tolist())
    return df, years


# ---------------------------------------------------------------------
# Section A: raw sums per month, straight from mis_vessel_master
# ---------------------------------------------------------------------
def compute_section_a_month(df, fin_year, month_idx):
    m = df[(df["fin_year"] == fin_year) & (df["fy_month_idx"] == month_idx)]

    vessels_sailed = len(m)  # one row per vessel call

    def tonnes(category, direction):
        sub = m[(m["broad_category"] == category) & (m["direction"] == direction)]
        return round(float(sub["quantity"].sum()), 3)

    a = {
        "Vessel Discharge (Including Restow)": 0.0,
        "Vessel Load (Including Restow)": 0.0,
        "Tonnage": 0.0,
        "Vessels Sailed": vessels_sailed,
        "Pre_berthing Waiting Time-on Port a/c (Total)": round(float(m["waiting_port"].sum()) * 24, 3),
        "Pre_berthing Waiting Time-on Non-Port a/c (Total)": round(float(m["waiting_non_port"].sum()) * 24, 3),
        "Total Berth Stay of all vessels (For Berth Productivity)": round(float(m["stay_at_berth"].sum()) * 24, 3),
        "Total Crane deplyoed hours (For Crane Productivity)": 0.0,
        "Vessel Inward movement (Total)": round(float(m["inward_movement"].sum()) * 24, 3),
        "Vessel Outward movement (Total)": round(float(m["outward_movement"].sum()) * 24, 3),
        "Idle time at working berth on Port A/c.": round(float(m["non_working_port"].sum()) * 24, 3),
        "Idle time at working berth on Non-Port A/c.": round(float(m["non_working_non_port"].sum()) * 24, 3),
        "Idle time at Non-working berth on Port A/c.": 0.0,
        "Idle time at Non-working berth on Non-Port A/c.": 0.0,
        "Shifting Time": 0.0,
        "Total No. of Moves for calculating  Berth / Crane Productivity": 0.0,
        "Total No. of TEUs for calculating  Crane Productivity": 0.0,
        "Rail Load": 0.0,
        "Rail Discharge": 0.0,
        "No. of Rakes handled": 0.0,
        "Days in a month": days_in_month(fin_year, month_idx),
        "No. of berths for % berth occupancy": NO_OF_BERTHS,
        "Dry Bulk traffic - Import": tonnes("DRY BULK", "Import"),
        "Dry Bulk traffic - Export": tonnes("DRY BULK", "Export"),
        "Break Bulk traffic - Import": tonnes("BREAK BULK", "Import"),
        "Break Bulk traffic - Export": tonnes("BREAK BULK", "Export"),
        "Liquid - Import": tonnes("LIQUID", "Import"),
        "Liquid - Export": tonnes("LIQUID", "Export"),
    }
    return a


def compute_section_a_fy(df, fin_year):
    """Same as compute_section_a_month, but aggregated across the WHOLE
    financial year instead of a single month — this becomes the real FY
    Total for Section A, and (fed into compute_section_b_month /
    compute_section_c_month) the real FY ratios for Sections B and C."""
    m = df[df["fin_year"] == fin_year]

    vessels_sailed = len(m)

    def tonnes(category, direction):
        sub = m[(m["broad_category"] == category) & (m["direction"] == direction)]
        return round(float(sub["quantity"].sum()), 3)

    total_days = sum(days_in_month(fin_year, idx) for idx in range(12))

    a = {
        "Vessel Discharge (Including Restow)": 0.0,
        "Vessel Load (Including Restow)": 0.0,
        "Tonnage": 0.0,
        "Vessels Sailed": vessels_sailed,
        "Pre_berthing Waiting Time-on Port a/c (Total)": round(float(m["waiting_port"].sum()) * 24, 3),
        "Pre_berthing Waiting Time-on Non-Port a/c (Total)": round(float(m["waiting_non_port"].sum()) * 24, 3),
        "Total Berth Stay of all vessels (For Berth Productivity)": round(float(m["stay_at_berth"].sum()) * 24, 3),
        "Total Crane deplyoed hours (For Crane Productivity)": 0.0,
        "Vessel Inward movement (Total)": round(float(m["inward_movement"].sum()) * 24, 3),
        "Vessel Outward movement (Total)": round(float(m["outward_movement"].sum()) * 24, 3),
        "Idle time at working berth on Port A/c.": round(float(m["non_working_port"].sum()) * 24, 3),
        "Idle time at working berth on Non-Port A/c.": round(float(m["non_working_non_port"].sum()) * 24, 3),
        "Idle time at Non-working berth on Port A/c.": 0.0,
        "Idle time at Non-working berth on Non-Port A/c.": 0.0,
        "Shifting Time": 0.0,
        "Total No. of Moves for calculating  Berth / Crane Productivity": 0.0,
        "Total No. of TEUs for calculating  Crane Productivity": 0.0,
        "Rail Load": 0.0,
        "Rail Discharge": 0.0,
        "No. of Rakes handled": 0.0,
        "Days in a month": total_days,
        "No. of berths for % berth occupancy": NO_OF_BERTHS,
        "Dry Bulk traffic - Import": tonnes("DRY BULK", "Import"),
        "Dry Bulk traffic - Export": tonnes("DRY BULK", "Export"),
        "Break Bulk traffic - Import": tonnes("BREAK BULK", "Import"),
        "Break Bulk traffic - Export": tonnes("BREAK BULK", "Export"),
        "Liquid - Import": tonnes("LIQUID", "Import"),
        "Liquid - Export": tonnes("LIQUID", "Export"),
    }
    return a


def _safe_div(n, d):
    return (n / d) if d else 0.0


def compute_section_b_month(a):
    total_traffic_tons = (
        a["Dry Bulk traffic - Import"] + a["Dry Bulk traffic - Export"]
        + a["Break Bulk traffic - Import"] + a["Break Bulk traffic - Export"]
        + a["Liquid - Import"] + a["Liquid - Export"]
    )
    vs = a["Vessels Sailed"]
    berth_stay_hrs = a["Total Berth Stay of all vessels (For Berth Productivity)"]

    b = {
        "Vessels Sailed": vs,
        "Total Traffic Throughputs (TEUs)": a["Vessel Discharge (Including Restow)"] + a["Vessel Load (Including Restow)"],
        "Total traffic throughputs (Tons)": round(total_traffic_tons, 3),
        "Parcel Size": 0.0,  # TEU-based, no source
        "Avg. Pre-berthing Waiting Time-Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"],
            24 * vs), 6),
        "Avg. Pre-berthing Waiting Time-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Port a/c (Total)"], 24 * vs), 6),
        "Avg. Pre-berthing Waiting Time-Non-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], 24 * vs), 6),
        "Avg. Berth stay": round(_safe_div(berth_stay_hrs, 24 * vs), 6),
        "Avg. Turn around time - Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"]
            + berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], 24 * vs), 6),
        "Avg. Turn around time - Port A/c.": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + berth_stay_hrs
            + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], 24 * vs), 6),
        "Avg. Turn around time - Non- Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], 24 * vs), 6),
        "Avg. Turn around time - Pilot Boarding to De-boarding-Total": round(_safe_div(
            berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], 24 * vs), 6),
        "Berth Occupancy": round(_safe_div(berth_stay_hrs, a["Days in a month"] * 24 * a["No. of berths for % berth occupancy"]), 6),
        "Idle time": round(_safe_div(
            a["Idle time at working berth on Port A/c."] + a["Idle time at working berth on Non-Port A/c."],
            berth_stay_hrs), 6),
        "Gross Berth Productivity": 0.0,
        "Gross Crane Productivity (Moves)": None,
        "Gross Crane Productivity (TEUs)": None,
        "Ship Output per Day (TEUs)": 0.0,
        "No. of Rakes handled": a["No. of Rakes handled"],
        "Total Rail traffic": a["Rail Load"] + a["Rail Discharge"],
        "% wrt to Total Thoughput": None,
    }
    b["Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c."] = b["Avg. Turn around time - Pilot Boarding to De-boarding-Total"]
    b["Ship Output per Day (Tonnes)"] = round(_safe_div(b["Total traffic throughputs (Tons)"], b["Avg. Berth stay"]), 3) if b["Avg. Berth stay"] else 0.0
    return b


def compute_section_c_month(a, b):
    vs = a["Vessels Sailed"]
    berth_stay_hrs = a["Total Berth Stay of all vessels (For Berth Productivity)"]

    c = {
        "Avg. Pre-berthing Waiting Time-Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], vs), 6),
        "Avg. Pre-berthing Waiting Time-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Port a/c (Total)"], vs), 6),
        "Avg. Pre-berthing Waiting Time-Non-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], vs), 6),
        "Avg. Berth stay": round(_safe_div(berth_stay_hrs, vs), 6),
        "Avg. Turn around time - Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"]
            + berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], vs), 6),
        "Avg. Turn around time - Port A/c.": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + berth_stay_hrs
            + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], vs), 6),
        "Avg. Turn around time - Non- Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], vs), 6),
        "Avg. Turn around time - Pilot Boarding to De-boarding-Total": round(_safe_div(
            berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], vs), 6),
        "Avg. TRT for 1000 tonnes (Days)": 0.0,
        "Avg. TRT for 1000 tonnes (Hrs)": 0.0,
        "Avg. TRT for 1000 TEUs (Days)": None,
        "Avg. TRT for 1000 TEUs (Hrs)": None,
    }
    c["Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c."] = c["Avg. Turn around time - Pilot Boarding to De-boarding-Total"]

    tat_total_days = b["Avg. Turn around time - Total"]
    total_tons = b["Total traffic throughputs (Tons)"]
    trt_days = round(_safe_div(tat_total_days * vs * 1000, total_tons), 6) if total_tons else 0.0
    c["Avg. TRT for 1000 tonnes (Days)"] = trt_days
    c["Avg. TRT for 1000 tonnes (Hrs)"] = round(trt_days * 24, 6)
    return c


REASONS_PORT = [
    "1. Non Availability of Berth", "2. Non Availability of Tugs", "3. Non Availability of Pilot",
    "4. Non availability of equipment.", "5. Equipment breakdown", "6. Absence of workers",
    "7. Strike / Stoppage", "8. Power failure", "9. Labour holidays",
    "10. Night navigation restrictions", "11. Draft Restriction", "12. Others",
]
REASONS_NON_PORT = [
    "1. Ships account", "2.Shippers account", "3. Agents options",
    "4. Absence of non-port workers", "5. For want of cargo", "6. Departure formalities",
    "7. Weather restrictions", "8. Lack of storage", "9. Tidal",
    "10. Documents not Ready", "11. Power Failure Grid", "12. Not in Schedule", "13. Others",
]
# ---------------------------------------------------------------------
# FY Total override for rows marked "Formula need to check"
# For these specific rows, the FY Total column should NOT be a
# recomputed ratio from yearly-aggregated Section A. Instead it should
# be a plain SUM of the 12 monthly values already shown on screen.
# All other rows keep the original ratio-based FY Total behavior.
# ---------------------------------------------------------------------
SUM_FY_LABELS_B = [
    "Avg. Pre-berthing Waiting Time-Total",
    "Avg. Pre-berthing Waiting Time-Port A/c.",
    "Avg. Pre-berthing Waiting Time-Non-Port A/c.",
    "Avg. Berth stay",
    "Avg. Turn around time - Total",
    "Avg. Turn around time - Port A/c.",
    "Avg. Turn around time - Non- Port A/c.",
    "Avg. Turn around time - Pilot Boarding to De-boarding-Total",
    "Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c.",
    "Berth Occupancy",
    "Idle time",
]

SUM_FY_LABELS_C = [
    "Avg. Pre-berthing Waiting Time-Total",
    "Avg. Pre-berthing Waiting Time-Port A/c.",
    "Avg. Pre-berthing Waiting Time-Non-Port A/c.",
    "Avg. Berth stay",
    "Avg. Turn around time - Total",
    "Avg. Turn around time - Port A/c.",
    "Avg. Turn around time - Non- Port A/c.",
    "Avg. Turn around time - Pilot Boarding to De-boarding-Total",
    "Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c.",
    "Avg. TRT for 1000 tonnes (Days)",
    "Avg. TRT for 1000 tonnes (Hrs)",
]


def _apply_sum_fy_override(fy_dict, monthly_list, labels):
    """
    Given the FY dict (originally ratio-based) and the list of 12 monthly
    dicts, overrides fy_dict[label] with a plain SUM of the 12 monthly
    values — but only for the labels marked "Formula need to check".
    Every other key in fy_dict is left untouched (still the real
    ratio-based FY value computed from yearly-aggregated Section A).
    """
    for label in labels:
        vals = [m.get(label) for m in monthly_list if m.get(label) is not None]
        if vals:
            fy_dict[label] = round(sum(vals), 6)
    return fy_dict


@bp.route("/module/RP01/report11/")
@login_required
def report11_index():
    return render_template("report11/report11.html", port_name="JJLTPL")


@bp.route("/api/module/RP01/report11/meta")
@login_required
def report11_api_meta():
    try:
        _, years = _get_df_and_years()
        return jsonify({"years": years, "port_name": "JJLTPL"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report11/report")
@login_required
def report11_api_report():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1] if years else None)
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        months = month_options_for(fin_year)
        a_by_month, b_by_month, c_by_month = [], [], []
        for mo in months:
            a = compute_section_a_month(df, fin_year, mo["idx"])
            b = compute_section_b_month(a)
            c = compute_section_c_month(a, b)
            a_by_month.append(a)
            b_by_month.append(b)
            c_by_month.append(c)

        # Real FY totals — Section A summed over the whole year, then B/C
        # derived from that FY-level Section A (same ratio formulas, just
        # fed yearly inputs instead of one month's inputs).
        a_fy = compute_section_a_fy(df, fin_year)
        b_fy = compute_section_b_month(a_fy)
        c_fy = compute_section_c_month(a_fy, b_fy)

        # NEW — For rows marked "Formula need to check", replace the
        # ratio-based FY Total with a plain SUM of the 12 monthly values.
        b_fy = _apply_sum_fy_override(b_fy, b_by_month, SUM_FY_LABELS_B)
        c_fy = _apply_sum_fy_override(c_fy, c_by_month, SUM_FY_LABELS_C)

        return jsonify({
            "port_name": "JJLTPL",
            "fin_year": fin_year,
            "months": [m["label"] for m in months],
            "section_a": a_by_month,
            "section_b": b_by_month,
            "section_c": c_by_month,
            "fy_total": {
                "section_a": a_fy,
                "section_b": b_fy,
                "section_c": c_fy,
            },
        })
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


# ---------------------------------------------------------------------
# Excel export — same grid layout as the uploaded workbook:
#   B=SR.NO, C=PARTICULARS, D=Units, E:P=Apr..Mar, Q=FY Total
#
# WHY THIS VERSION IS DIFFERENT FROM THE FORMULA-BASED ONE:
#   Every openpyxl formula cell is written to disk as
#       <f>SOME_FORMULA</f><v></v>
#   i.e. the formula text with an EMPTY cached value. Full desktop Excel
#   *usually* recalculates that automatically on open — but Excel Online,
#   WPS Office, mobile Excel, LibreOffice with auto-recalc off, and most
#   quick-preview panes just read the cached <v> and never run a calc
#   engine at all, so every formula cell renders blank regardless of any
#   fullCalcOnLoad setting. That's exactly the "Section B/C empty" bug.
#
#   Fix: don't ask Excel to compute anything. Section B and C are
#   computed in Python already (compute_section_b_month /
#   compute_section_c_month — the same functions the on-screen report
#   uses), so we just write those literal numbers straight into the
#   cells. Guaranteed to show correct data in any viewer, with zero
#   dependency on Excel's recalculation behavior. The FY Total column
#   uses compute_section_a_fy -> compute_section_b_month/c_month, i.e.
#   the real FY-aggregated ratios — never a sum/average of the 12
#   monthly ratios.
# ---------------------------------------------------------------------
def _write_row(ws, row_i, sr, label, unit, values, fy_value, fmt, thin_border):
    ws[f"B{row_i}"] = sr
    ws[f"C{row_i}"] = label
    ws[f"D{row_i}"] = unit
    for i, v in enumerate(values):
        col = get_column_letter(5 + i)  # E=5
        cell = ws[f"{col}{row_i}"]
        cell.value = v
        cell.number_format = fmt
        cell.border = thin_border
    # FY total column (Q = col 17) — literal computed value, not a formula
    q = ws[f"Q{row_i}"]
    q.value = fy_value
    q.number_format = fmt
    q.border = thin_border
    for c in ("B", "C", "D"):
        ws[f"{c}{row_i}"].border = thin_border



@bp.route("/api/module/RP01/report11/export")
@login_required
def report11_api_export():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1] if years else None)
        port_name = request.args.get("port_name", "JJLTPL")

        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        months = month_options_for(fin_year)
        a_list, b_list, c_list = [], [], []
        for mo in months:
            a = compute_section_a_month(df, fin_year, mo["idx"])
            b = compute_section_b_month(a)
            c = compute_section_c_month(a, b)
            a_list.append(a)
            b_list.append(b)
            c_list.append(c)

        # Real FY totals — identical logic to the on-screen report: Section A
        # aggregated over the whole year first, then B/C derived from that.
        a_fy = compute_section_a_fy(df, fin_year)
        b_fy = compute_section_b_month(a_fy)
        c_fy = compute_section_c_month(a_fy, b_fy)

        # NEW — Same override applied here so the Excel export matches
        # the on-screen report for rows marked "Formula need to check".
        b_fy = _apply_sum_fy_override(b_fy, b_list, SUM_FY_LABELS_B)
        c_fy = _apply_sum_fy_override(c_fy, c_list, SUM_FY_LABELS_C)

        wb = Workbook()
        ws = wb.active
        ws.title = "Bulk Terminal Performance"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=13)
        header_font = Font(bold=True)
        section_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["B2"] = "Bulk Terminal - Liquid JJLTPL"
        ws["B2"].font = title_font
        ws["B3"] = f"Performance Report {fin_year}"
        ws["B3"].font = bold
        ws["B4"] = f"Port / Terminal - {port_name}"
        ws["B4"].font = bold

        header_row = 6
        ws[f"B{header_row}"] = "SR.NO."
        ws[f"C{header_row}"] = "PARTICULARS"
        ws[f"D{header_row}"] = "Units"
        for i, mo in enumerate(months):
            ws[f"{get_column_letter(5 + i)}{header_row}"] = mo["label"]
        ws[f"Q{header_row}"] = f"FY {fin_year}"
        for col_idx in range(2, 18):
            cell = ws[f"{get_column_letter(col_idx)}{header_row}"]
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

        row_i = header_row + 1

        # ---------------- Section A ----------------
        ws[f"B{row_i}"] = "A)"
        ws[f"C{row_i}"] = "TERMINAL PERFOMANCE (To be filled by Terminals)"
        ws[f"B{row_i}"].font = section_font
        ws[f"C{row_i}"].font = section_font
        row_i += 1

        a_rows = [
            ("Vessel Discharge (Including Restow)", "TEUs", "0.0"),
            ("Vessel Load (Including Restow)", "TEUs", "0.0"),
            ("Tonnage", "Tonnes", "0.0"),
            ("Vessels Sailed", "Nos", "0"),
            ("Pre_berthing Waiting Time-on Port a/c (Total)", "Hrs", "0.00"),
            ("Pre_berthing Waiting Time-on Non-Port a/c (Total)", "Hrs.", "0.00"),
            ("Total Berth Stay of all vessels (For Berth Productivity)", "Hrs", "0.00"),
            ("Total Crane deplyoed hours (For Crane Productivity)", "Hrs", "0.0"),
            ("Vessel Inward movement (Total)", "Hrs", "0.00"),
            ("Vessel Outward movement (Total)", "Hrs", "0.00"),
            ("Idle time at working berth on Port A/c.", "Hrs.", "0.00"),
            ("Idle time at working berth on Non-Port A/c.", "Hrs.", "0.00"),
            ("Idle time at Non-working berth on Port A/c.", "Hrs.", "0.0"),
            ("Idle time at Non-working berth on Non-Port A/c.", "Hrs.", "0.0"),
            ("Shifting Time", "Hrs.", "0.0"),
            ("Total No. of Moves for calculating  Berth / Crane Productivity", "Moves", "0.0"),
            ("Total No. of TEUs for calculating  Crane Productivity", "TEUs", "0.0"),
            ("Rail Load", "TEUs", "0.0"),
            ("Rail Discharge", "TEUs", "0.0"),
            ("No. of Rakes handled", "Nos", "0.0"),
            ("Days in a month", "Nos", "0"),
            ("No. of berths for % berth occupancy", "Nos", "0"),
            ("Dry Bulk traffic - Import", "Tonnes", "0.000"),
            ("Dry Bulk traffic - Export", "Tonnes", "0.000"),
            ("Break Bulk traffic - Import", "Tonnes", "0.000"),
            ("Break Bulk traffic - Export", "Tonnes", "0.000"),
            ("Liquid - Import", "Tonnes", "0.000"),
            ("Liquid - Export", "Tonnes", "0.000"),
        ]
        for sr, (label, unit, fmt) in enumerate(a_rows, start=1):
            values = [a_list[i][label] for i in range(12)]
            fy_val = NO_OF_BERTHS if label == "No. of berths for % berth occupancy" else a_fy[label]
            _write_row(ws, row_i, sr, label, unit, values, fy_val, fmt, thin_border)
            row_i += 1

        row_i += 1
        # ---------------- Section B ----------------
        ws[f"B{row_i}"] = "B)"
        ws[f"C{row_i}"] = "PRODUCTIVITY PARAMETERS (Derived - Not to be filled)"
        ws[f"B{row_i}"].font = section_font
        ws[f"C{row_i}"].font = section_font
        row_i += 1

        b_rows = [
            ("Vessels Sailed", "Nos.", "0"),
            ("Total Traffic Throughputs (TEUs)", "TEUs", "0"),
            ("Total traffic throughputs (Tons)", "Tons", "0.000"),
            ("Parcel Size", "TEUs", "0"),
            ("Avg. Pre-berthing Waiting Time-Total", "Days", "0.000000"),
            ("Avg. Pre-berthing Waiting Time-Port A/c.", "Days", "0.000000"),
            ("Avg. Pre-berthing Waiting Time-Non-Port A/c.", "Days", "0.000000"),
            ("Avg. Berth stay", "Days", "0.000000"),
            ("Avg. Turn around time - Total", "Days", "0.000000"),
            ("Avg. Turn around time - Port A/c.", "Days", "0.000000"),
            ("Avg. Turn around time - Non- Port A/c.", "Days", "0.000000"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Total", "Days", "0.000000"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c.", "Days", "0.000000"),
            ("Berth Occupancy", "%", "0.00%"),
            ("Idle time", "%", "0.00%"),
            ("Gross Berth Productivity", "Moves /Hrs", "0"),
            ("Gross Crane Productivity (Moves)", "Moves /Hrs", "0"),
            ("Gross Crane Productivity (TEUs)", "TEUs/Hrs", "0"),
            ("Ship Output per Day (TEUs)", "TEUs", "0"),
            ("Ship Output per Day (Tonnes)", "Tonnes", "0.000"),
        ]
        for sr, (label, unit, fmt) in enumerate(b_rows, start=1):
            values = [b_list[i].get(label) for i in range(12)]
            fy_val = b_fy.get(label)
            _write_row(ws, row_i, sr, label, unit, values, fy_val, fmt, thin_border)
            row_i += 1

        row_i += 1
        for label, unit, fmt in [
            ("No. of Rakes handled", "Nos", "0"),
            ("Total Rail traffic", "TEUs", "0"),
            ("% wrt to Total Thoughput", "%", "0.0%"),
        ]:
            ws[f"C{row_i}"] = label
            ws[f"D{row_i}"] = unit
            row_i += 1

        row_i += 1
        # ---------------- Section C ----------------
        ws[f"B{row_i}"] = "C)"
        ws[f"C{row_i}"] = "PRODUCTIVITY PARAMETERS (Derived - Not to be filled)"
        ws[f"B{row_i}"].font = section_font
        ws[f"C{row_i}"].font = section_font
        row_i += 1

        c_rows = [
            ("Avg. Pre-berthing Waiting Time-Total", "Hrs.", "0.000000"),
            ("Avg. Pre-berthing Waiting Time-Port A/c.", "Hrs.", "0.000000"),
            ("Avg. Pre-berthing Waiting Time-Non-Port A/c.", "Hrs.", "0.000000"),
            ("Avg. Berth stay", "Hrs.", "0.000000"),
            ("Avg. Turn around time - Total", "Hrs.", "0.000000"),
            ("Avg. Turn around time - Port A/c.", "Hrs.", "0.000000"),
            ("Avg. Turn around time - Non- Port A/c.", "Hrs.", "0.000000"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Total", "Hrs.", "0.000000"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c.", "Hrs.", "0.000000"),
            ("Avg. TRT for 1000 tonnes (Days)", "Days", "0.000000"),
            ("Avg. TRT for 1000 tonnes (Hrs)", "Hrs.", "0.000000"),
        ]
        for sr, (label, unit, fmt) in enumerate(c_rows, start=1):
            values = [c_list[i].get(label) for i in range(12)]
            fy_val = c_fy.get(label)
            _write_row(ws, row_i, sr, label, unit, values, fy_val, fmt, thin_border)
            row_i += 1

        row_i += 1
        for label, unit in [("Avg. TRT for 1000 TEUs", "Days"), ("Avg. TRT for 1000 TEUs", "Hrs.")]:
            ws[f"C{row_i}"] = label
            ws[f"D{row_i}"] = unit
            row_i += 1

        # ---------------- Reasons block (static legend) ----------------
        row_i += 1
        ws[f"C{row_i}"] = "Reasons to be Considered"
        ws[f"C{row_i}"].font = bold
        row_i += 1
        ws[f"C{row_i}"] = "Port A/c."
        ws[f"D{row_i}"] = "Non-Port  A/c."
        ws[f"C{row_i}"].font = header_font
        ws[f"D{row_i}"].font = header_font
        row_i += 1
        for i in range(max(len(REASONS_PORT), len(REASONS_NON_PORT))):
            if i < len(REASONS_PORT):
                ws[f"C{row_i}"] = REASONS_PORT[i]
            if i < len(REASONS_NON_PORT):
                ws[f"D{row_i}"] = REASONS_NON_PORT[i]
            row_i += 1

        # ---- column widths -------------------------------------------------
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 48
        ws.column_dimensions["D"].width = 10
        for c in range(5, 18):
            ws.column_dimensions[get_column_letter(c)].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-11_Bulk_Terminal_Performance_{fin_year}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500
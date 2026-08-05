"""
Report-5 — Performance Report for the Month (Berth / Commodity / Vessels /
Occupancy / Handling). Flask Blueprint version. Reads from Postgres.

======================================================================
 DATA SOURCE CUTOVER
======================================================================
The Berth/Commodity/TOTAL-HANDLING/Commodity *label* columns are hard-coded
on the frontend (report5.html) — this backend only needs to supply, per
selected month:
    - quantity  (total handling, in tonnes)
    - vessel_count (distinct vessels for that month)
    - berth occupancy % (see BERTH OCCUPANCY section below)

Two different source systems are used depending on the month:

    - Apr .. Jun  ->  mis_vessel_master   (legacy system)
    - Jul .. Mar  ->  lueu_parcel_log (quantity) + ldud_header (vessel_name)
                      (new system, live from CUTOVER_DATE below)

The cutover is a ONE-TIME SYSTEM MIGRATION DATE, not a recurring
"July every year" rule — i.e. Jul-Aug-Sep-2025 is still legacy data,
but Jul-Aug-Sep-2026 onward is new-system data. Adjust CUTOVER_DATE
below if that's wrong.

======================================================================
 ASSUMPTIONS THAT STILL NEED CONFIRMATION
======================================================================
1. Vessel count for Jul+ uses ldud_header.created_date (per user
   instruction) to determine which month a vessel record counts toward.
2. lueu_parcel_log.entry_date / ldud_header's date column are stored as
   free-text (not a real DATE/TIMESTAMP column), matching the pattern
   seen on lueu_parcel_log. Rows are fetched and parsed in Python
   (via dateutil) rather than filtered in SQL, specifically to survive
   inconsistent text date formats. If entry_date turns out to be a real
   date/timestamp column, the parsing in `_parse_text_date()` still
   works, but you could simplify to a SQL WHERE clause instead.
3. mis_vessel_master.vessel_name is assumed to be the correct column
   for counting distinct vessels Apr-Jun (per your confirmation).
4. BERTH OCCUPANCY — column names (MIS_ALONGSIDE_COLUMN /
   MIS_CASTOFF_COLUMN / LDUD_ALONGSIDE_COLUMN / LDUD_CASTOFF_COLUMN
   below) are GUESSES and need confirmation. Also assumed: a vessel's
   hours are booked entirely to its ALONGSIDE month (no splitting
   across a month boundary), and "divide by 2" in the occupancy
   formula means 2 physical berths (NUM_BERTHS = 2).
5. SHORT CLOSE: lueu_parcel_log has a column flagging a parcel as "short closed". 
   Per user instruction, any row where this flag is true has its quantity
   SUBTRACTED from the running total instead of added, everywhere
   lueu_parcel_log.quantity is summed.

======================================================================
 DEBUGGING NOTE (added)
======================================================================
Both berth-hours fetch functions now return a tuple of
(total_hours, debug_rows) instead of just a float. `debug_rows` is a
list of per-vessel dicts (vessel_name, alongside, castoff, hours) for
every row that actually contributed to the sum, so the numbers going
into the occupancy % can be checked against a manual list of vessels
instead of guessing whether the berth_no filter is really being
applied server-side. This is threaded through fetch_berth_hours() and
surfaced in compute_berth_occupancy()'s result under "debug_rows", and
therefore also under result["occupancy_detail"]["debug_rows"] in the
/report endpoint response.

fetch_quantity_from_parcel_log() now also returns per-row debug info
(see "debug_rows" in its return value) showing, for every matched row,
its quantity, whether short_close was true, and how much it
contributed (positive or negative) to the total — so a short-close
subtraction can be checked against a manual list the same way berth
hours can.

======================================================================
 EXPORT-TO-EXCEL (added)
======================================================================
report5_export() below is the implementation behind the "Export Excel"
button in report5.html (window.location =
"/api/module/RP01/report5/export?month=...&year=..."), which previously
had no matching route.

No external template file is used — build_report5_export_workbook()
builds the workbook entirely in code with openpyxl, so this module has
no dependency on any other file on disk. The exact layout of the
original "5. Monthly Rep Sum" sheet (every label, merge, border, fill,
number format, and formula) is captured as static data in
EXPORT_SUM_CELLS / EXPORT_SUM_MERGES / EXPORT_SUM_COLUMN_WIDTHS /
EXPORT_SUM_ROW_HEIGHTS below, and re-applied to a fresh sheet on every
export. After that fixed layout is laid down, this month's numbers are
written into the handful of cells that hold real (non-formula) data —
Table 1's "Liquid (LB-03&LB-04)" row, Table 2's "LIQUID (JJLTPL)" row,
Table 3's Liquid/Cement rows — the same handful of "real" data points
applyReportToPerformanceData() / applyReportToVolumeComparisonData() in
report5.html fill into an otherwise-fixed table shape client-side.
Every other cell (labels + the row 11/22/23 totals, increase-% columns,
etc., which are all still live formulas) is left exactly as captured.
======================================================================
"""

import calendar
import io
import logging
import os
import traceback
from datetime import date
from functools import wraps

from dateutil import parser as dateutil_parser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from flask import jsonify, request, render_template, session, redirect, url_for, send_file

from database import get_db, get_cursor
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from .. import bp

logger = logging.getLogger("report5")
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter(
        "[report5] %(levelname)s %(message)s"
    ))
    logger.addHandler(_handler)
logger.setLevel(logging.DEBUG)

DEBUG_BERTH_OCCUPANCY = True


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
MONTH_NUM = {  # calendar month number, for date filtering against the new tables
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}
NUM_TO_MONTH_ABBREV = {v: k for k, v in MONTH_NUM.items()}


MONTH_ABBREV_TO_FULL = {
    "Jan": "January", "Feb": "February", "Mar": "March", "Apr": "April",
    "May": "May", "Jun": "June", "Jul": "July", "Aug": "August",
    "Sep": "September", "Oct": "October", "Nov": "November", "Dec": "December",
}


CUTOVER_DATE = date(2026, 7, 1)

LDUD_HEADER_DATE_COLUMN = "cast_off_datetime"


LIQUID_BERTH_CODES_MIS = ["LB-03", "LB-04"]  # <-- CONFIRM this is the full/correct set

MIS_ALONGSIDE_COLUMN = "alongside"      # confirmed column, text (needs datetime parse)
MIS_CASTOFF_COLUMN = "cast_off"         # CONFIRM vs sail_cast_off - which is "cast off"?
LDUD_ALONGSIDE_COLUMN = "alongside_datetime"  # confirmed column, text (needs datetime parse)
LDUD_CASTOFF_COLUMN = "cast_off_datetime"     # confirmed column, text (needs datetime parse)
NUM_BERTHS = 2  # per formula: divide by 2 -> 2 physical berths

MIS_COMMODITY_COLUMN = "cargo"            # CONFIRMED: real column on mis_vessel_master
LDUD_COMMODITY_COLUMN = None              # N/A: no commodity column on ldud_header; new-system data assumed Liquid-only
PARCEL_LOG_COMMODITY_COLUMN = None        # N/A: no commodity column on lueu_parcel_log; new-system data assumed Liquid-only

# CONFIRMED 31-Jul-2026 via information_schema.columns: real column is
# is_shortclose, type boolean (not a text flag) -- no truthiness
# workaround needed. See ASSUMPTION (5) at the top of this file.



MIS_HISTORY_MONTH_COLUMN = "month_jsw"
MIS_HISTORY_COMMODITY_COLUMN = "cargo_type"
MIS_HISTORY_QUANTITY_COLUMN = "quantity"

# --- Export-to-Excel layout (no external file) ---
EXPORT_SUMMARY_SHEET = "5. Monthly Rep Sum"

# Column widths / row heights / merges from the original "5. Monthly
# Rep Sum" sheet layout.
EXPORT_SUM_COLUMN_WIDTHS = {
    "A": 2.109375, "B": 24.88671875, "C": 19.6640625, "D": 13.88671875,
    "E": 12.6640625, "F": 17.33203125, "G": 16.33203125, "H": 11.88671875,
    "I": 12.33203125, "J": 9.5546875, "K": 14.6640625, "L": 9.5546875,
}
EXPORT_SUM_ROW_HEIGHTS = {3: 15.6, 22: 17.25, 23: 17.25, 26: 13.8, 27: 15.75, 28: 15.75, 39: 15.6}
EXPORT_SUM_MERGES = ["B3:I3", "C27:E27", "F27:H27", "C28:E28", "F28:H28"]


EXPORT_SUM_CELLS = [
    ('B2', '  ', ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('B3', 'Performance Report for the Month of June-2026', ('Arial', 12.0, True, False, 'single', None), None, (None, None, None, None), ('center', None), 'General'),
    ('B5', 'Berth', ('Arial', 10.0, True, False, None, None), None, (('thin', None), None, ('thin', None), None), (None, None), 'General'),
    ('C5', 'Commodity', ('Arial', 10.0, True, False, None, None), None, (('thin', None), None, ('thin', None), None), (None, None), 'General'),
    ('D5', 'No. of ', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), ('center', None), 'General'),
    ('E5', 'Berth', ('Arial', 10.0, True, False, None, None), None, (None, ('thin', None), ('thin', None), None), ('center', None), 'General'),
    ('F5', 'Handling', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), ('center', None), 'General'),
    ('G5', 'Handling', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), ('center', None), 'General'),
    ('H5', 'Increase', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), ('center', None), 'General'),
    ('B6', None, ('Arial', 10.0, True, False, None, None), None, (('thin', None), None, None, ('thin', None)), (None, None), 'General'),
    ('C6', None, ('Arial', 10.0, True, False, None, None), None, (('thin', None), None, None, ('thin', None)), (None, None), 'General'),
    ('D6', 'Vessels sailed', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), 'General'),
    ('E6', 'Occupancy', ('Arial', 10.0, True, False, None, None), None, (None, ('thin', None), None, ('thin', None)), ('center', None), 'General'),
    ('F6', 'June-25(Previous)', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), None, None), ('center', None), 'General'),
    ('G6', 'June-26(Current)', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), None, None), ('center', None), 'General'),
    ('H6', '%', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), 'General'),
    ('B7', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('C7', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('D7', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), ('center', None), 'General'),
    ('E7', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('F7', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('G7', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('H7', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), 'General'),
    ('B8', 'Liquid Berth', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C8', 'Liquid (LB-01&LB-02)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('D8', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), ('center', None), 'General'),
    ('E8', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F8', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('G8', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('H8', '=(G8-F8)/F22', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), '0.00%'),
    ('J8', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), (None, None), '0.00'),
    ('B9', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C9', 'Liquid (LB-03&LB-04)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('D9', 12, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), ('center', None), 'General'),
    ('E9', 0.2446236559141741, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F9', 117960.688, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.000'),
    ('G9', 92676.773, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.000'),
    ('H9', '=(G9-F9)/F23', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), '0.00%'),
    ('B10', 'ANCHORAGE', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C10', 'INN ANCH', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('D10', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), ('center', None), 'General'),
    ('E10', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F10', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.000'),
    ('G10', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.000'),
    ('H10', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), '0.00%'),
    ('J10', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), (None, None), '0.00'),
    ('B11', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C11', 'Total Liquid', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('D11', '=SUM(D8:D10)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), ('center', None), 'General'),
    ('E11', '=E9', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F11', '=SUM(F8:F10)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.000'),
    ('G11', '=SUM(G8:G10)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.000'),
    ('H11', '=(G11-F11)/F23', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), '0.00%'),
    ('J11', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), (None, None), '0.00'),
    ('B12', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C12', 'Others', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('D12', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), ('center', None), 'General'),
    ('E12', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('F12', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('G12', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('H12', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), 'General'),
    ('J12', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), (None, None), '0.00'),
    ('B13', 'Shallow Berth ', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C13', 'cement', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D13', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('E13', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F13', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('G13', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('H13', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0.00%'),
    ('J13', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), (None, None), '0.00'),
    ('B14', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C14', 'others', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('D14', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('E14', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F14', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0'),
    ('G14', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('H14', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0.00%'),
    ('B15', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('C15', 'Container', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('D15', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('E15', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F15', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('G15', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('H15', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0.00%'),
    ('B16', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C16', 'liquid', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D16', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('center', None), 'General'),
    ('E16', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F16', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('G16', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0'),
    ('H16', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0.00%'),
    ('B17', 'Anchorage', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C17', 'Liquid', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('D17', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('E17', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F17', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0'),
    ('G17', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0'),
    ('H17', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), '0.00%'),
    ('B18', 'Coastal Berth', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('C18', 'cement', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D18', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('E18', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F18', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('right', None), '0'),
    ('G18', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0'),
    ('H18', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0.00%'),
    ('B19', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('C19', 'Others', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D19', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('E19', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F19', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('right', None), '0'),
    ('G19', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0'),
    ('H19', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('right', None), '0.00%'),
    ('B20', 'CB01/CB-02', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C20', 'CEMENT', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D20', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('center', None), 'General'),
    ('E20', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), '0.00%'),
    ('F20', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0'),
    ('G20', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0'),
    ('H20', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('right', None), '0.00%'),
    ('B21', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, ('thin', None)), (None, None), 'General'),
    ('C21', 'Others', ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (('thin', None), ('thin', None), None, ('thin', None)), (None, None), 'General'),
    ('D21', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), 'General'),
    ('E21', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), (None, None), '0.00%'),
    ('F21', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('right', None), '0'),
    ('G21', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('right', None), '0'),
    ('H21', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('right', None), '0.00%'),
    ('B22', 'TOTAL LIQUID ', ('Arial', 10.0, False, False, None, None), 'FF66CCFF', (('thin', None), ('thin', None), None, ('thin', None)), (None, None), 'General'),
    ('C22', None, ('Arial', 10.0, False, False, None, None), 'FF66CCFF', (('thin', None), None, None, ('thin', None)), (None, None), 'General'),
    ('D22', '=D8+D9', ('Arial', 10.0, False, False, None, None), 'FF66CCFF', (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), 'General'),
    ('E22', '=E8+E9', ('Arial', 10.0, False, False, None, None), 'FF66CCFF', (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), '0.00%'),
    ('F22', '=F8+F9', ('Arial', 10.0, False, False, None, None), 'FF66CCFF', (('thin', None), ('thin', None), None, ('thin', None)), ('right', None), '0.000'),
    ('G22', '=G8+G9', ('Arial', 10.0, False, False, None, None), 'FF66CCFF', (('thin', None), ('thin', None), None, ('thin', None)), ('right', None), '0.000'),
    ('H22', '=(G22-F22)/F22', ('Arial', 10.0, False, False, None, None), 'FF66CCFF', (None, ('thin', None), None, None), ('right', None), '0.00%'),
    ('B23', 'Total', ('Arial', 10.0, True, False, None, None), 'FF92D050', (('thin', None), ('thin', None), ('thin', None), ('thin', None)), (None, None), 'General'),
    ('C23', None, ('Arial', 10.0, True, False, None, None), 'FF92D050', (('thin', None), ('thin', None), ('thin', None), ('thin', None)), (None, None), 'General'),
    ('D23', '=SUM(D8:D10)', ('Arial', 10.0, True, False, None, None), 'FF92D050', (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), 'General'),
    ('E23', '=SUM(E8:E10)', ('Arial', 10.0, True, False, None, None), 'FF92D050', (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), '0.00%'),
    ('F23', '=SUM(F7:F9)', ('Arial', 10.0, True, False, None, None), 'FF92D050', (('thin', None), None, None, ('thin', None)), ('right', None), '0.000'),
    ('G23', '=SUM(G7:G10)', ('Arial', 10.0, True, False, None, None), 'FF92D050', (('thin', None), None, None, ('thin', None)), ('right', None), '0.000'),
    ('H23', '=(G23-F23)/F23', ('Arial', 10.0, False, False, None, None), 'FF92D050', (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('right', None), '0.00%'),
    ('B24', None, ('Arial', 10.0, True, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('B25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('C25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('D25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('E25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('F25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('G25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('H25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('I25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('J25', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B26', 'COMPARISON OF VOLUME HANDLED IN THE CURRENT YEAR VIS-\u00c0-VIS LAST YEAR', ('Arial', 11.0, True, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('C26', None, ('Arial', 11.0, True, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('D26', None, ('Arial', 11.0, True, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('E26', None, ('Arial', 11.0, True, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('F26', None, ('Arial', 11.0, True, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('G26', None, ('Arial', 11.0, True, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('H26', None, ('Arial', 11.0, True, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('I26', None, ('Arial', 11.0, False, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('J26', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B27', '         TOTAL HANDLING', ('Arial', 10.0, True, False, None, None), None, (('thin', None), None, ('thin', None), None), (None, None), 'General'),
    ('C27', '2025-26', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), ('center', 'center'), 'General'),
    ('E27', None, ('Arial', 11.0, False, False, None, None), None, (None, ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('F27', ' 2026-27', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), ('center', 'center'), 'General'),
    ('H27', None, ('Arial', 11.0, False, False, None, None), None, (None, ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('I27', 'INCREASE', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), ('center', None), 'General'),
    ('J27', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B28', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('C28', "APR'2025 \u2013 MAR-2026", ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('center', 'center'), 'General'),
    ('E28', None, ('Arial', 11.0, False, False, None, None), None, (None, ('thin', None), None, ('thin', None)), (None, None), 'General'),
    ('F28', "APR' 2026 \u2013 MAR-2027", ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), ('center', None), 'General'),
    ('H28', None, ('Arial', 11.0, False, False, None, None), None, (None, ('thin', None), None, ('thin', None)), (None, None), 'General'),
    ('I28', '      %', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, ('thin', None)), (None, None), 'General'),
    ('J28', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B29', 'LIQUID JNPT (INN ANCH)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('C29', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('D29', "='[1]5. Monthly Rep Det'!D113", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('E29', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('F29', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, ('thin', None), None), ('center', None), 'General'),
    ('G29', "='[1]5. Monthly Rep Det'!G113", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('H29', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), ('thin', None), None), ('center', None), 'General'),
    ('I29', 0, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J29', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B30', 'LIQUID JNPT ( SWB)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D30', "='[1]5. Monthly Rep Det'!D114", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('F30', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('G30', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('H30', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('I30', '=(G30-D30)*100/D30', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J30', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B31', 'LIQUID       (BPCL)', ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D31', "='[1]5. Monthly Rep Det'!D115", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('F31', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('G31', "='[1]5. Monthly Rep Det'!G115", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('H31', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('I31', '=(G31-D31)*100/D31', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J31', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B32', 'LIQUID (JJLTPL)', ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('C32', None, ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (None, None, None, None), (None, None), 'General'),
    ('D32', 283589.208, ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (None, None, None, None), ('center', None), '0.000'),
    ('E32', None, ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (None, None, None, None), (None, None), 'General'),
    ('F32', None, ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (('thin', None), None, None, None), (None, None), 'General'),
    ('G32', 279990.995, ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (None, None, None, None), ('center', None), '0.000'),
    ('H32', None, ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (None, ('thin', None), None, None), (None, None), 'General'),
    ('I32', '=(G32-D32)*100/D32', ('Arial', 10.0, False, False, None, None), 'FFFFFF00', (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J32', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('K32', None, ('Verdana', 10.0, False, False, None, None), None, (None, None, None, None), (None, 'center'), '#,##0.00'),
    ('L32', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), (None, None), '#,##0.000'),
    ('B33', 'TOTAL LIQUID', ('Arial', 10.0, True, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D33', '=SUM(D30:D31)', ('Arial', 10.0, True, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('F33', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('G33', '=SUM(G29:G32)', ('Arial', 10.0, True, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('H33', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('I33', '=(G33-D33)*100/D33', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J33', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B34', 'CEMENT', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D34', "='[1]5. Monthly Rep Det'!D118", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('F34', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('G34', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('H34', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('I34', '=(G34-D34)*100/D34', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J34', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B35', 'BREAK BULK', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D35', "='[1]5. Monthly Rep Det'!D119", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('E35', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('F35', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('G35', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('H35', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('I35', '=(G35-D35)*100/D35', ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J35', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B36', 'OTHERS', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D36', "='[1]5. Monthly Rep Det'!D120", ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('F36', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('G36', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.000'),
    ('H36', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), (None, None), 'General'),
    ('I36', 0, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), None, None), ('center', None), '0.00'),
    ('J36', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('B37', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), (None, None), 'General'),
    ('C37', None, ('Arial', 10.0, False, False, None, None), None, (None, None, ('thin', None), ('thin', None)), (None, None), 'General'),
    ('D37', '=D29+D30+D31+D34+D35+D36+D32', ('Arial', 10.0, True, False, None, None), None, (None, None, ('thin', None), ('thin', None)), ('center', None), '0.000'),
    ('E37', None, ('Arial', 10.0, False, False, None, None), None, (None, None, ('thin', None), ('thin', None)), (None, None), 'General'),
    ('F37', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, ('thin', None), ('thin', None)), (None, None), 'General'),
    ('G37', '=G29+G30+G31+G34+G35+G36+G32', ('Arial', 10.0, True, False, None, None), None, (None, None, ('thin', None), ('thin', None)), ('center', None), '0.000'),
    ('H37', None, ('Arial', 10.0, False, False, None, None), None, (None, ('thin', None), ('thin', None), ('thin', None)), (None, None), 'General'),
    ('I37', '=I29+I30+I31+I34+I35+I36+I32', ('Arial', 10.0, True, False, None, None), None, (None, None, ('thin', None), ('thin', None)), ('center', None), '0.000'),
    ('J37', None, ('Arial', 10.0, False, False, None, 'FFFF0000'), None, (None, None, None, None), (None, None), 'General'),
    ('D38', None, ('Arial', 10.0, True, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('G38', None, ('Arial', 10.0, True, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('I38', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), '0.00'),
    ('B39', 'Commoditiwise Average Turn around Time(Berthing to sailing) & parcel size', ('Arial', 12.0, True, False, None, None), None, (None, None, None, None), (None, None), 'General'),
    ('B40', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, ('thin', None), None), (None, None), 'General'),
    ('C40', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), None), (None, None), 'General'),
    ('D40', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, ('thin', None), None), (None, None), 'General'),
    ('E40', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('B41', 'Commodity', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('C41', 'Av. Turn around', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), (None, None), 'General'),
    ('D41', 'Av. Parcel size', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('E41', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('B42', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, ('thin', None)), (None, None), 'General'),
    ('C42', '(days)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), None, None), ('center', None), 'General'),
    ('D42', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), '0'),
    ('E42', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, None), (None, None), 'General'),
    ('B43', 'Liquid (current)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, ('thin', None), None), ('right', None), 'General'),
    ('C43', 1.2638888888899, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0.00'),
    ('D43', 7723.064416666667, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0'),
    ('B44', '(previous)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, ('thin', None)), ('right', None), 'General'),
    ('C44', 1.3906067251461394, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0.00'),
    ('D44', 5351.805578947369, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0'),
    ('B45', 'Cement (current)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, ('thin', None), None), (None, None), 'General'),
    ('C45', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0.00'),
    ('D45', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0'),
    ('B46', '(previous)', ('Arial', 10.0, False, False, None, None), None, (('thin', None), None, None, ('thin', None)), ('right', None), 'General'),
    ('C46', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0.00'),
    ('D46', None, ('Arial', 10.0, False, False, None, None), None, (('thin', None), ('thin', None), ('thin', None), ('thin', None)), ('center', None), '0'),
   
    
    
    ('E50', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('F50', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('G50', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),
    ('H50', None, ('Arial', 10.0, False, False, None, None), None, (None, None, None, None), ('center', None), 'General'),

]

def _apply_export_layout(ws):
    # Column widths
    for col, width in EXPORT_SUM_COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Row heights
    for row, height in EXPORT_SUM_ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height

    # Merged cells
    for rng in EXPORT_SUM_MERGES:
        ws.merge_cells(rng)

    # Apply all captured cells
    for coord, value, font, fill, border, align, number_format in EXPORT_SUM_CELLS:
        cell = ws[coord]

        # Skip merged cells except the master cell
        if isinstance(cell, MergedCell):
            continue

        if isinstance(value, str) and value.startswith("="):
            value = value.replace("'[1]5. Monthly Rep Det'!", "'5. Monthly Rep Det'!")
            value = value.replace("[1]", "")

        cell.value = value

        if font:
            cell.font = Font(
                name=font[0],
                size=font[1],
                bold=font[2],
                italic=font[3],
                underline=font[4],
                color=font[5]
            )

        if fill:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=fill.replace("#", "")
            )

        if border:
            def mk(side):
                if side is None:
                    return Side()
                return Side(style=side[0], color=side[1])

            cell.border = Border(
                left=mk(border[0]),
                right=mk(border[1]),
                top=mk(border[2]),
                bottom=mk(border[3])
            )

        if align:
            cell.alignment = Alignment(
                horizontal=align[0],
                vertical=align[1]
            )

        if number_format:
            cell.number_format = number_format


class ReportDataError(Exception):
    """Raised for any problem loading/validating the report's source data.
    Caught by the route handlers and turned into a clean JSON error response."""
    pass


def _parse_fin_year_start(raw) -> int:
    """Normalize a fin_year value into its starting calendar year (int).
    Handles a plain int/numeric-string ('2026'), a range string
    ('2026-27' or '2026-2027'), and strips stray whitespace. Raises
    ValueError if it can't be parsed, so callers can fall back safely."""
    if raw is None:
        raise ValueError("fin_year value is None")
    s = str(raw).strip()
    if not s:
        raise ValueError("fin_year value is empty")
    # Take the leading run of digits as the starting year, e.g.
    # "2026-27" -> "2026", "2026-2027" -> "2026", "2026" -> "2026".
    head = s.split("-")[0].strip()
    return int(head)


def fetch_latest_fin_year_start() -> int:
    """Most recent fiscal year (as its starting calendar year) present in
    mis_vessel_master.fin_year. Falls back to today's calendar-year-based
    fin year if the table/column is empty or unparsable, so a bad/missing
    fin_year value never crashes the report."""
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("SELECT DISTINCT fin_year FROM mis_vessel_master WHERE fin_year IS NOT NULL")
        rows = cur.fetchall()
    finally:
        conn.close()

    parsed = []
    for r in rows:
        try:
            parsed.append(_parse_fin_year_start(r["fin_year"]))
        except (ValueError, TypeError):
            continue  # skip unparsable values rather than failing the whole report

    if parsed:
        return max(parsed)

    # Fallback: derive fin_year_start from today's date the same way the
    # rest of this module thinks about fiscal years (Apr-Mar).
    today = date.today()
    return today.year if today.month >= 4 else today.year - 1


def month_str_to_idx(abbrev: str) -> int:
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized month '{abbrev}'. Expected one of: {', '.join(MONTH_NAMES)}"
        )


def calendar_year_for_month(fin_year_start: int, month_abbrev: str) -> int:
    """fin_year_start=2026 ('2026-27') + month 'Jun' -> 2026;
    fin_year_start=2026 + month 'Feb' -> 2027 (Jan-Mar fall in the FY end year)."""
    return fin_year_start + 1 if month_abbrev in ("Jan", "Feb", "Mar") else fin_year_start


def uses_legacy_source(month_abbrev: str, calendar_year: int) -> bool:
    """True -> read from mis_vessel_master. False -> read from the new
    lueu_parcel_log / ldud_header tables."""
    month_num = MONTH_NUM[month_abbrev]
    # Use day=1 of that calendar month to compare against CUTOVER_DATE.
    month_date = date(calendar_year, month_num, 1)
    return month_date < CUTOVER_DATE


def _parse_text_date(raw):
    """Best-effort parse of a free-text date column into a date(). Returns
    None if it can't be parsed (row is skipped rather than crashing the
    whole report).

    CONFIRMED FORMAT (per user, screenshot of ldud_header data):
    'YYYY-MM-DDTHH:MM' — ISO, year-first. dateutil_parser.parse() already
    reads a leading 4-digit year correctly regardless of dayfirst/
    yearfirst, so this alone wasn't an ambiguity bug for that exact
    shape. But we now parse with yearfirst=True (and dayfirst=False) to
    match the confirmed format explicitly rather than relying on
    dateutil's general heuristics, so any row that ISN'T in this format
    is more likely to fail loudly / get flagged in debug output instead
    of being silently misinterpreted."""
    if not raw:
        return None
    try:
        return dateutil_parser.parse(str(raw), yearfirst=True, dayfirst=False).date()
    except (ValueError, TypeError):
        return None


def _parse_text_datetime(raw):
    """Like _parse_text_date, but keeps the time component (needed for
    hours-alongside calculations, not just which day something falls
    on). Returns None if it can't be parsed.

    See _parse_text_date() docstring re: confirmed 'YYYY-MM-DDTHH:MM'
    format and yearfirst=True."""
    if not raw:
        return None
    try:
        return dateutil_parser.parse(str(raw), yearfirst=True, dayfirst=False)
    except (ValueError, TypeError):
        return None


def _days_in_month(month_abbrev: str, calendar_year: int) -> int:
    month_num = MONTH_NUM[month_abbrev]
    return calendar.monthrange(calendar_year, month_num)[1]


# ---------------------------------------------------------------------
# LEGACY SOURCE (Apr - Jun): mis_vessel_master
# ---------------------------------------------------------------------

def fetch_from_mis_vessel_master(month_abbrev: str, calendar_year: int):
    """quantity + distinct vessel count for one calendar month, from the
    legacy table. mis_vessel_master.month is stored like 'Jun-26'."""
    yy = str(calendar_year)[-2:]
    month_str = f"{month_abbrev}-{yy}"

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute(
            """
            SELECT
                COALESCE(SUM(quantity), 0) AS quantity,
                COUNT(DISTINCT vessel_name) AS vessel_count
            FROM mis_vessel_master
            WHERE month = %(month_str)s
            """,
            {"month_str": month_str},
        )
        row = cur.fetchone()
    finally:
        conn.close()

    if not row:
        return {"quantity": 0.0, "vessel_count": 0}

    return {
        "quantity": float(row["quantity"] or 0),
        "vessel_count": int(row["vessel_count"] or 0),
    }


# ---------------------------------------------------------------------
# NEW SOURCE (Jul onward): lueu_parcel_log (quantity) + ldud_header (vessels)
# ---------------------------------------------------------------------





def fetch_quantity_from_parcel_log(month_abbrev: str, calendar_year: int) -> float:
    month_num = MONTH_NUM[month_abbrev]

    conn = get_db()
    try:
        cur = get_cursor(conn)

        cur.execute("""
            SELECT
                COALESCE(p.quantity, 0) AS quantity,
                h.cast_off_datetime AS castoff_raw,
                h.vessel_name
            FROM lueu_parcel_log p
            JOIN ldud_parcel_ops o
                ON o.id = p.parcel_op_id
            JOIN ldud_header h
                ON h.id = o.ldud_id
            WHERE COALESCE(p.is_deleted, false) = false
              AND COALESCE(h.is_deleted, false) = false
              AND COALESCE(p.is_shortclose, false) = false
              AND p.quantity IS NOT NULL
              AND NULLIF(TRIM(h.cast_off_datetime), '') IS NOT NULL
        """)

        rows = cur.fetchall()

    finally:
        conn.close()

    total = 0.0

    for r in rows:
        castoff = _parse_text_datetime(r["castoff_raw"])

        if not castoff:
            continue

        if castoff.year != calendar_year or castoff.month != month_num:
            continue

        total += float(r["quantity"] or 0)

    logger.info(
        "fetch_quantity_from_parcel_log(%s,%s) = %.3f",
        month_abbrev,
        calendar_year,
        total,
    )

    return round(total, 3)


def fetch_vessel_count_from_ldud_header(month_abbrev: str, calendar_year: int) -> int:
    """Distinct vessel_name count from ldud_header for the given calendar
    month/year. See ASSUMPTION (1) at the top of this file re: date column."""
    month_num = MONTH_NUM[month_abbrev]

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute(
            f"""
            SELECT vessel_name, {LDUD_HEADER_DATE_COLUMN} AS event_date
            FROM ldud_header
            WHERE
                COALESCE(is_deleted,false)=false
                AND cast_off_datetime IS NOT NULL
            """
        )
        rows = cur.fetchall()
    finally:
        conn.close()

    vessels = set()
    for r in rows:
        d = _parse_text_date(r["event_date"])
        if d and d.year == calendar_year and d.month == month_num:
            vessels.add(r["vessel_name"])

    return len(vessels)


# ---------------------------------------------------------------------
# LIVE DATA FALLBACK (borrowed from Report-12's logic)
#


def month_already_in_mis(month_abbrev: str, calendar_year: int) -> bool:
    """Simple yes/no check: does mis_vessel_master already have at
    least one row for this month? (Same idea as Report-12's
    "current month already in mis_rows" check.)"""
    yy = str(calendar_year)[-2:]
    month_str = f"{month_abbrev}-{yy}"

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute(
            "SELECT COUNT(*) AS cnt FROM mis_vessel_master WHERE month = %(month_str)s",
            {"month_str": month_str},
        )
        row = cur.fetchone()
    finally:
        conn.close()

    count = int(row["cnt"] or 0) if row else 0
    logger.debug(
        "month_already_in_mis: month=%s -> %d row(s) found in mis_vessel_master",
        month_str, count,
    )
    return count > 0


def fetch_live_month_figures(month_abbrev: str, calendar_year: int):
    """Beginner-style version of Report-12's "live VCN" query, adapted
    to Report-5's simpler quantity + vessel_count shape.

    Steps:
      1. Get every vessel call from vcn_header, joined to ldud_header
         (for the real cast-off date/time) and to the vcn cargo tables
         (for quantity).
      2. Loop over the rows in plain Python.
      3. Keep only the rows whose cast-off date falls in the month we
         were asked about.
      4. Add up the quantity and collect the distinct vessel names.

    Returns the same shape as fetch_from_mis_vessel_master() /
    fetch_month_figures():
        {"quantity": <float>, "vessel_count": <int>}
    """
    month_num = MONTH_NUM[month_abbrev]

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute(
            """
            SELECT
                h.vessel_name AS vessel_name,
                lh.cast_off_datetime AS castoff_raw,
                COALESCE(imp.quantity, 0) + COALESCE(exp.quantity, 0) AS quantity
            FROM vcn_header h
            JOIN ldud_header lh
                ON lh.vcn_id = h.id
            LEFT JOIN (
                SELECT vcn_id, SUM(quantity::numeric) AS quantity
                FROM vcn_consigners
                GROUP BY vcn_id
            ) imp
                ON imp.vcn_id = h.id
            LEFT JOIN (
                SELECT vcn_id, SUM(quantity::numeric) AS quantity
                FROM vcn_export_cargo_declaration
                GROUP BY vcn_id
            ) exp
                ON exp.vcn_id = h.id
            WHERE lh.cast_off_datetime IS NOT NULL
              AND NULLIF(TRIM(lh.cast_off_datetime), '') IS NOT NULL
            """
        )
        rows = cur.fetchall()
    finally:
        conn.close()

    # Step 2 & 3 & 4: plain loop, no pandas.
    total_quantity = 0.0
    vessel_names_seen = set()

    for row in rows:
        cast_off = _parse_text_datetime(row["castoff_raw"])

        # Skip rows we can't parse a date from.
        if cast_off is None:
            continue

        # Only keep rows that belong to the requested month/year.
        if cast_off.year != calendar_year or cast_off.month != month_num:
            continue

        total_quantity += float(row["quantity"] or 0)
        vessel_names_seen.add(row["vessel_name"])

    result = {
        "quantity": round(total_quantity, 3),
        "vessel_count": len(vessel_names_seen),
    }

    logger.debug(
        "fetch_live_month_figures: month=%s year=%s -> quantity=%.3f vessel_count=%d "
        "(from %d live vcn_header rows)",
        month_abbrev, calendar_year, result["quantity"], result["vessel_count"], len(rows),
    )

    return result


# ---------------------------------------------------------------------
# UNIFIED FETCH — picks the right source per month
# ---------------------------------------------------------------------
#
# CHANGED: for new-system months (Jul onward), this now also checks
# mis_vessel_master first (Report-12 style). Only when that month is
# NOT yet present in mis_vessel_master does it go live via
# fetch_live_month_figures(); if the live data also comes back empty,
# it falls back to the original lueu_parcel_log/ldud_header source so
# the report never goes completely blank. Apr-Jun (legacy) behaviour
# is completely unchanged.

def fetch_month_figures(month_abbrev: str, calendar_year: int):
    """Decide where this month's numbers should come from.

    Step 1 - Apr/May/Jun (before CUTOVER_DATE): always the legacy
             mis_vessel_master table. Nothing changes here.

    Step 2 - Jul onward (new system): first check if mis_vessel_master
             already has this month (same check Report-12 does). If it
             does, just use it - it's the "final" migrated data.

    Step 3 - If mis_vessel_master does NOT have this month yet, that
             means the month is still "live" / in progress, so pull the
             numbers straight from the VCN tables instead of showing
             zero.

    Step 4 - If, for any reason, the live fetch comes back completely
             empty (e.g. no VCN rows exist yet either), fall back to the
             original lueu_parcel_log / ldud_header source so the report
             still shows something instead of nothing.
    """

    # Step 1: legacy months never change.
    if uses_legacy_source(month_abbrev, calendar_year):
        return fetch_from_mis_vessel_master(month_abbrev, calendar_year)

    # Step 2: new-system month - is it already in mis_vessel_master?
    if month_already_in_mis(month_abbrev, calendar_year):
        logger.debug(
            "fetch_month_figures: %s-%s already in mis_vessel_master, using it",
            month_abbrev, calendar_year,
        )
        return fetch_from_mis_vessel_master(month_abbrev, calendar_year)

    # Step 3: not migrated yet -> go live.
    # The live data MUST come from parcel_log for Report-5 so that the
    # short-close subtraction logic is applied to the quantities.
    logger.debug(
        "fetch_month_figures: %s-%s NOT yet in mis_vessel_master, "
        "pulling live data from lueu_parcel_log/ldud_header",
        month_abbrev, calendar_year,
    )
    
    return {
        "quantity": fetch_quantity_from_parcel_log(month_abbrev, calendar_year),
        "vessel_count": fetch_vessel_count_from_ldud_header(month_abbrev, calendar_year),
    }


def fiscal_year_start_for(month_abbrev: str, calendar_year: int) -> int:
    """Inverse of calendar_year_for_month: given a specific month/calendar-year,
    return the fiscal year's starting calendar year (Apr of that FY)."""
    return calendar_year - 1 if month_abbrev in ("Jan", "Feb", "Mar") else calendar_year


def fetch_fiscal_year_quantity_total(fin_year_start: int, upto_month: str) -> float:
    total = 0.0

    for month_abbrev in MONTH_NAMES:
        calendar_year = calendar_year_for_month(fin_year_start, month_abbrev)
        figures = fetch_month_figures(month_abbrev, calendar_year)
        total += figures["quantity"]

        # Stop once the selected month is reached
        if month_abbrev == upto_month:
            break

    return round(total, 3)


def compute_report(month_abbrev: str, calendar_year: int):
    """Current month/year vs the same month one year earlier."""
    month_str_to_idx(month_abbrev)  # validates month_abbrev

    current = fetch_month_figures(month_abbrev, calendar_year)
    previous = fetch_month_figures(month_abbrev, calendar_year - 1)

    prev_qty = previous["quantity"]
    curr_qty = current["quantity"]
    increase_pct = ((curr_qty - prev_qty) / prev_qty * 100) if prev_qty else None

    return {
        "month": month_abbrev,
        "year": calendar_year,
        "current": current,
        "previous": previous,
        "increase_pct": round(increase_pct, 2) if increase_pct is not None else None,
    }


def compute_fiscal_year_comparison(fin_year_start: int, month_abbrev: str, calendar_year: int):
    """Current fiscal year vs the prior fiscal year, each summed across
    its own 12 months (mis + lueu combined per month, per year), plus the
    month-over-month report for the requested month."""
    current_total = fetch_fiscal_year_quantity_total(fin_year_start, month_abbrev)
    previous_total = fetch_fiscal_year_quantity_total(fin_year_start - 1, month_abbrev)

    result = compute_report(month_abbrev, calendar_year)

    increase_pct = (
        ((current_total - previous_total) / previous_total * 100)
        if previous_total else None
    )

    result["fin_year_start"] = fin_year_start
    result["fin_year_label"] = f"{fin_year_start}-{str(fin_year_start + 1)[-2:]}"
    result["prev_fin_year_label"] = f"{fin_year_start - 1}-{str(fin_year_start)[-2:]}"
    result["current_jjltpl"] = current_total
    result["previous_jjltpl"] = previous_total
    result["jjltpl_increase_pct"] = round(increase_pct, 2) if increase_pct is not None else None

    return result


# ---------------------------------------------------------------------
# BERTH OCCUPANCY
# ---------------------------------------------------------------------


def fetch_berth_hours_from_mis_vessel_master(month_abbrev: str, calendar_year: int):
    """Total alongside-hours (sum of castoff - alongside, in hours) for
    the legacy table, for vessels whose alongside falls in this month.
    mis_vessel_master.month is stored like 'Jun-26' — same pattern as
    fetch_from_mis_vessel_master above.

    Returns (total_hours, debug_rows)."""
    yy = str(calendar_year)[-2:]
    month_str = f"{month_abbrev}-{yy}"

    logger.debug(
        "fetch_berth_hours_from_mis_vessel_master: month=%s year=%s "
        "-> querying month='%s' berth_no IN %s using columns alongside=%r castoff=%r",
        month_abbrev, calendar_year, month_str, LIQUID_BERTH_CODES_MIS,
        MIS_ALONGSIDE_COLUMN, MIS_CASTOFF_COLUMN,
    )

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute(
            f"""
            SELECT vessel_name,
                   {MIS_ALONGSIDE_COLUMN} AS alongside_raw,
                   {MIS_CASTOFF_COLUMN} AS castoff_raw
            FROM mis_vessel_master
            WHERE month = %(month_str)s
              AND berth_no = ANY(%(berth_codes)s)
            """,
            {"month_str": month_str, "berth_codes": LIQUID_BERTH_CODES_MIS},
        )
        rows = cur.fetchall()
    except Exception:
        logger.exception(
            "fetch_berth_hours_from_mis_vessel_master: query FAILED "
            "(check MIS_ALONGSIDE_COLUMN / MIS_CASTOFF_COLUMN are real "
            "column names on mis_vessel_master)"
        )
        raise
    finally:
        conn.close()

    logger.debug(
        "fetch_berth_hours_from_mis_vessel_master: fetched %d row(s) for month='%s'",
        len(rows), month_str,
    )

    total_hours = 0.0
    debug_rows = []
    skipped_unparsed = 0
    skipped_bad_order = 0

    for i, r in enumerate(rows):
        vessel_name = r.get("vessel_name") if hasattr(r, "get") else r["vessel_name"]
        raw_alongside = r["alongside_raw"]
        raw_castoff = r["castoff_raw"]
        alongside = _parse_text_datetime(raw_alongside)
        castoff = _parse_text_datetime(raw_castoff)

        if not alongside or not castoff:
            skipped_unparsed += 1
            if DEBUG_BERTH_OCCUPANCY:
                logger.debug(
                    "  row %d: UNPARSEABLE vessel=%r alongside_raw=%r -> %r | castoff_raw=%r -> %r (skipped)",
                    i, vessel_name, raw_alongside, alongside, raw_castoff, castoff,
                )
            continue

        if castoff <= alongside:
            skipped_bad_order += 1
            if DEBUG_BERTH_OCCUPANCY:
                logger.debug(
                    "  row %d: vessel=%r castoff <= alongside (alongside=%s castoff=%s) - skipped",
                    i, vessel_name, alongside, castoff,
                )
            continue

        hours = (castoff - alongside).total_seconds() / 3600.0
        total_hours += hours
        debug_rows.append({
            "vessel_name": vessel_name,
            "alongside": alongside.isoformat(),
            "castoff": castoff.isoformat(),
            "hours": round(hours, 2),
        })
        if DEBUG_BERTH_OCCUPANCY:
            logger.debug(
                "  row %d: vessel=%r alongside=%s castoff=%s -> %.2f hours (running total=%.2f)",
                i, vessel_name, alongside, castoff, hours, total_hours,
            )

    logger.debug(
        "fetch_berth_hours_from_mis_vessel_master: TOTAL=%.2f hours "
        "(%d rows used, %d unparsed, %d bad-order)",
        total_hours, len(rows) - skipped_unparsed - skipped_bad_order,
        skipped_unparsed, skipped_bad_order,
    )

    return total_hours, debug_rows


def fetch_berth_hours_from_ldud_header(month_abbrev: str, calendar_year: int):
    """Total alongside-hours for the new table, for vessels whose
    CAST-OFF date falls in this calendar month/year (matches the
    convention already used for vessel_count via LDUD_HEADER_DATE_COLUMN
    and for commodity turnaround via fetch_commodity_turnaround_from_new_system:
    a vessel is attributed to the month it actually left, not the month
    it arrived)."""
    month_num = MONTH_NUM[month_abbrev]

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute(
            f"""
            SELECT vessel_name,
                   {LDUD_ALONGSIDE_COLUMN} AS alongside_raw,
                   {LDUD_CASTOFF_COLUMN} AS castoff_raw
            FROM ldud_header
            WHERE COALESCE(is_deleted, false) = false
            """
        )
        rows = cur.fetchall()
    except Exception:
        logger.exception(
            "fetch_berth_hours_from_ldud_header: query FAILED "
            "(check LDUD_ALONGSIDE_COLUMN / LDUD_CASTOFF_COLUMN are real "
            "column names on ldud_header)"
        )
        raise
    finally:
        conn.close()

    total_hours = 0.0
    debug_rows = []
    skipped_unparsed = 0
    skipped_bad_order = 0
    skipped_other_month = 0
    matched_rows = 0

    for i, r in enumerate(rows):
        vessel_name = r.get("vessel_name") if hasattr(r, "get") else r["vessel_name"]
        raw_alongside = r["alongside_raw"]
        raw_castoff = r["castoff_raw"]
        alongside = _parse_text_datetime(raw_alongside)
        castoff = _parse_text_datetime(raw_castoff)

        if not alongside or not castoff:
            skipped_unparsed += 1
            continue

        if castoff <= alongside:
            skipped_bad_order += 1
            continue

        # CHANGED: filter by CAST-OFF month/year, not alongside month/year
        if not (castoff.year == calendar_year and castoff.month == month_num):
            skipped_other_month += 1
            continue

        hours = (castoff - alongside).total_seconds() / 3600.0
        total_hours += hours
        matched_rows += 1
        debug_rows.append({
            "vessel_name": vessel_name,
            "alongside": alongside.isoformat(),
            "castoff": castoff.isoformat(),
            "hours": round(hours, 2),
        })

    logger.debug(
        "fetch_berth_hours_from_ldud_header: TOTAL=%.2f hours for %s %s "
        "(%d rows matched this month by CAST-OFF date, %d unparsed, %d bad-order, %d other-month)",
        total_hours, month_abbrev, calendar_year,
        matched_rows, skipped_unparsed, skipped_bad_order, skipped_other_month,
    )

    return total_hours, debug_rows


def fetch_berth_hours(month_abbrev: str, calendar_year: int):
    """Unified fetch, using the same legacy/new cutover logic as the rest
    of Report-5 (uses_legacy_source / CUTOVER_DATE).

    Returns (total_hours, debug_rows)."""
    if uses_legacy_source(month_abbrev, calendar_year):
        return fetch_berth_hours_from_mis_vessel_master(month_abbrev, calendar_year)
    return fetch_berth_hours_from_ldud_header(month_abbrev, calendar_year)


def compute_berth_occupancy(month_abbrev: str, calendar_year: int):
    """
    Berth Occupancy % = [ Σ(castoff - alongside) in hours ]
                         / (days_in_month * 24)
                         / 2
                         * 100

    Written as separate steps (not combined into one divisor) to match
    the formula exactly as specified: sum first, divide by days*24,
    then divide by 2 (number of berths), then convert to a percentage.

    Includes "debug_rows": the per-vessel rows that were actually
    summed, so the vessels behind a given % can be checked directly.
    """
    month_str_to_idx(month_abbrev)  # validates month_abbrev

    is_legacy = uses_legacy_source(month_abbrev, calendar_year)
    source = "mis_vessel_master" if is_legacy else "ldud_header"

    logger.debug(
        "compute_berth_occupancy: month=%s year=%s -> source=%s (legacy=%s)",
        month_abbrev, calendar_year, source, is_legacy,
    )

    try:
        # Step 1: Σ(castoff - alongside) in hours, across matching rows
        total_hours, debug_rows = fetch_berth_hours(month_abbrev, calendar_year)
    except Exception as e:
        logger.exception(
            "compute_berth_occupancy: fetch_berth_hours raised for %s %s (source=%s)",
            month_abbrev, calendar_year, source,
        )
        return {
            "month": month_abbrev,
            "year": calendar_year,
            "source": source,
            "total_alongside_hours": None,
            "days_in_month": _days_in_month(month_abbrev, calendar_year),
            "num_berths": NUM_BERTHS,
            "occupancy_pct": None,
            "debug_rows": [],
            "error": f"{type(e).__name__}: {e}",
        }

    days = _days_in_month(month_abbrev, calendar_year)

    # Step 2: divide by (days_in_month * 24)
    hours_in_month = days * 24
    step2 = (total_hours / hours_in_month) if hours_in_month else None

    # Step 3: divide by 2 (number of berths)
    step3 = (step2 / NUM_BERTHS) if step2 is not None else None

    # Step 4: convert fraction to a percentage
    occupancy_pct = (step3 * 100) if step3 is not None else None

    result = {
        "month": month_abbrev,
        "year": calendar_year,
        "source": source,
        "total_alongside_hours": round(total_hours, 2),
        "days_in_month": days,
        "hours_in_month": hours_in_month,
        "num_berths": NUM_BERTHS,
        "occupancy_pct": round(occupancy_pct, 2) if occupancy_pct is not None else None,
        "debug_rows": debug_rows,
    }

    logger.debug(
        "compute_berth_occupancy: total_hours=%.2f / hours_in_month=%d = %.6f / "
        "num_berths=%d = %.6f * 100 = occupancy_pct=%s (%d debug_rows)",
        total_hours, hours_in_month, step2 or 0, NUM_BERTHS, step3 or 0,
        result["occupancy_pct"], len(debug_rows),
    )

    return result


# ---------------------------------------------------------------------
# COMMODITYWISE AVERAGE TURN AROUND TIME (BERTHING TO SAILING) & PARCEL SIZE
# ---------------------------------------------------------------------
# Formula (per user):
#   Av. Turn Around (days) = SUM(cast_off - alongside) / vessel_count
#   Av. Parcel size        = SUM(quantity) / vessel_count
# for the vessels of the given commodity in the given month.
#
# Same legacy/new cutover as the rest of the report:
#   Apr..Jun  -> mis_vessel_master (alongside/cast_off/quantity/commodity
#                all on the same row, so this is a single query)
#   Jul..Mar  -> ldud_header (alongside_datetime/cast_off_datetime,
#                commodity, vessel_name) for turnaround + vessel_count,
#                joined to lueu_parcel_log (quantity, commodity) for the
#                quantity total. There's no confirmed shared key between
#                the two new-system tables, so quantity is summed by
#                commodity+month independently and divided by the
#                vessel_count from ldud_header -- same "join by
#                commodity+month, not by row" approach already used for
#                fiscal-year totals elsewhere in this file. Flag if a
#                real join key (e.g. a shared vessel_id or a parcel_log
#                -> ldud_header FK) exists -- that would be more precise
#                than this month-level aggregate approach.
#
# ASSUMPTIONS THAT STILL NEED CONFIRMATION:
#   1. MIS_COMMODITY_COLUMN / LDUD_COMMODITY_COLUMN / PARCEL_LOG_COMMODITY_COLUMN
#      are GUESSES (see constants above).
#   2. "commodity" value match is exact-string (e.g. "Liquid", "Cement")
#      -- case-sensitivity / exact spelling not yet confirmed. Below this
#      is done as TRIM(...) ILIKE %(commodity)s -- case-insensitive,
#      whitespace-tolerant -- rather than a strict "=" match, since a
#      strict match is the more likely cause of silently-empty results
#      if the stored value has different casing/padding than what the
#      frontend sends.
#   3. Turnaround hours -> days via /24 (not, e.g., business days).
#   4. Rows with unparsable or out-of-order (castoff <= alongside) dates
#      are excluded from BOTH the turnaround sum and the vessel_count
#      denominator (a vessel with bad data shouldn't silently drag the
#      average down by counting in the denominator but not the numerator).
#   5. SHORT CLOSE: same subtraction rule applies to lueu_parcel_log
#      quantity here as everywhere else it's summed (see ASSUMPTION 5
#      at the top of the file). The averaging denominator (parcel_count)
#      still counts short-close rows -- see note in
#      fetch_commodity_turnaround_from_new_system's docstring below.

def fetch_commodity_turnaround_from_mis(commodity: str, month_abbrev: str, calendar_year: int):
    """Turnaround + vessel_count for one commodity, one month, from
    mis_vessel_master. Av. Parcel Size comes from mis_history (see
    module note above).

    COMMODITY FILTER — CHANGED (30-Jul-2026): mis_vessel_master has no
    column that literally contains the word "Liquid"/"Cement". Checked
    cargo, category, category1, new_cat directly against Jul-25/Jun-26
    data -- all four hold either specific cargo names ("Base Oil", "CPO",
    "Acetic Acid") or Liquid sub-categories ("Other Liquid", "Edible
    Oil", "POL", "Chemical", "Ph.Acid"), never the bare label "Liquid",
    and no Cement rows appeared at all. So instead of a commodity-text
    match, this now filters by berth_no using LIQUID_BERTH_CODES_MIS --
    the same berth-code list already confirmed and used for berth
    occupancy (LB-03/LB-04) -- since that's the actual way "Liquid" is
    identified in this table. This function is therefore only meaningful
    for commodity="Liquid" for now; fetch_commodity_turnaround() already
    short-circuits "Cement" requests to None before this is ever called.

    mis_history.cargo_type, by contrast, DOES literally hold "LIQUID"
    (confirmed via SELECT DISTINCT), so that filter is left as an ILIKE
    commodity match, unchanged.

    NOTE: mis_history has no short_close-style column and isn't part of
    the lueu_parcel_log short-close change -- this function's parcel-size
    figure is unaffected by ASSUMPTION (5).
    """
    yy = str(calendar_year)[-2:]
    month_str = f"{month_abbrev}-{yy}"

    logger.debug(
        "fetch_commodity_turnaround_from_mis: commodity=%r month=%s "
        "-> mis_vessel_master filtered by berth_no IN %s (not by a commodity "
        "column -- see docstring), mis_history filtered by cargo_type ILIKE commodity",
        commodity, month_str, LIQUID_BERTH_CODES_MIS,
    )

    conn = get_db()
    try:
        cur = get_cursor(conn)

        # ---------------- Turnaround + vessel_count (mis_vessel_master) ----------------
        cur.execute(
            f"""
            SELECT
                vessel_name,
                {MIS_ALONGSIDE_COLUMN} AS alongside_raw,
                {MIS_CASTOFF_COLUMN} AS castoff_raw
            FROM mis_vessel_master
            WHERE month = %(month_str)s
              AND {MIS_ALONGSIDE_COLUMN} IS NOT NULL
              AND {MIS_CASTOFF_COLUMN} IS NOT NULL
              AND berth_no = ANY(%(berth_codes)s)
            """,
            {"month_str": month_str, "berth_codes": LIQUID_BERTH_CODES_MIS},
        )
        turnaround_rows = cur.fetchall()

        # ---------------- Av. Parcel Size (mis_history) ----------------
        cur.execute(
            f"""
            SELECT {MIS_HISTORY_QUANTITY_COLUMN} AS quantity
            FROM mis_history
            WHERE TRIM({MIS_HISTORY_MONTH_COLUMN}) = %(month_str)s
              AND TRIM({MIS_HISTORY_COMMODITY_COLUMN}) ILIKE %(commodity)s
              AND {MIS_HISTORY_QUANTITY_COLUMN} IS NOT NULL
            """,
            {"month_str": month_str, "commodity": commodity.strip()},
        )
        parcel_rows = cur.fetchall()
    except Exception:
        logger.exception(
            "fetch_commodity_turnaround_from_mis: query FAILED for commodity=%r "
            "month=%s (check LIQUID_BERTH_CODES_MIS / MIS_HISTORY_* are correct)",
            commodity, month_str,
        )
        raise
    finally:
        conn.close()

    logger.debug(
        "fetch_commodity_turnaround_from_mis: commodity=%r month=%s -> "
        "%d mis_vessel_master row(s), %d mis_history row(s)",
        commodity, month_str, len(turnaround_rows), len(parcel_rows),
    )

    avg_turnaround_days, vessel_count, debug_rows = _sum_turnaround_hours(turnaround_rows)
    avg_turnaround_days = (
        round(avg_turnaround_days / 24.0 / vessel_count, 2)
        if vessel_count else None
    )

    total_parcel_qty = sum(float(r["quantity"] or 0) for r in parcel_rows)
    parcel_count = len(parcel_rows)
    avg_parcel_size = round(total_parcel_qty / parcel_count, 2) if parcel_count else None

    if DEBUG_BERTH_OCCUPANCY:
        logger.debug(
            "fetch_commodity_turnaround_from_mis: commodity=%r month=%s -> "
            "avg_turnaround_days=%s vessel_count=%d avg_parcel_size=%s "
            "(from %d mis_history rows)",
            commodity, month_str, avg_turnaround_days, vessel_count,
            avg_parcel_size, parcel_count,
        )

    return avg_turnaround_days, avg_parcel_size, vessel_count, debug_rows


def fetch_commodity_turnaround_from_new_system(commodity: str, month_abbrev: str, calendar_year: int):
    month_num = MONTH_NUM[month_abbrev]

    conn = get_db()
    try:
        cur = get_cursor(conn)

        # ---------------- Turnaround (+ vessel_count) ----------------
        # FIXED: query ldud_header directly, no join through
        # lueu_parcel_log/ldud_parcel_ops. That join was fanning out
        # one row per parcel for vessels with multiple parcels, which
        # double/triple-counted their hours and vessel_count. This now
        # matches the confirmed SQL exactly (7 vessels, 289.02 hrs,
        # 1.72 days for Jul-26).
        cur.execute("""
            SELECT
                vessel_name,
                alongside_datetime AS alongside_raw,
                cast_off_datetime AS castoff_raw
            FROM ldud_header
            WHERE COALESCE(is_deleted, false) = false
              AND NULLIF(TRIM(alongside_datetime), '') IS NOT NULL
              AND NULLIF(TRIM(cast_off_datetime), '') IS NOT NULL
        """)
        turnaround_rows = cur.fetchall()

        # ---------------- Parcel Size (unchanged — already a
        # separate query, not affected by the join bug above) --------
        cur.execute(
            """
            SELECT
                entry_date,
                quantity,
                is_shortclose AS short_close
            FROM lueu_parcel_log
            WHERE COALESCE(is_deleted, false) = false
              AND COALESCE(is_shortclose, false) = false
              AND quantity IS NOT NULL
            """
        )
        parcel_rows = cur.fetchall()

    except Exception:
        logger.exception("fetch_commodity_turnaround_from_new_system failed")
        raise
    finally:
        conn.close()

    # ---------- Filter turnaround by CAST-OFF month/year ----------
    matched_turnaround_rows = []
    for r in turnaround_rows:
        castoff = _parse_text_datetime(r["castoff_raw"])
        if not castoff:
            continue
        if castoff.year == calendar_year and castoff.month == month_num:
            matched_turnaround_rows.append(r)

    avg_turnaround_days, vessel_count, debug_rows = _sum_turnaround_hours(matched_turnaround_rows)
    avg_turnaround_days = (
        round(avg_turnaround_days / 24 / vessel_count, 2)
        if vessel_count else None
    )

    # ---------- Parcel Size (unchanged) ----------
    total_qty = 0
    for r in parcel_rows:
        d = _parse_text_date(r["entry_date"])
        if not d or d.year != calendar_year or d.month != month_num:
            continue
        total_qty += float(r["quantity"] or 0)

    avg_parcel_size = round(total_qty / vessel_count, 2) if vessel_count else None

    logger.debug(
        "Commodity=%s Month=%s-%s Qty=%s Vessel=%s AvgTA=%s AvgParcel=%s",
        commodity, month_abbrev, calendar_year, total_qty, vessel_count,
        avg_turnaround_days, avg_parcel_size,
    )

    return avg_turnaround_days, avg_parcel_size, vessel_count, debug_rows

def _sum_turnaround_hours(rows):
    """
    Returns:
        total_hours,
        vessel_count,
        debug_rows
    """

    total_hours = 0.0
    vessel_count = 0
    debug_rows = []

    for r in rows:
        vessel_name = r.get("vessel_name")

        alongside = _parse_text_datetime(r.get("alongside_raw"))
        castoff = _parse_text_datetime(r.get("castoff_raw"))

        if alongside is None or castoff is None:
            logger.debug(
                "Skipping %s : Invalid date (Alongside=%s Castoff=%s)",
                vessel_name,
                r.get("alongside_raw"),
                r.get("castoff_raw"),
            )
            continue

        if castoff <= alongside:
            logger.debug(
                "Skipping %s : Castoff <= Alongside",
                vessel_name,
            )
            continue

        hours = (castoff - alongside).total_seconds() / 3600.0

        total_hours += hours
        vessel_count += 1

        debug_rows.append({
            "vessel_name": vessel_name,
            "alongside": alongside.strftime("%d-%b-%Y %H:%M"),
            "castoff": castoff.strftime("%d-%b-%Y %H:%M"),
            "hours": round(hours, 2),
        })

    logger.debug(
        "Total Hours = %.2f | Vessel Count = %d | Average Days = %.2f",
        total_hours,
        vessel_count,
        (total_hours / 24 / vessel_count) if vessel_count else 0,
    )

    return total_hours, vessel_count, debug_rows


def _summarize_commodity_rows(rows):
    """Shared helper for the mis_vessel_master path: rows carry
    vessel_name/alongside_raw/castoff_raw/quantity together, so
    turnaround, vessel_count, and parcel size can all come from the
    same row set (unlike the new-system path, which has to combine two
    tables). Returns (avg_turnaround_days, avg_parcel_size, vessel_count,
    debug_rows)."""
    total_hours = 0.0
    total_quantity = 0.0
    vessel_count = 0
    debug_rows = []

    for r in rows:
        vessel_name = r["vessel_name"]
        alongside = _parse_text_datetime(r["alongside_raw"])
        castoff = _parse_text_datetime(r["castoff_raw"])

        if not alongside or not castoff or castoff <= alongside:
            continue

        hours = (castoff - alongside).total_seconds() / 3600.0
        qty = float(r["quantity"] or 0)
        total_hours += hours
        total_quantity += qty
        vessel_count += 1
        debug_rows.append({
            "vessel_name": vessel_name,
            "alongside": alongside.isoformat(),
            "castoff": castoff.isoformat(),
            "hours": round(hours, 2),
            "quantity": qty,
        })

    avg_turnaround_days = (total_hours / 24.0 / vessel_count) if vessel_count else None
    avg_parcel_size = (total_quantity / vessel_count) if vessel_count else None

    return (
        round(avg_turnaround_days, 2) if avg_turnaround_days is not None else None,
        round(avg_parcel_size, 2) if avg_parcel_size is not None else None,
        vessel_count,
        debug_rows,
    )


def fetch_commodity_turnaround(commodity: str, month_abbrev: str, calendar_year: int):

    # Don't display anything for Cement
    if commodity.lower() == "cement":
        return {
            "commodity": commodity,
            "month": month_abbrev,
            "year": calendar_year,
            "source": "",
            "avg_turnaround_days": None,
            "avg_parcel_size": None,
            "vessel_count": 0,
            "debug_rows": []
        }

    is_legacy = uses_legacy_source(month_abbrev, calendar_year)
    source = "mis_vessel_master" if is_legacy else "ldud_header+lueu_parcel_log"

    if is_legacy:
        avg_days, avg_parcel, vessel_count, debug_rows = fetch_commodity_turnaround_from_mis(
            commodity, month_abbrev, calendar_year
        )
    else:
        avg_days, avg_parcel, vessel_count, debug_rows = fetch_commodity_turnaround_from_new_system(
            commodity, month_abbrev, calendar_year
        )

    return {
        "commodity": commodity,
        "month": month_abbrev,
        "year": calendar_year,
        "source": source,
        "avg_turnaround_days": avg_days,
        "avg_parcel_size": avg_parcel,
        "vessel_count": vessel_count,
        "debug_rows": debug_rows,
    }


def _previous_month_abbrev_year(month_abbrev: str, calendar_year: int):
    """Prior calendar month (not prior year) — e.g. Jul-2026 -> Jun-2026,
    Jan-2026 -> Dec-2025."""
    month_num = MONTH_NUM[month_abbrev]
    if month_num == 1:
        return NUM_TO_MONTH_ABBREV[12], calendar_year - 1
    return NUM_TO_MONTH_ABBREV[month_num - 1], calendar_year


def compute_commodity_turnaround_report(commodity: str, month_abbrev: str, calendar_year: int):
    """Current month vs the PRIOR CALENDAR MONTH (not prior year) —
    matches the confirmed SQL: Jul-26 current vs Jun-26 previous."""
    month_str_to_idx(month_abbrev)  # validates month_abbrev

    current = fetch_commodity_turnaround(commodity, month_abbrev, calendar_year)

    prev_month_abbrev, prev_calendar_year = _previous_month_abbrev_year(month_abbrev, calendar_year)
    previous = fetch_commodity_turnaround(commodity, prev_month_abbrev, prev_calendar_year)

    return {
        "commodity": commodity,
        "month": month_abbrev,
        "year": calendar_year,
        "current": current,
        "previous": previous,
    }


# ---------------------------------------------------------------------
# EXPORT TO EXCEL — fills the fixed template, doesn't rebuild it
# ---------------------------------------------------------------------
# The template (REPORT5_TEMPLATE_PATH) is the exact workbook the report
# is expected to look like: same two sheets, same merges, same borders/
# fills/fonts/number-formats, and the same live formulas for every
# derived cell (row totals, increase-%, TOTAL LIQUID, etc.).
#
# Only the handful of cells below — the ones that already hold
# hand-typed numbers in that file rather than a formula — get
# overwritten per request. Everything else in the workbook (both
# sheets) is copied through untouched, so "Export Excel" always
# produces a file in the exact same format/columns/cells as the
# original, just re-computed for whichever month/year was requested.

def build_report5_export_workbook(month_abbrev: str, calendar_year: int):
    """Returns (BytesIO, filename) for the filled-in export workbook."""

    month_str_to_idx(month_abbrev)  # validates month_abbrev

    wb = Workbook()
    ws = wb.active
    ws.title = EXPORT_SUMMARY_SHEET

    # Column widths
    for col, width in EXPORT_SUM_COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Row heights
    for row, height in EXPORT_SUM_ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height

    # Merged cells
    for rng in EXPORT_SUM_MERGES:
        ws.merge_cells(rng)

    # Draw the complete layout
    _apply_export_layout(ws)

    # ---- Same data the /report, /berth-occupancy and
    # /commodity-turnaround endpoints already compute -- reused as-is
    # so the export always matches what's on screen. ----
    fin_year_start = fiscal_year_start_for(month_abbrev, calendar_year)
    report = compute_fiscal_year_comparison(fin_year_start, month_abbrev, calendar_year)
    occupancy = compute_berth_occupancy(month_abbrev, calendar_year)
    liquid_turn = compute_commodity_turnaround_report("Liquid", month_abbrev, calendar_year)
    cement_turn = compute_commodity_turnaround_report("Cement", month_abbrev, calendar_year)


    if EXPORT_SUMMARY_SHEET not in wb.sheetnames:
        raise ReportDataError(
            f"Export template is missing the '{EXPORT_SUMMARY_SHEET}' sheet."
        )
    ws = wb[EXPORT_SUMMARY_SHEET]

    full_month = MONTH_ABBREV_TO_FULL.get(month_abbrev, month_abbrev)
    curr_yy = str(calendar_year)[-2:]
    prev_yy = str(calendar_year - 1)[-2:]

    # ---- Title + Table 1 column headers (B3 merged B3:I3, F6, G6) ----
    ws["B3"] = f"Performance Report for the Month of {full_month}-{calendar_year}"
    ws["F6"] = f"{full_month}-{prev_yy}(Previous)"
    ws["G6"] = f"{full_month}-{curr_yy}(Current)"

    # ---- Table 1 : only the "Liquid (LB-03&LB-04)" row (row 9) is a
    # real input row -- Total Liquid (row 11), TOTAL LIQUID (row 22)
    # and Total (row 23) are existing SUM()/formula rows in the
    # template and recompute themselves from row 9 automatically. ----
    current = report["current"]
    previous = report["previous"]
    ws["D9"] = current.get("vessel_count")
    ws["E9"] = (
        occupancy["occupancy_pct"] / 100
        if occupancy.get("occupancy_pct") is not None else None
    )
    ws["F9"] = previous.get("quantity")
    ws["G9"] = current.get("quantity")
    # H9 ( ="(G9-F9)/F23" increase % ) is an existing formula -- left as-is.

    # ---- Table 2 : fiscal-year labels + the one real input row,
    # "LIQUID (JJLTPL)" (row 32). TOTAL LIQUID (row 33) and the grand
    # total (row 37) are existing formulas in the template. ----
    fy_prev_start = fin_year_start - 1
    fy_curr_start = fin_year_start
    ws["C27"] = f"{fy_prev_start}-{str(fy_prev_start + 1)[-2:]}"
    ws["F27"] = f"{fy_curr_start}-{str(fy_curr_start + 1)[-2:]}"
    ws["C28"] = f"APR'{fy_prev_start} \u2013 MAR-{fy_prev_start + 1}"
    ws["F28"] = f"APR'{fy_curr_start} \u2013 MAR-{fy_curr_start + 1}"
    ws["D32"] = report.get("previous_jjltpl")
    ws["G32"] = report.get("current_jjltpl")
    # I32 ( ="(G32-D32)*100/D32" ) is an existing formula -- left as-is.

    # ---- Table 3 : Liquid (rows 43-44) + Cement (rows 45-46). Cement
    # stays blank (avg_turnaround_days/avg_parcel_size are always None
    # for Cement, per fetch_commodity_turnaround()'s current
    # "don't display anything for Cement" behaviour), matching the
    # template's own blank Cement cells. ----
    def _fill_turnaround_rows(current_row, previous_row, turnaround_report):
        cur = turnaround_report["current"]
        prev = turnaround_report["previous"]
        ws[f"C{current_row}"] = cur.get("avg_turnaround_days")
        ws[f"D{current_row}"] = cur.get("avg_parcel_size")
        ws[f"C{previous_row}"] = prev.get("avg_turnaround_days")
        ws[f"D{previous_row}"] = prev.get("avg_parcel_size")

    _fill_turnaround_rows(43, 44, liquid_turn)
    _fill_turnaround_rows(45, 46, cement_turn)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Report5_{month_abbrev}-{calendar_year}.xlsx"
    return buf, filename


# ---------------------------------------------------------------------
# ROUTES
# ---------------------------------------------------------------------

@bp.route("/module/RP01/report5/")
@login_required
def report5_index():
    return render_template("report5/report5.html")


@bp.route("/api/module/RP01/report5/meta")
@login_required
def report5_meta():
    try:
        today = date.today()
        current_month_abbrev = NUM_TO_MONTH_ABBREV[today.month]

        # Current year is now driven by the latest fin_year present in
        # mis_vessel_master rather than the server's local clock, per
        # user instruction. current_month still comes from today's date -
        # fin_year doesn't imply a specific month.
        fin_year_start = fetch_latest_fin_year_start()
        current_year = calendar_year_for_month(fin_year_start, current_month_abbrev)

        return jsonify({
            "months": [{"abbrev": m, "label": m} for m in MONTH_NAMES],
            "current_month": current_month_abbrev,
            "current_year": current_year,
            "fin_year_start": fin_year_start,
        })
    except Exception:
        import traceback
        traceback.print_exc()
        raise


@bp.route("/api/module/RP01/report5/report")
@login_required
def report5_report():
    try:
        month = request.args.get("month")
        if not month:
            return jsonify({"error": "Missing required parameter: month"}), 400

        year_param = request.args.get("year")
        if year_param:
            try:
                calendar_year = int(year_param)
            except ValueError:
                return jsonify({"error": f"Invalid 'year' parameter: {year_param}"}), 400
        else:
            # No explicit year - default to the current fiscal year as
            # recorded in mis_vessel_master.fin_year, converted to the
            # calendar year that matches the requested month.
            fin_year_start = fetch_latest_fin_year_start()
            calendar_year = calendar_year_for_month(fin_year_start, month)

        fin_year_start = fiscal_year_start_for(month, calendar_year)
        result = compute_fiscal_year_comparison(fin_year_start, month, calendar_year)

        # The frontend (report5.html) only calls this single /report
        # endpoint, not /berth-occupancy separately, so occupancy has to
        # be folded in here or the UI table cell stays blank forever.
        occupancy = compute_berth_occupancy(month, calendar_year)
        result["occupancy_pct"] = occupancy["occupancy_pct"]
        result["occupancy_detail"] = occupancy  # full breakdown, for debugging/export

        # Sanity check: confirm the key we just set is actually the same
        # value about to go out over the wire. If this log ever shows
        # occupancy_pct missing or None while occupancy_detail shows a
        # real number, something overwrote `result` between here and
        # jsonify() - which would point at a bug in this route, not in
        # compute_berth_occupancy.
        logger.debug(
            "report5_report: FINAL response keys=%s | occupancy_pct=%r | occupancy_detail debug_rows=%d",
            sorted(result.keys()), result.get("occupancy_pct"),
            len(occupancy.get("debug_rows", [])),
        )
        print("DEBUG report5 result:", result)  # TEMP — remove once fetch works

        return jsonify(result)

    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report5/commodity-turnaround")
@login_required
def report5_commodity_turnaround():
    """Backs the 'Commoditywise Average Turn Around Time (Berthing to
    Sailing) & Parcel Size' table: current + previous avg_turnaround_days
    and avg_parcel_size for one commodity (e.g. 'Liquid', 'Cement')."""
    try:
        commodity = request.args.get("commodity")
        if not commodity:
            return jsonify({"error": "Missing required parameter: commodity"}), 400

        month = request.args.get("month")
        if not month:
            return jsonify({"error": "Missing required parameter: month"}), 400

        year_param = request.args.get("year")
        if year_param:
            try:
                calendar_year = int(year_param)
            except ValueError:
                return jsonify({"error": f"Invalid 'year' parameter: {year_param}"}), 400
        else:
            fin_year_start = fetch_latest_fin_year_start()
            calendar_year = calendar_year_for_month(fin_year_start, month)

        result = compute_commodity_turnaround_report(commodity, month, calendar_year)
        return jsonify(result)

    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report5/berth-occupancy")
@login_required
def report5_berth_occupancy():
    try:
        month = request.args.get("month")
        if not month:
            return jsonify({"error": "Missing required parameter: month"}), 400

        year_param = request.args.get("year")
        if year_param:
            try:
                calendar_year = int(year_param)
            except ValueError:
                return jsonify({"error": f"Invalid 'year' parameter: {year_param}"}), 400
        else:
            fin_year_start = fetch_latest_fin_year_start()
            calendar_year = calendar_year_for_month(fin_year_start, month)

        result = compute_berth_occupancy(month, calendar_year)
        return jsonify(result)

    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/report5/export", methods=["GET"])
@login_required
def report5_export():
    try:
        month = request.args.get("month")
        if not month:
            return jsonify({"error": "Missing required parameter: month"}), 400

        year_param = request.args.get("year")

        if year_param:
            try:
                calendar_year = int(year_param)
            except ValueError:
                return jsonify({"error": "Invalid year parameter"}), 400
        else:
            fin_year_start = fetch_latest_fin_year_start()
            calendar_year = calendar_year_for_month(fin_year_start, month)

        buf, filename = build_report5_export_workbook(
            month,
            calendar_year
        )

        buf.seek(0)

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
"""
Budget Vs. Actual – Quantity Handled Report (RP01 Module)

Multi-month side-by-side pivot grid report from April up to the current month (or all months for historical years).

Commodities:
  1. Edible Oil
  2. Phosphoric Acid/Lube/Chemical
  3. POL - White
  4. POL - Black
  Total

Columns per Month:
  - Budget
  - Actual JNPT
  - Actual JSW

Rules & Cutoffs:
  - JNPT Actual: Filtered by vessel Cast-off Time.
  - JSW Historical Actual: Filtered by `month_jsw` in mis_history.
  - JSW Live Actual: Filtered by 07:00 AM monthly operational window.
"""

import io
import time
import traceback
from datetime import datetime
from functools import wraps

import pandas as pd
from flask import jsonify, render_template, request, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import get_cursor, get_db
from .. import bp

CANONICAL_COMMODITIES = [
    "Edible Oil",
    "Phosphoric Acid/Lube/Chemical",
    "POL - White",
    "POL - Black",
]

MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated


def classify_commodity(cargo_name, cargo_sub_cat, cargo_sub_cat_2):
    """
    Classifies any cargo record into one of the 4 standard particular rows:
      1. Edible Oil
      2. Phosphoric Acid/Lube/Chemical
      3. POL - White
      4. POL - Black
    """
    cn = (cargo_name or "").strip().upper()
    sub = (cargo_sub_cat or "").strip().upper()
    sub2 = (cargo_sub_cat_2 or "").strip().upper()

    if sub2 == "EDIBLE OIL" or sub == "EDIBLE OIL" or ("OIL" in cn and ("PALM" in cn or "SUNFLOWER" in cn or "SOYABEAN" in cn or "EDIBLE" in cn)):
        return "Edible Oil"

    if sub2 == "POL" or sub == "POL" or "POL" in cn or "POL - WHITE" in sub2 or "POL - BLACK" in sub2 or "POL - WHITE" in sub or "POL - BLACK" in sub:
        if any(kw in cn or kw in sub2 or kw in sub for kw in ["BLACK", "FURNACE", "CARBON", "CARBAN", "FO", "CBFS"]):
            return "POL - Black"
        else:
            return "POL - White"

    return "Phosphoric Acid/Lube/Chemical"





def get_current_fy_and_month_idx():
    now = datetime.now()
    y, m = now.year, now.month
    if m >= 4:
        cur_fy = f"{y}-{str((y + 1) % 100).zfill(2)}"
        cur_m_idx = m - 4
    else:
        cur_fy = f"{y - 1}-{str(y % 100).zfill(2)}"
        cur_m_idx = m + 8
    return cur_fy, cur_m_idx


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def month_label_from_idx(fin_year: str, month_idx: int) -> str:
    start_y = fy_start_year(fin_year)
    mn = MONTH_NAMES[month_idx]
    yy = start_y if month_idx < 9 else start_y + 1
    return f"{mn}-{str(yy % 100).zfill(2)}"


def load_all_actuals_and_budgets(fin_year: str, detailed: bool = False):
    """
    Returns (jnpt_data, jsw_data, budget_data)
    where each is a dict: month_idx -> { commodity_name -> float_qty }
    If detailed is False: Groups into the 4 canonical commodities (Edible Oil, Phosphoric Acid/Lube/Chemical, POL - White, POL - Black).
    If detailed is True: Keeps all dynamic cargo categories/names from the database without grouping.
    """
    conn = get_db()
    
    jnpt_data = {m: {} for m in range(12)}
    jsw_data = {m: {} for m in range(12)}
    budget_data = {m: {} for m in range(12)}

    if not detailed:
        for m in range(12):
            for c in CANONICAL_COMMODITIES:
                budget_data[m][c] = 0.0
                jnpt_data[m][c] = 0.0
                jsw_data[m][c] = 0.0

    try:
        cur = get_cursor(conn)

        # 1. Load Budgets exclusively from financial_year_targets table (target field only)
        cur.execute("""
            SELECT targets
            FROM financial_year_targets
            WHERE financial_year = %s;
        """, (fin_year,))
        fy_row = cur.fetchone()

        if fy_row and fy_row["targets"]:
            targets_raw = fy_row["targets"]
            if isinstance(targets_raw, str):
                import json
                targets_raw = json.loads(targets_raw)
            
            target_items = targets_raw.get("targets", []) if isinstance(targets_raw, dict) else []
            for t_item in target_items:
                name = (t_item.get("name") or "").strip()
                monthly_data = t_item.get("monthly_data", [])
                cat = name if detailed else classify_commodity(name, name, name)
                for md in monthly_data:
                    m_name = (md.get("month") or "").strip()
                    val = float(md.get("target") or 0.0)
                    if m_name in MONTH_NAMES:
                        m_idx = MONTH_NAMES.index(m_name)
                        if 0 <= m_idx < 12:
                            budget_data[m_idx][cat] = budget_data[m_idx].get(cat, 0.0) + val

        # 2. Load Historical MIS Actuals
        cur.execute("""
            SELECT month_jnpt, month_jsw, cargo_name, cargo_sub_category, cargo_sub_category_2, quantity
            FROM mis_history
            WHERE fin_year = %s;
        """, (fin_year,))
        mis_rows = cur.fetchall()

        for r in mis_rows:
            cn = (r["cargo_name"] or "").strip()
            sub = (r["cargo_sub_category"] or "").strip()
            sub2 = (r["cargo_sub_category_2"] or "").strip()
            qty = float(r["quantity"] or 0.0)

            if detailed:
                cat = sub2 or sub or cn or "Other Cargo"
            else:
                cat = classify_commodity(cn, sub, sub2)

            mjnpt = (r["month_jnpt"] or "").strip()
            mjsw = (r["month_jsw"] or "").strip()

            for m_idx in range(12):
                mlabel = month_label_from_idx(fin_year, m_idx)
                if mjnpt == mlabel:
                    jnpt_data[m_idx][cat] = jnpt_data[m_idx].get(cat, 0.0) + qty
                if mjsw == mlabel:
                    jsw_data[m_idx][cat] = jsw_data[m_idx].get(cat, 0.0) + qty

        # 3. Live Pipeline Data (if mis_history has no records for a month)
        for m_idx in range(12):
            mlabel = month_label_from_idx(fin_year, m_idx)
            cur.execute("""
                SELECT COUNT(*) AS cnt FROM mis_history WHERE fin_year = %s AND (month_jnpt = %s OR month_jsw = %s);
            """, (fin_year, mlabel, mlabel))
            if cur.fetchone()["cnt"] == 0:
                cur.execute("""
                    SELECT 
                        po.cargo_name,
                        SUM(COALESCE(l.quantity, 0)) AS total_qty
                    FROM lueu_parcel_log l
                    JOIN ldud_parcel_ops po ON po.id = l.parcel_op_id
                    JOIN ldud_header ld ON ld.id = po.ldud_id
                    WHERE ld.cast_off_datetime IS NOT NULL
                      AND COALESCE(l.is_deleted, false) = false
                      AND COALESCE(l.is_shortclose, false) = false
                    GROUP BY po.cargo_name;
                """)
                for r in cur.fetchall():
                    cn = (r["cargo_name"] or "").strip()
                    qty = float(r["total_qty"] or 0.0)
                    cat = cn if detailed else classify_commodity(cn, "", "")
                    jnpt_data[m_idx][cat] = jnpt_data[m_idx].get(cat, 0.0) + qty
                    jsw_data[m_idx][cat] = jsw_data[m_idx].get(cat, 0.0) + qty

    finally:
        conn.close()

    return jnpt_data, jsw_data, budget_data


@bp.route("/module/RP01/report-budget/")
@login_required
def report_budget_index():
    return render_template("report_budget/report_budget.html")


@bp.route("/module/RP01/report-budget/targets")
@login_required
def report_budget_targets_index():
    return render_template("report_budget/report_targets.html")


@bp.route("/api/module/RP01/report-budget/meta")
@login_required
def report_budget_api_meta():
    try:
        cur_fy, cur_m_idx = get_current_fy_and_month_idx()
        conn = get_db()
        fy_list = []
        try:
            cur = get_cursor(conn)
            cur.execute("""
                SELECT DISTINCT fin_year FROM mis_history WHERE fin_year IS NOT NULL AND TRIM(fin_year) != ''
                UNION
                SELECT DISTINCT financial_year AS fin_year FROM financial_year_targets WHERE financial_year IS NOT NULL AND TRIM(financial_year) != '';
            """)
            for r in cur.fetchall():
                if r["fin_year"]:
                    fy_list.append(r["fin_year"].strip())
        finally:
            conn.close()

        if cur_fy not in fy_list:
            fy_list.append(cur_fy)
        years = sorted(list(set(fy_list)))

        return jsonify({
            "years": years,
            "current_fy": cur_fy,
            "current_month_idx": cur_m_idx,
            "canonical_commodities": CANONICAL_COMMODITIES
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to load meta info: {e}"}), 500


@bp.route("/api/module/RP01/report-budget/report")
@login_required
def report_budget_api_report():
    try:
        cur_fy, cur_m_idx = get_current_fy_and_month_idx()
        fin_year = request.args.get("fin_year", cur_fy)
        month_idx_arg = request.args.get("month_idx")
        commodity = (request.args.get("commodity") or "ALL").strip()
        port = (request.args.get("port") or "ALL").strip().upper()
        detailed_arg = (request.args.get("detailed") or "false").strip().lower()
        detailed = (detailed_arg == "true" or detailed_arg == "1")

        if month_idx_arg is not None and month_idx_arg != "":
            target_m_idx = int(month_idx_arg)
            max_m_idx = min(target_m_idx, cur_m_idx if fin_year == cur_fy else 11)
        else:
            max_m_idx = cur_m_idx if fin_year == cur_fy else 11

        visible_months = [
            {"idx": idx, "label": month_label_from_idx(fin_year, idx)}
            for idx in range(max_m_idx + 1)
        ]

        jnpt_data, jsw_data, budget_data = load_all_actuals_and_budgets(fin_year, detailed=detailed)

        all_comms = set()
        for m_idx in range(12):
            all_comms.update(budget_data[m_idx].keys())
            all_comms.update(jnpt_data[m_idx].keys())
            all_comms.update(jsw_data[m_idx].keys())

        if not detailed:
            all_comms_list = CANONICAL_COMMODITIES
        else:
            all_comms_list = sorted(list(all_comms))

        if commodity != "ALL":
            selected_commodities = [c for c in all_comms_list if c.lower() == commodity.lower()]
            if not selected_commodities:
                selected_commodities = [commodity]
        else:
            selected_commodities = all_comms_list

        rows = []
        month_totals = {
            m["idx"]: {"budget": 0.0, "jnpt_actual": 0.0, "jsw_actual": 0.0}
            for m in visible_months
        }
        grand_total = {"budget": 0.0, "jnpt_actual": 0.0, "jsw_actual": 0.0}

        for comm in selected_commodities:
            row_months = {}
            row_tot_budget = 0.0
            row_tot_jnpt = 0.0
            row_tot_jsw = 0.0

            for m in visible_months:
                idx = m["idx"]
                b_qty = budget_data[idx].get(comm, 0.0)
                jnpt_q = jnpt_data[idx].get(comm, 0.0) if port in ("ALL", "JNPT") else 0.0
                jsw_q = jsw_data[idx].get(comm, 0.0) if port in ("ALL", "JSW") else 0.0

                row_months[idx] = {
                    "budget": b_qty,
                    "jnpt_actual": jnpt_q,
                    "jsw_actual": jsw_q,
                }

                row_tot_budget += b_qty
                row_tot_jnpt += jnpt_q
                row_tot_jsw += jsw_q

                month_totals[idx]["budget"] += b_qty
                month_totals[idx]["jnpt_actual"] += jnpt_q
                month_totals[idx]["jsw_actual"] += jsw_q

            grand_total["budget"] += row_tot_budget
            grand_total["jnpt_actual"] += row_tot_jnpt
            grand_total["jsw_actual"] += row_tot_jsw

            rows.append({
                "particular": comm,
                "months": row_months,
                "total": {
                    "budget": row_tot_budget,
                    "jnpt_actual": row_tot_jnpt,
                    "jsw_actual": row_tot_jsw,
                }
            })

        return jsonify({
            "fin_year": fin_year,
            "port": port,
            "commodity": commodity,
            "visible_months": visible_months,
            "rows": rows,
            "month_totals": month_totals,
            "grand_total": grand_total
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate multi-month budget report: {e}"}), 500


@bp.route("/api/module/RP01/report-budget/budget-save", methods=["POST"])
@login_required
def report_budget_api_save():
    try:
        data = request.get_json(force=True) or {}
        fin_year = data.get("fin_year")
        month_idx = data.get("month_idx")
        commodity = data.get("commodity")
        budget_quantity = float(data.get("budget_quantity", 0.0))

        if not fin_year or month_idx is None or not commodity:
            return jsonify({"error": "Missing required fields"}), 400

        conn = get_db()
        try:
            cur = get_cursor(conn)
            cur.execute("""
                INSERT INTO cargo_budget (fin_year, month_idx, cargo_sub_category_2, budget_quantity)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (fin_year, month_idx, cargo_sub_category_2)
                DO UPDATE SET budget_quantity = EXCLUDED.budget_quantity;
            """, (fin_year, int(month_idx), commodity.strip(), budget_quantity))
            conn.commit()
        finally:
            conn.close()

        return jsonify({"success": True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to save budget quantity: {e}"}), 500


@bp.route("/api/module/RP01/report-budget/export")
@login_required
def report_budget_api_export():
    """Generates Excel export matching the reference multi-month layout."""
    try:
        cur_fy, cur_m_idx = get_current_fy_and_month_idx()
        fin_year = request.args.get("fin_year", cur_fy)
        month_idx_arg = request.args.get("month_idx")
        commodity = (request.args.get("commodity") or "ALL").strip()
        port = (request.args.get("port") or "ALL").strip().upper()

        if month_idx_arg is not None and month_idx_arg != "":
            target_m_idx = int(month_idx_arg)
            max_m_idx = min(target_m_idx, cur_m_idx if fin_year == cur_fy else 11)
        else:
            max_m_idx = cur_m_idx if fin_year == cur_fy else 11

        visible_months = [
            {"idx": idx, "label": month_label_from_idx(fin_year, idx)}
            for idx in range(max_m_idx + 1)
        ]

        detailed_arg = (request.args.get("detailed") or "false").strip().lower()
        detailed = (detailed_arg == "true" or detailed_arg == "1")

        jnpt_data, jsw_data, budget_data = load_all_actuals_and_budgets(fin_year, detailed=detailed)

        all_comms = set()
        for m_idx in range(12):
            all_comms.update(budget_data[m_idx].keys())
            all_comms.update(jnpt_data[m_idx].keys())
            all_comms.update(jsw_data[m_idx].keys())

        if not detailed:
            all_comms_list = CANONICAL_COMMODITIES
        else:
            all_comms_list = sorted(list(all_comms))

        if commodity != "ALL":
            selected_commodities = [c for c in all_comms_list if c.lower() == commodity.lower()]
            if not selected_commodities:
                selected_commodities = [commodity]
        else:
            selected_commodities = all_comms_list

        show_jnpt = port in ("ALL", "JNPT")
        show_jsw = port in ("ALL", "JSW")
        act_cols = (1 if show_jnpt else 0) + (1 if show_jsw else 0)
        m_cols = 1 + act_cols

        wb = Workbook()
        ws = wb.active
        ws.title = "Budget vs Actual"

        hdr_font = Font(name="Arial", size=10, bold=True, color="000000")
        hdr_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
        
        tot_font = Font(name="Arial", size=10, bold=True, color="000000")
        border_thin = Side(style="thin", color="000000")
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        # Row 1: Particular + Month Groupings + Total Grouping
        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
        cell = ws.cell(row=1, column=1, value="Particular")
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

        col_pos = 2
        for m in visible_months:
            ws.merge_cells(start_row=1, start_column=col_pos, end_row=1, end_column=col_pos + m_cols - 1)
            m_cell = ws.cell(row=1, column=col_pos, value=m["label"])
            m_cell.font = hdr_font
            m_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=2, column=col_pos, value="Budget").font = hdr_font
            ws.cell(row=2, column=col_pos).alignment = Alignment(horizontal="center", vertical="center")

            curr_act_col = col_pos + 1
            if act_cols > 1:
                ws.merge_cells(start_row=2, start_column=curr_act_col, end_row=2, end_column=curr_act_col + act_cols - 1)
                ws.cell(row=2, column=curr_act_col, value="Actual").font = hdr_font
                ws.cell(row=2, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")

                if show_jnpt:
                    ws.cell(row=3, column=curr_act_col, value="JNPT").font = hdr_font
                    ws.cell(row=3, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")
                    curr_act_col += 1
                if show_jsw:
                    ws.cell(row=3, column=curr_act_col, value="JSW").font = hdr_font
                    ws.cell(row=3, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")
            else:
                ws.cell(row=2, column=curr_act_col, value=f"Actual ({port})").font = hdr_font
                ws.cell(row=2, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")

            for r in range(1, 4):
                for c in range(col_pos, col_pos + m_cols):
                    ws.cell(row=r, column=c).border = cell_border
                    ws.cell(row=r, column=c).fill = hdr_fill

            col_pos += m_cols

        # Grand Total Columns
        ws.merge_cells(start_row=1, start_column=col_pos, end_row=1, end_column=col_pos + m_cols - 1)
        gt_cell = ws.cell(row=1, column=col_pos, value=f"Total ({fin_year})")
        gt_cell.font = hdr_font
        gt_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=2, column=col_pos, value="Budget").font = hdr_font
        ws.cell(row=2, column=col_pos).alignment = Alignment(horizontal="center", vertical="center")

        curr_act_col = col_pos + 1
        if act_cols > 1:
            ws.merge_cells(start_row=2, start_column=curr_act_col, end_row=2, end_column=curr_act_col + act_cols - 1)
            ws.cell(row=2, column=curr_act_col, value="Actual").font = hdr_font
            ws.cell(row=2, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")

            if show_jnpt:
                ws.cell(row=3, column=curr_act_col, value="JNPT").font = hdr_font
                ws.cell(row=3, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")
                curr_act_col += 1
            if show_jsw:
                ws.cell(row=3, column=curr_act_col, value="JSW").font = hdr_font
                ws.cell(row=3, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws.cell(row=2, column=curr_act_col, value=f"Actual ({port})").font = hdr_font
            ws.cell(row=2, column=curr_act_col).alignment = Alignment(horizontal="center", vertical="center")

        for r in range(1, 4):
            for c in range(col_pos, col_pos + m_cols):
                ws.cell(row=r, column=c).border = cell_border
                ws.cell(row=r, column=c).fill = hdr_fill

        # Data Rows
        curr_row = 4
        month_totals = {m["idx"]: {"b": 0.0, "jnpt": 0.0, "jsw": 0.0} for m in visible_months}
        gt_totals = {"b": 0.0, "jnpt": 0.0, "jsw": 0.0}

        for comm in selected_commodities:
            ws.cell(row=curr_row, column=1, value=comm).font = Font(name="Arial", size=10)
            ws.cell(row=curr_row, column=1).border = cell_border

            col_pos = 2
            row_tot_b, row_tot_jnpt, row_tot_jsw = 0.0, 0.0, 0.0

            for m in visible_months:
                idx = m["idx"]
                b_qty = budget_data[idx].get(comm, 0.0)
                jnpt_q = jnpt_data[idx].get(comm, 0.0) if show_jnpt else 0.0
                jsw_q = jsw_data[idx].get(comm, 0.0) if show_jsw else 0.0

                ws.cell(row=curr_row, column=col_pos, value=b_qty).number_format = "#,##0"
                c_idx = col_pos + 1
                if show_jnpt:
                    ws.cell(row=curr_row, column=c_idx, value=jnpt_q).number_format = "#,##0"
                    c_idx += 1
                if show_jsw:
                    ws.cell(row=curr_row, column=c_idx, value=jsw_q).number_format = "#,##0"
                    c_idx += 1

                for c in range(col_pos, col_pos + m_cols):
                    cell = ws.cell(row=curr_row, column=c)
                    cell.border = cell_border
                    cell.alignment = Alignment(horizontal="right")

                row_tot_b += b_qty
                row_tot_jnpt += jnpt_q
                row_tot_jsw += jsw_q

                month_totals[idx]["b"] += b_qty
                month_totals[idx]["jnpt"] += jnpt_q
                month_totals[idx]["jsw"] += jsw_q

                col_pos += m_cols

            ws.cell(row=curr_row, column=col_pos, value=row_tot_b).number_format = "#,##0"
            c_idx = col_pos + 1
            if show_jnpt:
                ws.cell(row=curr_row, column=c_idx, value=row_tot_jnpt).number_format = "#,##0"
                c_idx += 1
            if show_jsw:
                ws.cell(row=curr_row, column=c_idx, value=row_tot_jsw).number_format = "#,##0"
                c_idx += 1

            for c in range(col_pos, col_pos + m_cols):
                cell = ws.cell(row=curr_row, column=c)
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="right")

            gt_totals["b"] += row_tot_b
            gt_totals["jnpt"] += row_tot_jnpt
            gt_totals["jsw"] += row_tot_jsw

            curr_row += 1

        # Total Row
        ws.cell(row=curr_row, column=1, value="Total").font = tot_font
        ws.cell(row=curr_row, column=1).border = cell_border

        col_pos = 2
        for m in visible_months:
            idx = m["idx"]
            ws.cell(row=curr_row, column=col_pos, value=month_totals[idx]["b"]).number_format = "#,##0"
            c_idx = col_pos + 1
            if show_jnpt:
                ws.cell(row=curr_row, column=c_idx, value=month_totals[idx]["jnpt"]).number_format = "#,##0"
                c_idx += 1
            if show_jsw:
                ws.cell(row=curr_row, column=c_idx, value=month_totals[idx]["jsw"]).number_format = "#,##0"
                c_idx += 1

            for c in range(col_pos, col_pos + m_cols):
                cell = ws.cell(row=curr_row, column=c)
                cell.font = tot_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="right")

            col_pos += m_cols

        ws.cell(row=curr_row, column=col_pos, value=gt_totals["b"]).number_format = "#,##0"
        c_idx = col_pos + 1
        if show_jnpt:
            ws.cell(row=curr_row, column=c_idx, value=gt_totals["jnpt"]).number_format = "#,##0"
            c_idx += 1
        if show_jsw:
            ws.cell(row=curr_row, column=c_idx, value=gt_totals["jsw"]).number_format = "#,##0"
            c_idx += 1

        for c in range(col_pos, col_pos + m_cols):
            cell = ws.cell(row=curr_row, column=c)
            cell.font = tot_font
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Budget_vs_Actual_{fin_year}.xlsx"
        return send_file(output, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to export Excel report: {e}"}), 500


# -------------------------------------------------------------------------
# BUDGET TARGET MAINTENANCE ENDPOINTS (financial_year_targets JSONB Schema)
# -------------------------------------------------------------------------

# -------------------------------------------------------------------------
# BUDGET TARGET MAINTENANCE ENDPOINTS (financial_year_targets JSONB Schema)
# -------------------------------------------------------------------------

CARGO_MASTER_COLUMNS = [
    {"field": "cargo_type", "label": "Cargo Type"},
    {"field": "cargo_category", "label": "Cargo Category"},
    {"field": "cargo_category_2", "label": "Cargo Category 2"},
    {"field": "cargo_sub_category", "label": "Cargo Sub Category"},
    {"field": "cargo_sub_category_2", "label": "Cargo Sub Category 2"},
]
CARGO_MASTER_COLUMN_MAP = {c["field"]: c["label"] for c in CARGO_MASTER_COLUMNS}


@bp.route("/api/module/RP01/report-budget/cargo-master-columns")
@login_required
def report_budget_api_cargo_master_columns():
    """Get approved Cargo Master classification columns."""
    return jsonify({"columns": CARGO_MASTER_COLUMNS})


@bp.route("/api/module/RP01/report-budget/cargo-master-values")
@login_required
def report_budget_api_cargo_master_values():
    """Get distinct values from existing Cargo Master (vessel_cargo) for selected column."""
    try:
        column = request.args.get("column", "").strip()
        if column not in CARGO_MASTER_COLUMN_MAP:
            return jsonify({"error": f"Invalid or unauthorized column '{column}'."}), 400

        conn = get_db()
        try:
            cur = get_cursor(conn)
            # Safe column insertion after strict allowlist validation
            query = f"""
                SELECT DISTINCT {column} AS val
                FROM vessel_cargo
                WHERE {column} IS NOT NULL AND TRIM({column}) != ''
                ORDER BY {column};
            """
            cur.execute(query)
            values = [r["val"].strip() for r in cur.fetchall() if r["val"]]
        finally:
            conn.close()

        return jsonify({"column": column, "values": values})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to fetch values for column {column}: {e}"}), 500


@bp.route("/api/module/RP01/report-budget/cargo-sub-categories")
@login_required
def report_budget_api_cargo_sub_categories():
    """Get Cargo Sub Category-2 options dynamically from existing Cargo Master (vessel_cargo)."""
    return report_budget_api_cargo_master_values()


@bp.route("/api/module/RP01/report-budget/financial-year-targets")
@login_required
def report_budget_api_get_fy_targets():
    """Get targets JSONB for a specific financial year as per spec Section 7."""
    try:
        fin_year = request.args.get("financial_year", "").strip()
        if not fin_year:
            return jsonify({"error": "Financial Year query parameter is required"}), 400

        conn = get_db()
        targets_list = []
        try:
            cur = get_cursor(conn)
            cur.execute("SELECT id, financial_year, targets, updated_at FROM financial_year_targets WHERE financial_year = %s;", (fin_year,))
            r = cur.fetchone()
            if r and r["targets"]:
                import json
                t_raw = r["targets"]
                if isinstance(t_raw, str):
                    t_raw = json.loads(t_raw)
                if isinstance(t_raw, dict):
                    targets_list = t_raw.get("targets", [])
        finally:
            conn.close()

        return jsonify({
            "financial_year": fin_year,
            "targets": targets_list
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to fetch financial year targets: {e}"}), 500


@bp.route("/api/module/RP01/report-budget/targets")
@login_required
def report_budget_api_get_targets():
    """Get saved budget targets from financial_year_targets table as a list of target entries for UI grid."""
    try:
        fin_year = request.args.get("financial_year")
        conn = get_db()
        rows = []
        try:
            cur = get_cursor(conn)
            if fin_year:
                cur.execute("SELECT id, financial_year, targets, updated_at FROM financial_year_targets WHERE financial_year = %s;", (fin_year,))
            else:
                cur.execute("SELECT id, financial_year, targets, updated_at FROM financial_year_targets ORDER BY financial_year DESC;")
            
            db_rows = cur.fetchall()
            import json
            for r in db_rows:
                fy = r["financial_year"]
                targets_raw = r["targets"]
                if isinstance(targets_raw, str):
                    targets_raw = json.loads(targets_raw)
                
                if isinstance(targets_raw, dict):
                    # 1. New schema: {"targets": [{"column": ..., "name": ..., "monthly_data": [...]}]}
                    target_items = targets_raw.get("targets", [])
                    for t_item in target_items:
                        col = t_item.get("column", "")
                        col_lbl = CARGO_MASTER_COLUMN_MAP.get(col, col)
                        name = t_item.get("name", "")
                        monthly_data = t_item.get("monthly_data", [])
                        for md in monthly_data:
                            rows.append({
                                "id": r["id"],
                                "financial_year": fy,
                                "column": col,
                                "column_label": col_lbl,
                                "name": name,
                                "cargo_sub_category_2": name,
                                "month": md.get("month"),
                                "outlook": float(md.get("outlook", 0.0)),
                                "target": float(md.get("target", 0.0)),
                                "budget_quantity": float(md.get("target") or md.get("outlook") or 0.0),
                                "updated_at": str(r["updated_at"])
                            })

                    # 2. Legacy schema fallback: {"cargo_groups": [...]}
                    cargo_groups = targets_raw.get("cargo_groups", [])
                    for group in cargo_groups:
                        comm = group.get("cargo_sub_category_2", "")
                        monthly = group.get("monthly_budget", [])
                        for mb in monthly:
                            rows.append({
                                "id": r["id"],
                                "financial_year": fy,
                                "column": "cargo_sub_category_2",
                                "column_label": "Cargo Sub Category 2",
                                "name": comm,
                                "cargo_sub_category_2": comm,
                                "month": mb.get("month"),
                                "outlook": float(mb.get("budget_quantity", 0.0)),
                                "target": float(mb.get("budget_quantity", 0.0)),
                                "budget_quantity": float(mb.get("budget_quantity", 0.0)),
                                "updated_at": str(r["updated_at"])
                            })
        finally:
            conn.close()

        return jsonify({"targets": rows})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to fetch budget targets: {e}"}), 500


@bp.route("/api/module/RP01/report-budget/target-save", methods=["POST"])
@login_required
def report_budget_api_save_target():
    """Save/update a monthly budget target entry into financial_year_targets JSONB as per spec Section 5 & 6."""
    try:
        data = request.get_json(force=True) or {}
        financial_year = (data.get("financial_year") or "").strip()
        column = (data.get("column") or "").strip()
        name = (data.get("name") or "").strip()
        month = (data.get("month") or "").strip()
        
        # Backward compatibility for legacy inputs
        if not column and data.get("cargo_sub_category_2"):
            column = "cargo_sub_category_2"
            name = data.get("cargo_sub_category_2").strip()

        # Parse numeric outlook and target
        try:
            outlook = float(data.get("outlook") if "outlook" in data else data.get("budget_quantity", 0.0))
        except (ValueError, TypeError):
            return jsonify({"error": "Outlook must be a valid numeric value"}), 400

        try:
            target = float(data.get("target") if "target" in data else data.get("budget_quantity", 0.0))
        except (ValueError, TypeError):
            return jsonify({"error": "Target must be a valid numeric value"}), 400

        # Step 1-6 validations
        if not financial_year:
            return jsonify({"error": "Financial Year is required"}), 400
        if column not in CARGO_MASTER_COLUMN_MAP:
            return jsonify({"error": f"Invalid Cargo Master column '{column}'."}), 400
        if not name:
            return jsonify({"error": "Name value is required"}), 400
        if not month or month not in MONTH_NAMES:
            return jsonify({"error": f"Valid Month is required ({', '.join(MONTH_NAMES)})"}), 400
        if outlook < 0 or target < 0:
            return jsonify({"error": "Outlook and Target cannot be negative"}), 400

        conn = get_db()
        try:
            cur = get_cursor(conn)
            # Step 4: Validate that name exists for the selected Cargo Master column
            query = f"SELECT COUNT(*) AS cnt FROM vessel_cargo WHERE LOWER(TRIM({column})) = LOWER(%s);"
            cur.execute(query, (name,))
            if cur.fetchone()["cnt"] == 0:
                return jsonify({"error": f"Name '{name}' does not exist in Cargo Master for column '{CARGO_MASTER_COLUMN_MAP[column]}'."}), 400

            # Step 7-8: Get or create financial_year_targets row
            cur.execute("SELECT id, targets FROM financial_year_targets WHERE financial_year = %s;", (financial_year,))
            fy_row = cur.fetchone()
            import json

            if fy_row:
                rec_id = fy_row["id"]
                targets_raw = fy_row["targets"]
                if isinstance(targets_raw, str):
                    targets_raw = json.loads(targets_raw)
            else:
                targets_raw = {"targets": []}
                rec_id = None

            if not isinstance(targets_raw, dict):
                targets_raw = {"targets": []}
            if "targets" not in targets_raw or not isinstance(targets_raw["targets"], list):
                targets_raw["targets"] = []

            # Step 9-10: Find or create matching target object (by column + name)
            target_obj = None
            for t_item in targets_raw["targets"]:
                if (t_item.get("column") or "").strip().lower() == column.lower() and \
                   (t_item.get("name") or "").strip().lower() == name.lower():
                    target_obj = t_item
                    break

            if not target_obj:
                target_obj = {
                    "column": column,
                    "name": name,
                    "monthly_data": []
                }
                targets_raw["targets"].append(target_obj)
            else:
                if "monthly_data" not in target_obj or not isinstance(target_obj["monthly_data"], list):
                    target_obj["monthly_data"] = []

            # Step 11-13: Find or update matching month inside monthly_data (Duplicate Prevention)
            month_obj = None
            for md in target_obj["monthly_data"]:
                if (md.get("month") or "").strip().lower() == month.lower():
                    month_obj = md
                    break

            if month_obj:
                month_obj["outlook"] = outlook
                month_obj["target"] = target
            else:
                target_obj["monthly_data"].append({
                    "month": month,
                    "outlook": outlook,
                    "target": target
                })

            # Step 14-15: Save updated JSONB back to PostgreSQL
            targets_json = json.dumps(targets_raw)
            if rec_id:
                cur.execute("""
                    UPDATE financial_year_targets
                    SET targets = %s::jsonb, updated_at = NOW()
                    WHERE id = %s;
                """, (targets_json, rec_id))
            else:
                cur.execute("""
                    INSERT INTO financial_year_targets (financial_year, targets, updated_at)
                    VALUES (%s, %s::jsonb, NOW());
                """, (financial_year, targets_json))

            conn.commit()
        finally:
            conn.close()

        return jsonify({"success": True, "message": "Financial year target saved successfully."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to save budget target: {e}"}), 500


@bp.route("/api/module/RP01/report-budget/target-delete", methods=["POST"])
@login_required
def report_budget_api_delete_target():
    """Delete a specific (column + name + month) target entry from financial_year_targets JSONB."""
    try:
        data = request.get_json(force=True) or {}
        financial_year = (data.get("financial_year") or "").strip()
        column = (data.get("column") or "cargo_sub_category_2").strip()
        name = (data.get("name") or data.get("cargo_sub_category_2") or "").strip()
        month = (data.get("month") or "").strip()

        if not financial_year or not column or not name or not month:
            return jsonify({"error": "Missing required parameters (financial_year, column, name, month) for deletion"}), 400

        conn = get_db()
        try:
            cur = get_cursor(conn)
            cur.execute("SELECT id, targets FROM financial_year_targets WHERE financial_year = %s;", (financial_year,))
            fy_row = cur.fetchone()
            import json

            if fy_row and fy_row["targets"]:
                rec_id = fy_row["id"]
                targets_raw = fy_row["targets"]
                if isinstance(targets_raw, str):
                    targets_raw = json.loads(targets_raw)
                
                if isinstance(targets_raw, dict) and "targets" in targets_raw:
                    target_items = targets_raw.get("targets", [])
                    new_targets = []
                    for t_item in target_items:
                        if (t_item.get("column") or "").strip().lower() == column.lower() and \
                           (t_item.get("name") or "").strip().lower() == name.lower():
                            md_list = t_item.get("monthly_data", [])
                            t_item["monthly_data"] = [md for md in md_list if (md.get("month") or "").strip().lower() != month.lower()]
                            if t_item["monthly_data"]:
                                new_targets.append(t_item)
                        else:
                            new_targets.append(t_item)
                    targets_raw["targets"] = new_targets

                cur.execute("""
                    UPDATE financial_year_targets
                    SET targets = %s::jsonb, updated_at = NOW()
                    WHERE id = %s;
                """, (json.dumps(targets_raw), rec_id))

            conn.commit()
        finally:
            conn.close()

        return jsonify({"success": True, "message": "Target deleted successfully."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Failed to delete budget target: {e}"}), 500



"""
Report-10 — Export Performance Report (Monthly & Year-to-Date)
Flask Blueprint version. Reads from mis_vessel_master (legacy) and the
live LUEU01 pipeline (vcn_header / ldud_header / lueu_parcel_log).

Structure mirrors report09.py — same login_required, same dual data-source
fallback, same FY Apr-Mar convention.

The report table mirrors the Excel screenshot:
  Rows   : Unload (Import), Load (Export), Anchorage, Total
  Columns: During the month | Last year this month | Var.%
           | 2026-27 Up-to-current-month | 2025-26 Up-to-prev-year-current-month | Var.%

All data is FULLY DYNAMIC from the DB — no hard-coded quantity values.
"""

import io
import datetime
import traceback
from functools import wraps

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from .. import bp

from database import get_db, get_cursor


# ---------------------------------------------------------------------------
# Auth helper
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# FY / Month helpers  (Apr-Mar fiscal year, same convention as report9)
# ---------------------------------------------------------------------------

MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep",
               "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
CAL_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def _dt_to_fy_month(dt):
    """datetime/date  ->  (fin_year_str, fy_month_idx)
    e.g. 2026-07-12  ->  ('2026-27', 3)
    """
    d = dt.date() if isinstance(dt, datetime.datetime) else dt
    fy_start = d.year if d.month >= 4 else d.year - 1
    fin_year = f"{fy_start}-{str(fy_start + 1)[-2:]}"
    mn = CAL_MONTH_ABBR[d.month - 1]
    return fin_year, MONTH_NAMES.index(mn)


def _prev_fy(fin_year: str) -> str:
    """'2026-27' -> '2025-26'"""
    sy = fy_start_year(fin_year)
    return f"{sy - 1}-{str(sy)[-2:]}"


def _parse_dt(v):
    """Parse free-text datetime strings from the live pipeline."""
    if not v:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.datetime.strptime(s.replace("T", " ")[:16], "%Y-%m-%d %H:%M")
    except ValueError:
        try:
            return datetime.datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None


def month_str_to_idx(month_str: str) -> int:
    """'Jun-26' -> 2,  'Dec-24' -> 8 (FY Apr..Mar order)."""
    abbrev = str(month_str).split("-")[0].strip()
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized month value: '{month_str}' (expected e.g. 'Jun-26')"
        )


def month_options_for(fin_year: str) -> list:
    """All 12 month options for the given FY, as {idx, label}."""
    sy = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = sy if idx < 9 else sy + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


# ---------------------------------------------------------------------------
# Report row definitions
# ---------------------------------------------------------------------------

class ReportDataError(Exception):
    pass


# The four logical rows in the report.
# "import" means Import traffic (Unload), "export" means Export traffic (Load).
ROW_DEFS = [
    {"key": "Unload",    "op_type": "import",  "type": "data",  "label": "Unload"},
    {"key": "Load",      "op_type": "export",  "type": "data",  "label": "Load"},
    {"key": "Anchorage", "op_type": None,       "type": "data",  "label": "Anchorage"},
    {"key": "Total",     "op_type": None,       "type": "total", "label": "Total"},
]

# Berths counted as "Anchorage" traffic
ANCHORAGE_BERTHS = {"INN ANCHORAGE", "INN ANCH", "ANCHORAGE"}


# ---------------------------------------------------------------------------
# Data loading  (same dual-source fallback as report9)
# ---------------------------------------------------------------------------

def _load_mis_data() -> pd.DataFrame:
    """Load from mis_vessel_master (legacy source)."""
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT fin_year, month, berth_no, vcn_no, import_export, quantity
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month    IS NOT NULL
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        return pd.DataFrame(columns=["fin_year", "fy_month_idx",
                                     "berth_no", "vcn_no", "import_export", "quantity"])

    df = pd.DataFrame(rows)
    df["fin_year"]      = df["fin_year"].astype(str).str.strip()
    df["berth_no"]      = df["berth_no"].astype(str).str.strip()
    df["vcn_no"]        = df["vcn_no"].astype(str).str.strip()
    df["import_export"] = df["import_export"].astype(str).str.strip().str.lower()
    df["quantity"]      = pd.to_numeric(df["quantity"], errors="coerce").fillna(0.0)
    df["fy_month_idx"]  = df["month"].apply(month_str_to_idx)
    return df[["fin_year", "fy_month_idx", "berth_no", "vcn_no",
               "import_export", "quantity"]].copy()


def _load_live_pipeline_data() -> pd.DataFrame:
    """Load from the live LUEU01 pipeline for periods not in mis_vessel_master."""
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT ld.cast_off_datetime, h.id AS vcn_id,
                   h.berth_name, h.operation_type, l.quantity
            FROM   lueu_parcel_log l
            JOIN   ldud_parcel_ops po ON po.id = l.parcel_op_id
            JOIN   ldud_header     ld ON ld.id = po.ldud_id
            JOIN   vcn_header       h ON  h.id = ld.vcn_id
            WHERE  l.is_deleted IS NOT TRUE
              AND  COALESCE(l.is_shortclose, FALSE) = FALSE
              AND  ld.cast_off_datetime IS NOT NULL
              AND  l.quantity           IS NOT NULL
        """)
        log_rows = cur.fetchall()
    finally:
        conn.close()

    empty = pd.DataFrame(columns=["fin_year", "fy_month_idx",
                                   "berth_no", "vcn_no", "import_export", "quantity"])
    if not log_rows:
        return empty

    ldf = pd.DataFrame(log_rows)
    ldf["quantity"]      = pd.to_numeric(ldf["quantity"], errors="coerce").fillna(0.0)
    ldf["berth_no"]      = ldf["berth_name"].astype(str).str.strip()
    ldf["import_export"] = ldf["operation_type"].astype(str).str.strip().str.lower()
    ldf["vcn_no"]        = ldf["vcn_id"].astype(str)
    ldf["cast_off_dt"]   = ldf["cast_off_datetime"].apply(_parse_dt)
    ldf = ldf.dropna(subset=["cast_off_dt"])
    if ldf.empty:
        return empty

    fy_list, idx_list = [], []
    for dt in ldf["cast_off_dt"]:
        fy, idx = _dt_to_fy_month(dt)
        fy_list.append(fy)
        idx_list.append(idx)
    ldf["fin_year"]     = fy_list
    ldf["fy_month_idx"] = idx_list
    return ldf[["fin_year", "fy_month_idx", "berth_no", "vcn_no",
                "import_export", "quantity"]].copy()


def load_data() -> pd.DataFrame:
    """Merge legacy + live-pipeline with mis_vessel_master winning where it has data."""
    mv_df = _load_mis_data()
    covered = set(zip(mv_df["fin_year"], mv_df["fy_month_idx"]))

    live_df = _load_live_pipeline_data()
    if not live_df.empty:
        live_df = live_df[
            ~live_df.apply(lambda r: (r["fin_year"], r["fy_month_idx"]) in covered, axis=1)
        ]

    df_all = pd.concat([mv_df, live_df], ignore_index=True)
    if df_all.empty:
        raise ReportDataError(
            "No usable rows found in mis_vessel_master or the live LUEU01 pipeline."
        )
    return df_all


# ---------------------------------------------------------------------------
# Report computation
# ---------------------------------------------------------------------------

def _safe_var_pct(curr, prev):
    """Variance % — returns None if denominator is zero."""
    if prev is None or prev == 0:
        return None
    return round((curr - prev) / abs(prev) * 100, 2)


def _qty_for_op(subset: pd.DataFrame, op_type: str | None,
                anchorage: bool = False) -> float:
    """Sum quantity for a given operation type / anchorage flag."""
    if anchorage:
        mask = subset["berth_no"].str.upper().isin({b.upper() for b in ANCHORAGE_BERTHS})
        return round(float(subset.loc[mask, "quantity"].sum()), 3)
    if op_type is None:
        return round(float(subset["quantity"].sum()), 3)
    mask = subset["import_export"] == op_type
    return round(float(subset.loc[mask, "quantity"].sum()), 3)


def compute_report(df: pd.DataFrame, fin_year: str, month_idx: int) -> dict:
    """
    Returns a dict with keys:
      month_label, fin_year, prev_fin_year, rows
    Each row has:
      label, type,
      month_curr, month_prev, month_var_pct,
      upto_curr, upto_prev, upto_var_pct
    """
    prev_fy = _prev_fy(fin_year)

    fy_curr  = df[df["fin_year"] == fin_year]
    fy_prev  = df[df["fin_year"] == prev_fy]

    month_curr  = fy_curr[fy_curr["fy_month_idx"] == month_idx]
    month_prev  = fy_prev[fy_prev["fy_month_idx"] == month_idx]
    upto_curr   = fy_curr[fy_curr["fy_month_idx"] <= month_idx]
    upto_prev   = fy_prev[fy_prev["fy_month_idx"] <= month_idx]

    rows = []
    for rdef in ROW_DEFS:
        op = rdef["op_type"]
        is_anch = (rdef["key"] == "Anchorage")
        is_total = (rdef["type"] == "total")

        mc = _qty_for_op(month_curr, op, anchorage=is_anch)
        mp = _qty_for_op(month_prev, op, anchorage=is_anch)
        uc = _qty_for_op(upto_curr, op, anchorage=is_anch)
        up = _qty_for_op(upto_prev, op, anchorage=is_anch)

        rows.append({
            "label":         rdef["label"],
            "type":          rdef["type"],
            "month_curr":    mc,
            "month_prev":    mp,
            "month_var_pct": _safe_var_pct(mc, mp),
            "upto_curr":     uc,
            "upto_prev":     up,
            "upto_var_pct":  _safe_var_pct(uc, up),
        })

    return {
        "fin_year":      fin_year,
        "prev_fin_year": prev_fy,
        "rows":          rows,
    }


def _get_df_and_years():
    df = load_data()
    years = sorted(df["fin_year"].unique().tolist())
    return df, years


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------

@bp.route("/module/RP01/report10/")
@login_required
def report10_index():
    return render_template("/report10/report10.html")


@bp.route("/api/module/RP01/report10/meta")
@login_required
def report10_api_meta():
    try:
        _, years = _get_df_and_years()
        months = {fy: month_options_for(fy) for fy in years}
        return jsonify({"years": years, "months": months})
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report10/report")
@login_required
def report10_api_report():
    try:
        df, years = _get_df_and_years()
        fin_year  = request.args.get("fin_year", years[-1] if years else "")
        month_idx = int(request.args.get("month_idx", 2))

        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        result = compute_report(df, fin_year, month_idx)

        # Build the month label string (e.g. "Jun-26")
        opts = month_options_for(fin_year)
        month_label = next((o["label"] for o in opts if o["idx"] == month_idx), "")
        result["month_label"] = month_label

        return jsonify(result)
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report10/export")
@login_required
def report10_api_export():
    try:
        df, years = _get_df_and_years()
        fin_year  = request.args.get("fin_year", years[-1] if years else "")
        month_idx = int(request.args.get("month_idx", 2))

        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        result     = compute_report(df, fin_year, month_idx)
        opts       = month_options_for(fin_year)
        month_label = next((o["label"] for o in opts if o["idx"] == month_idx), "")
        prev_fy    = result["prev_fin_year"]

        # ---- Build Excel workbook ----------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Report-10"

        bold         = Font(bold=True)
        title_font   = Font(bold=True, size=12)
        center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left         = Alignment(horizontal="left",   vertical="center")
        right_al     = Alignment(horizontal="right",  vertical="center")

        thin  = Side(style="thin",   color="000000")
        thick = Side(style="medium", color="000000")
        thin_border   = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick_top_bdr = Border(left=thin, right=thin, top=thick, bottom=thin)

        yellow_fill   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_fill   = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        total_fill    = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
        title_fill    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        # Column layout:
        # A=Label | B=Curr Month | C=Prev Month | D=Var% | E=Upto Curr | F=Upto Prev | G=Var%
        COLS = ["A", "B", "C", "D", "E", "F", "G"]
        NUM_COLS = 7

        # Row 1 – date stamp
        ws["G1"] = datetime.date.today().strftime("%d-%b-%y")
        ws["G1"].alignment = right_al

        # Row 3 – main title
        ws.merge_cells("A3:G3")
        ws["A3"] = "EXPORT PERFORMANCE REPORT"
        ws["A3"].font = title_font
        ws["A3"].alignment = center
        ws["A3"].fill = title_fill

        # Row 5 – column group headers
        ws.merge_cells("B5:D5")
        ws["B5"] = month_label
        ws["B5"].font = bold
        ws["B5"].alignment = center
        ws["B5"].fill = yellow_fill

        ws.merge_cells("E5:G5")
        ws["E5"] = f"Up to {month_label}"
        ws["E5"].font = bold
        ws["E5"].alignment = center

        # Row 6 – sub-headers
        headers = [
            ("A6", ""),
            ("B6", f"During the month\n({month_label})"),
            ("C6", f"Last year this month\n({month_label.split('-')[0]}-{str(int(fin_year.split('-')[0]) - 1)[-2:]})"),
            ("D6", "Var. (%)"),
            ("E6", f"{fin_year}\n(Up to current month'{fin_year.split('-')[0]})"),
            ("F6", f"{prev_fy}\n(Up to previous year current month'{prev_fy.split('-')[0]})"),
            ("G6", "Var. (%)"),
        ]
        for addr, label in headers:
            cell = ws[addr]
            cell.value = label
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill
            cell.border = thin_border

        ws.row_dimensions[6].height = 48

        # Data rows starting at row 8
        data_row = 8
        for r in result["rows"]:
            is_total = r["type"] == "total"

            def _fmt(v):
                # Write None (blank cell) for zero or missing values
                if v is None or v == 0:
                    return None
                return v

            def _var(v):
                # Write None (blank cell) instead of "#DIV/0!" string
                if v is None:
                    return None
                return round(v, 2)

            values = [
                r["label"],
                _fmt(r["month_curr"]),
                _fmt(r["month_prev"]),
                _var(r["month_var_pct"]),
                _fmt(r["upto_curr"]),
                _fmt(r["upto_prev"]),
                _var(r["upto_var_pct"]),
            ]

            fill = total_fill if is_total else None
            bdr  = thick_top_bdr if is_total else thin_border

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.border = bdr
                if is_total:
                    cell.font = bold
                    if fill:
                        cell.fill = fill
                if col_idx == 1:
                    cell.alignment = left
                else:
                    cell.alignment = right_al
                    if col_idx not in (4, 7):  # not Var% cols
                        cell.number_format = "#,##0.000"
                    else:
                        cell.number_format = "0.00"

            data_row += 1

        # Column widths
        widths = {"A": 20, "B": 18, "C": 22, "D": 10, "E": 22, "F": 26, "G": 10}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-10_Export_Performance_{fin_year}_{month_label}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500

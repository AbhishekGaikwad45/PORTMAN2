from flask import render_template, request, session, redirect, url_for, Response, jsonify
from functools import wraps
from datetime import date, datetime, timedelta
import io
import json

from .. import bp
from database import get_db, get_cursor

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants (same as rest of RP01) ──────────────────────────────────
XL_NORM_SZ = 11
_thin  = Side(style='thin',   color='000000')
_bdr   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr   = Alignment(horizontal='center', vertical='center', wrap_text=False)
_left  = Alignment(horizontal='left',   vertical='center', wrap_text=False)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, size=XL_NORM_SZ):
    return Font(name='Calibri', bold=bold, size=size)


def _parse_dt(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val))
    except Exception:
        return None


def _fmt_dt(val, strfmt='%d-%m-%Y %H:%M'):
    dt = _parse_dt(val)
    return dt.strftime(strfmt) if dt else ''


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── Routes ──────────────────────────────────────────────────────────────────

@bp.route('/module/RP01/daily-ops/')
@login_required
def daily_ops_index():
    return render_template('daily_ops/daily_ops.html', username=session.get('username'))


# ── Routes API (from CRM01 conveyor_routes) ─────────────────────────────────

@bp.route('/api/module/RP01/daily-ops/routes', methods=['GET'])
@login_required
def daily_ops_routes():
    """Return all active route names from conveyor_routes."""
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("SELECT route_name FROM conveyor_routes WHERE is_active = 1 ORDER BY route_name")
    routes = [r['route_name'] for r in cur.fetchall()]
    conn.close()
    return jsonify(routes)


# ── Cutoff API ──────────────────────────────────────────────────────────────

@bp.route('/api/module/RP01/daily-ops/cutoff', methods=['GET'])
@login_required
def daily_ops_cutoff_get():
    """Return the latest cutoff record (or empty defaults)."""
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("""
        SELECT id, cutoff_date, cutoff_values
        FROM daily_ops_cutoff
        ORDER BY id DESC LIMIT 1
    """)
    row = cur.fetchone()
    conn.close()
    if row:
        return jsonify({
            'id':            row['id'],
            'cutoff_date':   row['cutoff_date'],
            'cutoff_values': json.loads(row['cutoff_values']),
        })
    return jsonify({
        'id':            None,
        'cutoff_date':   '',
        'cutoff_values': {'mbc_cargo': {}, 'cargo_handled': {}},
    })


@bp.route('/api/module/RP01/daily-ops/cutoff', methods=['POST'])
@login_required
def daily_ops_cutoff_save():
    """Upsert cutoff: replace any existing row with the new values."""
    data = request.get_json(force=True)
    cutoff_date   = data.get('cutoff_date', '')
    cutoff_values = data.get('cutoff_values', {})

    if not cutoff_date:
        return Response('cutoff_date is required', status=400)

    values_json = json.dumps(cutoff_values)
    user = session.get('username', '')

    conn = get_db()
    cur  = get_cursor(conn)
    # Delete all existing rows (single-row table)
    cur.execute("DELETE FROM daily_ops_cutoff")
    cur.execute("""
        INSERT INTO daily_ops_cutoff (cutoff_date, cutoff_values, created_by)
        VALUES (%s, %s, %s)
    """, (cutoff_date, values_json, user))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ── Cutoff helper ───────────────────────────────────────────────────────────

def _load_cutoff():
    """Return (cutoff_date_str, cutoff_values_dict) or (None, {})."""
    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("""
        SELECT cutoff_date, cutoff_values
        FROM daily_ops_cutoff
        ORDER BY id DESC LIMIT 1
    """)
    row = cur.fetchone()
    conn.close()
    if row:
        return row['cutoff_date'], json.loads(row['cutoff_values'])
    return None, {}


# ── Data fetchers ───────────────────────────────────────────────────────────

def _fetch_data(report_date):
    """Fetch all non-closed vessels from LDUD01, with BL qty from VCN
    and unloaded qty from LUEU01."""
    window_end   = datetime(report_date.year, report_date.month, report_date.day, 7, 0, 0)
    window_start = window_end - timedelta(hours=24)
    ws_str = window_start.strftime('%Y-%m-%d %H:%M:%S')
    we_str = window_end.strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cur  = get_cursor(conn)

    # All non-closed vessels (Draft or Partial Close)
    cur.execute("""
        SELECT h.id, h.vcn_id, h.vessel_name, h.operation_type,
               h.nor_tendered, h.discharge_commenced, h.discharge_completed,
               h.doc_status
        FROM ldud_header h
        WHERE h.doc_status != 'Closed'
        ORDER BY h.discharge_commenced ASC NULLS LAST
    """)
    vessels = [dict(r) for r in cur.fetchall()]

    ldud_ids = [v['id']     for v in vessels]
    vcn_ids  = [v['vcn_id'] for v in vessels if v.get('vcn_id')]

    # BL quantities from VCN (both import and export)
    bl_import  = {}
    bl_export  = {}
    vcn_meta   = {}
    if vcn_ids:
        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity), 0) AS total
            FROM vcn_cargo_declaration WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        for r in cur.fetchall():
            bl_import[r['vcn_id']] = float(r['total'])

        cur.execute("""
            SELECT vcn_id, COALESCE(SUM(bl_quantity), 0) AS total
            FROM vcn_export_cargo_declaration WHERE vcn_id = ANY(%s) GROUP BY vcn_id
        """, (vcn_ids,))
        for r in cur.fetchall():
            bl_export[r['vcn_id']] = float(r['total'])

        cur.execute("""
            SELECT id, importer_exporter_name FROM vcn_header WHERE id = ANY(%s)
        """, (vcn_ids,))
        vcn_meta = {r['id']: r['importer_exporter_name'] or '' for r in cur.fetchall()}

    # Unloaded quantities from LUEU01 — total per source vessel
    lueu_total = {}
    lueu_24h   = {}
    if ldud_ids:
        # Total unloaded till report date 7AM
        cur.execute("""
            SELECT source_id, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE source_type = 'LDUD'
              AND source_id = ANY(%s)
              AND entry_date IS NOT NULL
              AND (entry_date || ' ' || COALESCE(from_time, '00:00')) < %s
            GROUP BY source_id
        """, (ldud_ids, we_str))
        for r in cur.fetchall():
            lueu_total[r['source_id']] = float(r['qty'])

        # 24h unloaded
        cur.execute("""
            SELECT source_id, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE source_type = 'LDUD'
              AND source_id = ANY(%s)
              AND entry_date IS NOT NULL
              AND (entry_date || ' ' || COALESCE(from_time, '00:00')) >= %s
              AND (entry_date || ' ' || COALESCE(from_time, '00:00')) < %s
            GROUP BY source_id
        """, (ldud_ids, ws_str, we_str))
        for r in cur.fetchall():
            lueu_24h[r['source_id']] = float(r['qty'])

    conn.close()

    for v in vessels:
        lid        = v['id']
        vid        = v.get('vcn_id')
        op         = v.get('operation_type', '')
        bl_qty     = (bl_export.get(vid, 0) if op == 'Export' else bl_import.get(vid, 0)) if vid else 0
        unloaded   = lueu_total.get(lid, 0)
        v['stevedore_group']        = vcn_meta.get(vid, '') if vid else ''
        v['bl_qty']                 = bl_qty
        v['ops_24h']                = lueu_24h.get(lid, 0)
        v['ops_till']               = unloaded
        v['balance']                = bl_qty - unloaded

    return vessels


def _fetch_cargo_handled(report_date):
    """Fetch cargo handled by route (day + month).
    Month values incorporate cutoff if the cutoff date falls within the report month.
    """
    window_end   = datetime(report_date.year, report_date.month, report_date.day, 7, 0, 0)
    window_start = window_end - timedelta(hours=24)
    month_start  = datetime(report_date.year, report_date.month, 1, 7, 0, 0)
    we_str  = window_end.strftime('%Y-%m-%d %H:%M:%S')
    ws_str  = window_start.strftime('%Y-%m-%d %H:%M:%S')

    # ── Load cutoff ─────────────────────────────────────────────────────
    cutoff_date_str, cutoff_vals = _load_cutoff()
    cargo_cutoff = cutoff_vals.get('cargo_handled', {})

    # Determine if cutoff applies to this month
    cutoff_7am = None
    if cutoff_date_str and cargo_cutoff:
        try:
            cd = datetime.strptime(cutoff_date_str, '%Y-%m-%d')
            cutoff_7am = datetime(cd.year, cd.month, cd.day, 7, 0, 0)
        except ValueError:
            pass

    use_cutoff = (cutoff_7am is not None
                  and month_start < cutoff_7am
                  and cutoff_7am <= window_end)

    conn = get_db()
    cur  = get_cursor(conn)

    def _period(start, end):
        cur.execute("""
            SELECT route_name, COALESCE(SUM(quantity), 0) AS qty
            FROM lueu_lines
            WHERE route_name IS NOT NULL AND route_name != ''
              AND entry_date IS NOT NULL
              AND (entry_date || ' ' || COALESCE(from_time, '00:00')) >= %s
              AND (entry_date || ' ' || COALESCE(from_time, '00:00')) < %s
            GROUP BY route_name
            ORDER BY route_name
        """, (start, end))
        return {r['route_name']: float(r['qty']) for r in cur.fetchall()}

    day_dict = _period(ws_str, we_str)

    if use_cutoff:
        # Query only from cutoff 7AM onwards for the month
        cutoff_str = cutoff_7am.strftime('%Y-%m-%d %H:%M:%S')
        live_dict = _period(cutoff_str, we_str)
        # Merge: cutoff values + live values
        month_dict = {}
        all_routes = set(list(cargo_cutoff.keys()) + list(live_dict.keys()))
        for route in all_routes:
            co_val   = float(cargo_cutoff.get(route, 0))
            live_val = live_dict.get(route, 0)
            month_dict[route] = co_val + live_val
    else:
        mth_str = month_start.strftime('%Y-%m-%d %H:%M:%S')
        month_dict = _period(mth_str, we_str)

    conn.close()

    day_rows   = sorted(day_dict.items())
    month_rows = sorted(month_dict.items())
    return day_rows, month_rows


def _fetch_tide_data(report_date):
    window_end   = datetime(report_date.year, report_date.month, report_date.day, 7, 0, 0)
    window_start = window_end - timedelta(hours=24)
    we_str = window_end.strftime('%Y-%m-%dT%H:%M')
    ws_str = window_start.strftime('%Y-%m-%dT%H:%M')

    conn = get_db()
    cur  = get_cursor(conn)
    cur.execute("""
        SELECT tide_datetime, tide_meters
        FROM tide_master
        WHERE tide_datetime >= %s AND tide_datetime < %s
        ORDER BY tide_datetime ASC
    """, (ws_str, we_str))
    rows = [(r['tide_datetime'], float(r['tide_meters'])) for r in cur.fetchall()]
    conn.close()
    return rows


def _fmt_tide_dt(dt_str):
    """'2026-01-27T16:00' -> '27/16:00'"""
    try:
        dt = datetime.fromisoformat(dt_str)
        return dt.strftime('%d/%H:%M')
    except Exception:
        return dt_str


# ── Excel builder ───────────────────────────────────────────────────────────

def _build_excel(vessels, report_date,
                 day_rows=None, month_rows=None, tide_rows=None):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Ops'
    day_rows   = day_rows   or []
    month_rows = month_rows or []
    tide_rows  = tide_rows  or []

    col_widths = {1: 30, 2: 35, 3: 35, 4: 35, 5: 35, 6: 35, 7: 10, 8: 32, 9: 22}
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    def _cell(r, c, val='', bold=False, fill='FFFFFF', align=_ctr):
        cell = ws.cell(r, c, val)
        cell.font      = _font(bold=bold)
        cell.fill      = _fill(fill)
        cell.alignment = align
        cell.border    = _bdr
        return cell

    def _merge_row(r, c1, c2, val='', bold=False, fill='FFFFFF', align=_ctr):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        for ci in range(c1, c2 + 1):
            b = Border(
                left   = _thin if ci == c1 else None,
                right  = _thin if ci == c2 else None,
                top    = _thin,
                bottom = _thin,
            )
            try:
                cell        = ws.cell(r, ci)
                cell.fill   = _fill(fill)
                cell.border = b
            except AttributeError:
                pass
        anchor           = ws.cell(r, c1)
        anchor.value     = val
        anchor.font      = _font(bold=bold)
        anchor.alignment = align

    def _merge_col(r1, r2, c, val='', bold=False, fill='FFFFFF', align=_ctr):
        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
        for ri in range(r1, r2 + 1):
            b = Border(
                left   = _thin,
                right  = _thin,
                top    = _thin if ri == r1 else None,
                bottom = _thin if ri == r2 else None,
            )
            try:
                cell        = ws.cell(ri, c)
                cell.fill   = _fill(fill)
                cell.border = b
            except AttributeError:
                pass
        anchor           = ws.cell(r1, c)
        anchor.value     = val
        anchor.font      = _font(bold=bold)
        anchor.alignment = align

    date_str  = f"{report_date.day}.{report_date.month}.{report_date.year}"
    title_str = f'Daily Report of JSW Dharamtar Port Operation : {date_str}'

    # Row 1
    ws.row_dimensions[1].height = 20
    _cell(1, 1, report_date.strftime('%d-%m-%Y'), align=_left)
    _merge_row(1, 2, 7, title_str, align=_ctr)
    _cell(1, 8, 'Doc No. | REV.02 | Issue no. 02', align=_left)
    _cell(1, 9, f'Issue Date: {report_date.strftime("%d-%m-%Y")}', align=_left)

    # Row 2: vessel name headers
    ws.row_dimensions[2].height = 20
    _cell(2, 1, '')
    for i, v in enumerate(vessels):
        _cell(2, 2 + i, f'Vessel {i + 1}: {v["vessel_name"]}', bold=True, align=_ctr)
    for i in range(len(vessels), 5):
        _cell(2, 2 + i, '')
    _cell(2, 7, '')
    _cell(2, 8, '')
    _cell(2, 9, '')

    label_discharge = 'Unloaded till Date (LUEU)'
    label_balance   = 'Balance'
    label_commenced = 'Disch Commenced'
    label_completed = 'Disch Completed'

    _q = lambda x: int(round(x)) if x else ''
    _n = lambda x: x if x else ''
    ROWS = [
        ('Stevedore Group',                 'stevedore_group',          None,       _left),
        ('BL Qty',                          'bl_qty',                   _q,         _ctr),
        ('24 hrs Discharge',                'ops_24h',                  _q,         _ctr),
        (label_discharge,                   'ops_till',                 _q,         _ctr),
        (label_balance,                     'balance',                  _q,         _ctr),
        ('Vsl Arrived/NOR',                 'nor_tendered',             _fmt_dt,    _ctr),
        (label_commenced,                   'discharge_commenced',      _fmt_dt,    _ctr),
        (label_completed,                   'discharge_completed',      _fmt_dt,    _ctr),
    ]

    for idx, (label, field, formatter, align) in enumerate(ROWS):
        r = 3 + idx
        ws.row_dimensions[r].height = 18

        if label is None:
            for ci in range(1, 10):
                _cell(r, ci, '')
            continue

        _cell(r, 1, label, bold=True, align=_left)
        for i, v in enumerate(vessels):
            raw = v.get(field)
            val = formatter(raw) if (formatter and raw is not None) else (raw or '')
            _cell(r, 2 + i, val, align=align)
        for i in range(len(vessels), 5):
            _cell(r, 2 + i, '')
        _cell(r, 7, '')
        _cell(r, 8, '')
        _cell(r, 9, '')

    # ── Cargo Handled section ────────────────────────────────────────────────
    cargo_start = 3 + len(ROWS)

    def _cargo_section(row_start, period_rows, period_label):
        r = row_start
        n = len(period_rows) + 1
        _merge_col(r, r + n - 1, 1, period_label, bold=True, align=_ctr)
        for route_name, qty in period_rows:
            _cell(r, 2, route_name, align=_left)
            _cell(r, 3, int(round(qty)) if qty else '', align=_ctr)
            for ci in range(4, 10):
                _cell(r, ci, '')
            ws.row_dimensions[r].height = 18
            r += 1
        total = sum(q for _, q in period_rows)
        _cell(r, 2, 'Total:', bold=True, align=_left)
        _cell(r, 3, int(round(total)) if total else '', bold=True, align=_ctr)
        for ci in range(4, 10):
            _cell(r, ci, '')
        ws.row_dimensions[r].height = 18
        r += 1
        return r

    r = cargo_start
    for ci in range(1, 10):
        _cell(r, ci, '')
    ws.row_dimensions[r].height = 18
    r += 1
    _merge_row(r, 1, 3, 'Cargo Handled', bold=True, align=_left)
    for ci in range(4, 10):
        _cell(r, ci, '')
    ws.row_dimensions[r].height = 18
    r += 1
    r = _cargo_section(r, day_rows, 'For the Day')
    r = _cargo_section(r, month_rows, 'For the Month')

    # ── Tide — Dharamtar Port section ────────────────────────────────────────
    for ci in range(1, 10):
        _cell(r, ci, '')
    ws.row_dimensions[r].height = 18
    r += 1
    _merge_row(r, 1, 2, 'Tide- Dharamtar Port', bold=False, align=_ctr)
    for ci in range(3, 10):
        ws.cell(r, ci).value = None
    ws.row_dimensions[r].height = 18
    r += 1
    _cell(r, 1, 'Time', align=_ctr)
    _cell(r, 2, 'Tide', align=_ctr)
    ws.row_dimensions[r].height = 18
    r += 1
    for td_str, td_m in tide_rows:
        _cell(r, 1, _fmt_tide_dt(td_str), align=_ctr)
        _cell(r, 2, td_m, align=_ctr)
        ws.row_dimensions[r].height = 18
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Download endpoint ───────────────────────────────────────────────────────

@bp.route('/api/module/RP01/daily-ops/download')
@login_required
def daily_ops_download():
    date_str = request.args.get('report_date', date.today().strftime('%Y-%m-%d'))

    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response('Invalid date', status=400)

    vessels = _fetch_data(report_date)
    if not vessels:
        return Response('No active (non-closed) vessels found', status=404)

    day_rows, month_rows = _fetch_cargo_handled(report_date)
    tide_rows            = _fetch_tide_data(report_date)
    buf = _build_excel(vessels, report_date,
                       day_rows, month_rows, tide_rows)
    fname = f'DailyOps_{date_str}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )
